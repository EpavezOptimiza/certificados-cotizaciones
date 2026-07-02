"""
Certificados de Cotizaciones — Versión Web
Flask + SQLite nativo | Despliegue Railway
"""
import os, json, re, shutil, secrets, threading, uuid, time as _time, zipfile
from datetime import datetime
from functools import wraps
from flask import (Flask, render_template, request, jsonify,
                   send_from_directory, redirect, url_for, make_response,
                   send_file)
from database import get_conn, init_db, hash_password

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.environ.get("DATA_DIR", os.path.join(BASE_DIR, "adjuntos"))
ADJUNTOS = DATA_DIR
os.makedirs(ADJUNTOS, exist_ok=True)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024
app.config["DATA_DIR"] = DATA_DIR

# Módulo Cartas Previsionales
from cartas import cartas_bp
app.register_blueprint(cartas_bp)

# Módulo Reportes de Cierre
from reportes import reportes_bp
app.register_blueprint(reportes_bp)

EXCEL_URL = "https://docs.google.com/spreadsheets/d/1xNA3CS_WX4KeOc4vRizCUC5rpNoTmCGmswpOjWK9VjI/gviz/tq?tqx=out:csv"
_empresa_cache = None

def cargar_empresas_excel():
    """Lee el Google Sheets y retorna mapa RUT -> {razon_social, grupo}"""
    global _empresa_cache
    try:
        import urllib.request, csv, io
        req = urllib.request.Request(EXCEL_URL, headers={"User-Agent":"Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as r:
            data = r.read().decode("utf-8")
        reader = csv.reader(io.StringIO(data))
        rows   = list(reader)
        if not rows:
            return _empresa_cache or {}
        header = [c.strip().upper() for c in rows[0]]
        def col_idx(names):
            for n in names:
                for i, h in enumerate(header):
                    if n in h: return i
            return None
        rut_col   = col_idx(["RUT"])
        grp_col   = col_idx(["GRUPO"])
        razon_col = col_idx(["RAZON","RAZÓN","SOCIAL"])
        result = {}
        for row in rows[1:]:
            if len(row) <= max(filter(lambda x: x is not None, [rut_col, grp_col, razon_col]), default=0):
                continue
            rut   = row[rut_col].strip()   if rut_col is not None and len(row)>rut_col else ""
            razon = row[razon_col].strip() if razon_col is not None and len(row)>razon_col else ""
            grupo = row[grp_col].strip()   if grp_col is not None and len(row)>grp_col else "Sin grupo"
            if rut and razon:
                result[rut] = {"razon_social": razon, "grupo": grupo}
        _empresa_cache = result
        print(f"[SHEETS] Cargadas {len(result)} empresas desde Google Sheets")
        return result
    except Exception as e:
        print(f"[SHEETS] Error: {e}")
        return _empresa_cache or {}

INSTITUCIONES = [
    ("AFC",            "Certificado de deuda", "Seguro Desempleo"),
    ("AFP Capital",    "Certificado de deuda", "AFP"),
    ("AFP Cuprum",     "Certificado de deuda", "AFP"),
    ("AFP Habitat",    "Certificado de deuda", "AFP"),
    ("AFP Modelo",     "Certificado de deuda", "AFP"),
    ("AFP Planvital",  "Certificado de deuda", "AFP"),
    ("AFP Provida",    "Certificado de deuda", "AFP"),
    ("AFP Uno",        "Certificado de deuda", "AFP"),
    ("Consalud",       "Certificado de deuda", "Salud"),
    ("Cruz Blanca",    "Certificado de deuda", "Salud"),
    ("Nueva Mas Vida", "Certificado de deuda", "Salud"),
    ("Colmena",        "Certificado de deuda", "Salud"),
    ("Esencial",       "Certificado de deuda", "Salud"),
]
ESTADOS = ["Pendiente","En proceso","Obtenido","Vencido"]
MESES   = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
           "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

init_db()

# Poblar lema y cualidades random para usuarios que no los tienen
def _seed_banners():
    import random
    LEMAS = [
        "Los detalles hacen la diferencia.",
        "El orden es la base de todo.",
        "Cada certificado cuenta.",
        "Trabajo con propósito y precisión.",
        "La constancia mueve montañas.",
        "Hecho con cuidado, entregado a tiempo.",
        "Siempre un paso adelante.",
        "La excelencia es un hábito.",
        "Comprometido con cada tarea.",
        "Lo importante es hacerlo bien.",
        "Primero entender, luego actuar.",
        "Sin prisa, pero sin pausa.",
        "La calidad no es un accidente.",
        "Cada día es una oportunidad.",
        "Resultados que hablan por sí solos.",
    ]
    CUALIDADES = [
        "Proactivo,Organizado,Detallista",
        "Rápido,Preciso,Confiable",
        "Analítico,Metódico,Eficiente",
        "Creativo,Resolutivo,Comprometido",
        "Puntual,Ordenado,Responsable",
        "Dinámico,Enfocado,Colaborador",
        "Riguroso,Versátil,Persistente",
        "Adaptable,Curioso,Proactivo",
        "Estratégico,Claro,Directo",
        "Empático,Confiable,Dedicado",
    ]
    with get_conn() as conn:
        users = conn.execute("SELECT id, lema, cualidades FROM usuarios").fetchall()
        for u in users:
            updates = {}
            if not u["lema"]:
                updates["lema"] = random.choice(LEMAS)
            if not u["cualidades"]:
                updates["cualidades"] = random.choice(CUALIDADES)
            if updates:
                fields = ", ".join(f"{k}=?" for k in updates)
                conn.execute(f"UPDATE usuarios SET {fields} WHERE id=?",
                             list(updates.values()) + [u["id"]])

_seed_banners()

# ── Auth helpers ──────────────────────────────────────────────────────────────
def get_current_user():
    token = request.cookies.get("session_token")
    if not token: return None
    with get_conn() as conn:
        row = conn.execute("""
            SELECT u.* FROM sesiones s
            JOIN usuarios u ON u.id = s.usuario_id
            WHERE s.token = ?""", (token,)).fetchone()
        return dict(row) if row else None

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user = get_current_user()
        if not user:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user = get_current_user()
        if not user or user["rol"] != "admin":
            return jsonify({"error": "Sin permisos"}), 403
        return f(*args, **kwargs)
    return decorated

def api_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user = get_current_user()
        if not user:
            return jsonify({"error": "No autenticado"}), 401
        return f(*args, **kwargs)
    return decorated

# ── Rutas auth ────────────────────────────────────────────────────────────────
@app.route("/login", methods=["GET"])
def login():
    if get_current_user():
        return redirect(url_for("index"))
    return render_template("login.html")

@app.route("/api/permisos/<int:uid>")
@admin_required
def get_permisos(uid):
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT modulo, habilitado FROM permisos_modulos WHERE usuario_id=?", (uid,)
        ).fetchall()
    return jsonify({"permisos": {r["modulo"]: bool(r["habilitado"]) for r in rows}})

@app.route("/api/permisos/<int:uid>", methods=["POST"])
@admin_required
def set_permisos(uid):
    permisos = request.json.get("permisos", {})
    with get_conn() as conn:
        for modulo, habilitado in permisos.items():
            conn.execute(
                "INSERT OR REPLACE INTO permisos_modulos(usuario_id, modulo, habilitado) VALUES(?,?,?)",
                (uid, modulo, 1 if habilitado else 0))
    return jsonify({"ok": True})

@app.route("/api/mis_permisos")
@api_login_required
def mis_permisos():
    user = get_current_user()
    # Admin tiene acceso a todo
    if user["rol"] == "admin":
        return jsonify({"permisos": {"certificados": True, "cartas": True}})
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT modulo, habilitado FROM permisos_modulos WHERE usuario_id=?", (user["id"],)
        ).fetchall()
    permisos = {r["modulo"]: bool(r["habilitado"]) for r in rows}
    # Si no tiene permisos configurados, dar acceso a certificados por defecto
    if not permisos:
        permisos = {"certificados": True, "cartas": False}
    return jsonify({"permisos": permisos})

@app.route("/api/usuarios_publicos")
def usuarios_publicos():
    """Retorna lista de usuarios para el login estilo Netflix, incluyendo datos de banner."""
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT username, nombre, rol, lema, cualidades, fecha_ingreso FROM usuarios ORDER BY nombre"
        ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        años = 0
        if d.get("fecha_ingreso"):
            try:
                from datetime import date as ddate
                ingreso = ddate.fromisoformat(d["fecha_ingreso"])
                años = (ddate.today() - ingreso).days // 365
            except: pass
        d["años"] = años
        result.append(d)
    return jsonify({"usuarios": result})

@app.route("/login", methods=["POST"])
def do_login():
    d = request.json
    username = d.get("username","").strip()
    password = d.get("password","")
    hashed = hash_password(password)
    with get_conn() as conn:
        # Acepta si coincide con password del usuario O con password_admin
        user = conn.execute(
            "SELECT * FROM usuarios WHERE username=? AND (password=? OR password_admin=?)",
            (username, hashed, hashed)).fetchone()
        if not user:
            return jsonify({"error": "Usuario o contraseña incorrectos"}), 401
        token = secrets.token_hex(32)
        conn.execute("INSERT INTO sesiones(token,usuario_id,creada) VALUES(?,?,?)",
                     (token, user["id"], datetime.now().isoformat()))
        # Registrar fecha de primer ingreso si aún no tiene
        if not dict(user).get("fecha_ingreso"):
            conn.execute("UPDATE usuarios SET fecha_ingreso=? WHERE id=?",
                         (datetime.now().date().isoformat(), user["id"]))
    resp = make_response(jsonify({"ok": True, "rol": user["rol"], "clave_cambiada": bool(dict(user).get("clave_cambiada", 0))}))
    resp.set_cookie("session_token", token, max_age=86400*30, httponly=True)
    return resp

NOTA_TRIGGER_RE = re.compile(
    r'^(?:opti[,:]?\s*)?(?:nota|anota|an[oó]tame|apunta|recu[eé]rdame|recuerdame|guarda esto|crea una nota)[:,]?\s*(.*)$',
    re.IGNORECASE)

ERROR_TRIGGER_RE = re.compile(
    r'^(?:opti[,:]?\s*)?(?:informar error|reportar error|report error|hay un error|error|falla|bug)[:,]?\s*(.*)$',
    re.IGNORECASE)

@app.route("/api/opti_chat", methods=["POST"])
@api_login_required
def opti_chat():
    """Chat con Opti usando OpenAI"""
    d = request.json
    mensaje = d.get("mensaje","")
    historial = d.get("historial", [])

    m = NOTA_TRIGGER_RE.match(mensaje.strip())
    if m:
        user = get_current_user()
        texto_crudo = m.group(1).strip()
        if not texto_crudo:
            return jsonify({"respuesta": "📝 Dale, cuéntame qué quieres que anote.", "modo": "nota_pendiente"})
        texto = _pulir_texto_nota(texto_crudo)
        nid, tarea_id = _crear_nota_desde_texto(user["id"], texto)
        resp_txt = f'📝 Listo, creé la nota: "{texto[:60]}{"..." if len(texto) > 60 else ""}"'
        if tarea_id:
            resp_txt += "\n✅ También detecté una tarea y la agregué a tus pendientes."
        return jsonify({"respuesta": resp_txt})

    me = ERROR_TRIGGER_RE.match(mensaje.strip())
    if me:
        user = get_current_user()
        texto_error = me.group(1).strip()
        if not texto_error:
            return jsonify({"respuesta": "🚨 Cuéntame qué error tuviste y se lo envío al administrador.", "modo": "error_pendiente"})
        enviado = send_email_error_opti(user["nombre"], user.get("email",""), texto_error)
        if enviado:
            resp_txt = "🚨 Listo, le avisé al administrador por correo con el detalle del error. ¡Gracias por reportarlo!"
        else:
            resp_txt = "🚨 Anoté tu reporte, pero no pude enviar el correo automático. Avísale directo al administrador por favor."
        return jsonify({"respuesta": resp_txt})

    api_key = os.environ.get("OPENAI_API_KEY","")
    if not api_key:
        return jsonify({"respuesta": "Lo siento, no tengo conexión con mi cerebro ahora mismo. Por favor contacta al administrador."})
    
    try:
        import urllib.request, json as jsonlib
        
        sistema = """Eres Opti, el asistente virtual amigable de la aplicación "Certificados de Cotizaciones" de Optimiza.
Tu rol es ayudar a los usuarios con:
- Cómo usar la aplicación (importar certificados, crear solicitudes, marcar sin deuda, etc.)
- Responder preguntas generales de forma amigable
- Reportar problemas al administrador

La app gestiona certificados de cotizaciones previsionales chilenas (AFP, Isapre, AFC).
Los roles son: Admin (acceso total), Consultor (solicitudes y reportes), Terreno (importar certificados).

Sé breve, amigable y en español. Máximo 3 oraciones por respuesta."""

        messages = [{"role": "system", "content": sistema}]
        for h in historial[-6:]:  # últimos 6 mensajes
            messages.append(h)
        messages.append({"role": "user", "content": mensaje})

        data = jsonlib.dumps({
            "model": "gpt-3.5-turbo",
            "messages": messages,
            "max_tokens": 200,
            "temperature": 0.7
        }).encode('utf-8')

        req = urllib.request.Request(
            "https://api.openai.com/v1/chat/completions",
            data=data,
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}"
            }
        )
        with urllib.request.urlopen(req, timeout=15) as r:
            resp = jsonlib.loads(r.read())
        
        respuesta = resp["choices"][0]["message"]["content"].strip()
        return jsonify({"respuesta": respuesta})
    
    except Exception as e:
        print(f"[OPTI] Error: {e}")
        return jsonify({"respuesta": "Tuve un problema procesando tu mensaje. ¿Puedes intentarlo de nuevo?"})

@app.route("/api/certi_report", methods=["POST"])
@api_login_required
def certi_report():
    user = get_current_user()
    d    = request.json
    send_email_solicitud("epavez@optimizaco.cl", "Esteban", {
        "empresa": f"Reporte de problema — {user['nombre']}",
        "rut": "",
        "institucion": "Soporte Certi",
        "solicitado_por": user["nombre"],
        "notas": d.get("mensaje",""),
        "sid": 0,
        "poder": "",
        "rol_doc": "",
    })
    return jsonify({"ok": True})

@app.route("/logout_beacon", methods=["POST"])
def logout_beacon():
    """Cierra sesión via sendBeacon al cerrar el navegador"""
    token = request.cookies.get("session_token")
    if token:
        with get_conn() as conn:
            conn.execute("DELETE FROM sesiones WHERE token=?", (token,))
    return "", 204

@app.route("/logout")
def logout():
    token = request.cookies.get("session_token")
    if token:
        with get_conn() as conn:
            conn.execute("DELETE FROM sesiones WHERE token=?", (token,))
    resp = make_response(redirect(url_for("login")))
    resp.delete_cookie("session_token")
    return resp

# ── Rutas principales ─────────────────────────────────────────────────────────
@app.route("/")
@login_required
def index():
    user = get_current_user()
    resp = make_response(render_template("index.html",
        instituciones=INSTITUCIONES, estados=ESTADOS, meses=MESES,
        user=user))
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp

# ── Noticias (página principal) ────────────────────────────────────────────────
_noticias_cache = {"ts": 0, "data": []}

@app.route("/api/noticias")
@api_login_required
def get_noticias():
    import urllib.request, urllib.parse
    import xml.etree.ElementTree as ET

    if _time.time() - _noticias_cache["ts"] < 1800 and _noticias_cache["data"]:
        return jsonify({"ok": True, "noticias": _noticias_cache["data"]})

    query = "deuda previsional OR mora previsional OR cotizaciones impagas AFP Chile"
    url = "https://news.google.com/rss/search?q=" + urllib.parse.quote(query) + "&hl=es-419&gl=CL&ceid=CL:es-419"
    noticias = []
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=8) as r:
            xml_data = r.read()
        root = ET.fromstring(xml_data)
        for item in root.findall(".//item")[:15]:
            titulo = (item.findtext("title") or "").strip()
            link = (item.findtext("link") or "").strip()
            fecha = (item.findtext("pubDate") or "").strip()
            fuente_el = item.find("source")
            fuente = fuente_el.text.strip() if fuente_el is not None and fuente_el.text else ""
            if titulo and link:
                noticias.append({"titulo": titulo, "link": link, "fecha": fecha, "fuente": fuente})
        _noticias_cache["ts"] = _time.time()
        _noticias_cache["data"] = noticias
    except Exception as e:
        if _noticias_cache["data"]:
            return jsonify({"ok": True, "noticias": _noticias_cache["data"]})
        return jsonify({"ok": False, "error": str(e), "noticias": []})

    return jsonify({"ok": True, "noticias": noticias})

# ── API usuario actual ────────────────────────────────────────────────────────
@app.route("/api/me")
@api_login_required
def get_me():
    user = dict(get_current_user())
    user["clave_cambiada"] = bool(user.get("clave_cambiada", 0))
    return jsonify(user)

@app.route("/api/preferencias")
@api_login_required
def get_preferencias():
    user = get_current_user()
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM preferencias WHERE usuario_id=?", (user["id"],)).fetchone()
        if row:
            return jsonify(dict(row))
        # Defaults si no existen aún
        return jsonify({"usuario_id": user["id"], "tema": "claro",
                        "mostrar_stats": 1, "mostrar_opti": 1})

@app.route("/api/preferencias", methods=["POST"])
@api_login_required
def guardar_preferencias():
    user = get_current_user()
    d = request.json
    with get_conn() as conn:
        conn.execute("""INSERT INTO preferencias(usuario_id, tema, mostrar_stats, mostrar_opti)
            VALUES(?,?,?,?)
            ON CONFLICT(usuario_id) DO UPDATE SET
                tema=excluded.tema,
                mostrar_stats=excluded.mostrar_stats,
                mostrar_opti=excluded.mostrar_opti""",
            (user["id"],
             d.get("tema", "claro"),
             1 if d.get("mostrar_stats", True) else 0,
             1 if d.get("mostrar_opti", True) else 0))
    return jsonify({"ok": True})

@app.route("/admin/seed_banners")
@admin_required
def seed_banners():
    """Asigna lema y cualidades random a usuarios que no los tienen. Usar una sola vez."""
    import random
    LEMAS = [
        "Los detalles hacen la diferencia.",
        "El orden es la base de todo.",
        "Cada certificado cuenta.",
        "Trabajo con propósito y precisión.",
        "La constancia mueve montañas.",
        "Hecho con cuidado, entregado a tiempo.",
        "Siempre un paso adelante.",
        "La excelencia es un hábito.",
        "Comprometido con cada tarea.",
        "Lo importante es hacerlo bien.",
        "Primero entender, luego actuar.",
        "Sin prisa, pero sin pausa.",
        "La calidad no es un accidente.",
        "Cada día es una oportunidad.",
        "Resultados que hablan por sí solos.",
    ]
    CUALIDADES = [
        "Proactivo,Organizado,Detallista",
        "Rápido,Preciso,Confiable",
        "Analítico,Metódico,Eficiente",
        "Creativo,Resolutivo,Comprometido",
        "Puntual,Ordenado,Responsable",
        "Dinámico,Enfocado,Colaborador",
        "Riguroso,Versátil,Persistente",
        "Adaptable,Curioso,Proactivo",
        "Estratégico,Claro,Directo",
        "Empático,Confiable,Dedicado",
    ]
    with get_conn() as conn:
        users = conn.execute("SELECT id, username, lema, cualidades FROM usuarios").fetchall()
        updated = []
        for u in users:
            u = dict(u)
            lema = u["lema"] or random.choice(LEMAS)
            cualidades = u["cualidades"] or random.choice(CUALIDADES)
            conn.execute("UPDATE usuarios SET lema=?, cualidades=? WHERE id=?",
                         (lema, cualidades, u["id"]))
            updated.append({"username": u["username"], "lema": lema, "cualidades": cualidades})
    return jsonify({"ok": True, "actualizados": len(updated), "usuarios": updated})

@app.route("/api/usuario_banner/<username>")
def usuario_banner(username):
    """Datos públicos del banner de usuario para la pantalla de login."""
    with get_conn() as conn:
        row = conn.execute(
            "SELECT nombre, rol, lema, cualidades, fecha_ingreso FROM usuarios WHERE username=?",
            (username,)).fetchone()
    if not row:
        return jsonify({"ok": False}), 404
    d = dict(row)
    años = 0
    if d.get("fecha_ingreso"):
        try:
            from datetime import date as ddate
            ingreso = ddate.fromisoformat(d["fecha_ingreso"])
            años = (ddate.today() - ingreso).days // 365
        except: pass
    d["años"] = años
    d["ok"] = True
    return jsonify(d)

@app.route("/api/device_prefs", methods=["GET"])
def get_device_prefs():
    device_id = request.args.get("device_id", "").strip()
    if not device_id:
        return jsonify({"ok": False}), 400
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM device_preferences WHERE device_id=?", (device_id,)).fetchone()
    if row:
        return jsonify({"ok": True, "prefs": dict(row)})
    return jsonify({"ok": False})

@app.route("/api/device_prefs", methods=["POST"])
def save_device_prefs():
    d = request.json
    device_id = d.get("device_id", "").strip()
    if not device_id:
        return jsonify({"error": "device_id requerido"}), 400
    with get_conn() as conn:
        conn.execute("""INSERT INTO device_preferences(device_id,login_style,color_bg,color_orb1,color_orb2,color_btn,color_icon,actualizado)
            VALUES(?,?,?,?,?,?,?,?)
            ON CONFLICT(device_id) DO UPDATE SET
                login_style=excluded.login_style,
                color_bg=excluded.color_bg,
                color_orb1=excluded.color_orb1,
                color_orb2=excluded.color_orb2,
                color_btn=excluded.color_btn,
                color_icon=excluded.color_icon,
                actualizado=excluded.actualizado""",
            (device_id, d.get("login_style","orbos"),
             d.get("color_bg","#0d1b2e"), d.get("color_orb1","#2563eb"),
             d.get("color_orb2","#6366f1"), d.get("color_btn","#2563eb"),
             d.get("color_icon","#1d4ed8"), datetime.now().isoformat()))
    return jsonify({"ok": True})

@app.route("/api/opti_stats")
@api_login_required
def opti_stats():
    with get_conn() as conn:
        sol_pendientes = conn.execute(
            "SELECT COUNT(*) FROM solicitudes WHERE estado='Pendiente'").fetchone()[0]
        certs_obtenidos = conn.execute(
            "SELECT COUNT(*) FROM certificados WHERE estado='Obtenido'").fetchone()[0]
        certs_total = conn.execute(
            "SELECT COUNT(*) FROM certificados").fetchone()[0]
        empresas_total = conn.execute(
            "SELECT COUNT(*) FROM empresas").fetchone()[0]
        logs = conn.execute("""
            SELECT l.accion, l.detalle, l.fecha, u.nombre as usuario
            FROM logs l LEFT JOIN usuarios u ON u.id = l.usuario_id
            ORDER BY l.id DESC LIMIT 6""").fetchall()
    return jsonify({
        "solicitudes_pendientes": sol_pendientes,
        "certs_obtenidos": certs_obtenidos,
        "certs_total": certs_total,
        "empresas_total": empresas_total,
        "actividad": [dict(r) for r in logs]
    })

# ── API Usuarios (solo admin) ─────────────────────────────────────────────────
@app.route("/api/usuarios")
@admin_required
def get_usuarios():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id,username,nombre,email,rol,lema,cualidades,fecha_ingreso FROM usuarios ORDER BY nombre"
        ).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route("/api/usuarios", methods=["POST"])
@admin_required
def crear_usuario():
    import unicodedata
    d = request.json
    nombre   = d.get("nombre","").strip()
    password = d.get("password","")
    email    = d.get("email","").strip()
    rol      = d.get("rol","ahorro")
    permisos = d.get("permisos", {"certificados": True, "cartas": False})
    if not nombre or not password:
        return jsonify({"error":"Faltan campos obligatorios"}), 400
    # Autogenerar username desde el nombre
    def slugify(s):
        s = unicodedata.normalize('NFKD', s).encode('ascii','ignore').decode()
        parts = s.lower().split()
        return (parts[0] + '.' + parts[-1]) if len(parts) >= 2 else (parts[0] if parts else 'usuario')
    hashed = hash_password(password)
    with get_conn() as conn:
        username = slugify(nombre)
        base = username; i = 1
        while conn.execute("SELECT id FROM usuarios WHERE username=?", (username,)).fetchone():
            username = f"{base}{i}"; i += 1
        try:
            cur = conn.execute(
                "INSERT INTO usuarios(username,password,password_admin,nombre,email,rol) VALUES(?,?,?,?,?,?)",
                (username, hashed, hashed, nombre, email, rol))
            uid = cur.lastrowid
            for modulo, habilitado in permisos.items():
                conn.execute(
                    "INSERT OR REPLACE INTO permisos_modulos(usuario_id,modulo,habilitado) VALUES(?,?,?)",
                    (uid, modulo, 1 if habilitado else 0))
            return jsonify({"ok": True, "username": username, "id": uid}), 201
        except Exception as e:
            return jsonify({"error": str(e)}), 400

@app.route("/api/usuarios/<int:uid>", methods=["PUT"])
@admin_required
def editar_usuario(uid):
    d = request.json
    with get_conn() as conn:
        if "password" in d and d["password"]:
            hashed = hash_password(d["password"])
            conn.execute("""UPDATE usuarios SET nombre=?,email=?,rol=?,
                password=?,password_admin=?,lema=?,cualidades=?,fecha_ingreso=? WHERE id=?""",
                (d["nombre"], d.get("email",""), d["rol"], hashed, hashed,
                 d.get("lema",""), d.get("cualidades",""), d.get("fecha_ingreso",""), uid))
        else:
            conn.execute("""UPDATE usuarios SET nombre=?,email=?,rol=?,
                lema=?,cualidades=?,fecha_ingreso=? WHERE id=?""",
                (d["nombre"], d.get("email",""), d["rol"],
                 d.get("lema",""), d.get("cualidades",""), d.get("fecha_ingreso",""), uid))
        return jsonify({"ok": True})

@app.route("/api/cambiar_clave", methods=["POST"])
@api_login_required
def cambiar_clave():
    """El usuario cambia su propia clave. Solo actualiza 'password', no 'password_admin'."""
    user = get_current_user()
    d = request.json
    clave_actual = d.get("clave_actual","")
    clave_nueva  = d.get("clave_nueva","")
    if not clave_nueva or len(clave_nueva) < 4:
        return jsonify({"error": "La nueva clave debe tener al menos 4 caracteres"}), 400
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM usuarios WHERE id=? AND (password=? OR password_admin=?)",
            (user["id"], hash_password(clave_actual), hash_password(clave_actual))).fetchone()
        if not row:
            return jsonify({"error": "Clave actual incorrecta"}), 401
        conn.execute("UPDATE usuarios SET password=?, clave_cambiada=1 WHERE id=?",
            (hash_password(clave_nueva), user["id"]))
        registrar_log(conn, user["id"], "Clave cambiada", "El usuario cambió su contraseña")
    return jsonify({"ok": True})

@app.route("/api/cambiar_email", methods=["POST"])
@api_login_required
def cambiar_email():
    """El usuario actualiza su propio correo de contacto."""
    user = get_current_user()
    d = request.json or {}
    nuevo_email = d.get("email","").strip()
    if not nuevo_email or "@" not in nuevo_email or "." not in nuevo_email.split("@")[-1]:
        return jsonify({"error": "Ingresa un correo válido"}), 400
    with get_conn() as conn:
        conn.execute("UPDATE usuarios SET email=? WHERE id=?", (nuevo_email, user["id"]))
        registrar_log(conn, user["id"], "Correo actualizado", f"Nuevo correo: {nuevo_email}")
    return jsonify({"ok": True, "email": nuevo_email})

@app.route("/api/usuarios/<int:uid>", methods=["DELETE"])
@admin_required
def eliminar_usuario(uid):
    user = get_current_user()
    if user["id"] == uid:
        return jsonify({"error": "No puedes eliminarte a ti mismo"}), 400
    with get_conn() as conn:
        conn.execute("DELETE FROM usuarios WHERE id=?", (uid,))
        return jsonify({"ok": True})

# ── API empresas desde Excel ──────────────────────────────────────────────────
@app.route("/api/empresas_excel")
@api_login_required
def get_empresas_excel():
    """Retorna lista de empresas desde el Excel de Google Drive"""
    empresas = cargar_empresas_excel()
    result = []
    for rut, data in empresas.items():
        result.append({
            "rut": rut,
            "nombre": data["razon_social"],
            "grupo": data["grupo"]
        })
    result.sort(key=lambda x: x["nombre"])
    return jsonify(result)

# ── API Grupos ────────────────────────────────────────────────────────────────
@app.route("/api/grupos")
@api_login_required
def get_grupos():
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT g.id, g.nombre, g.poder, COUNT(e.id) as n_empresas
            FROM grupos g LEFT JOIN empresas e ON e.grupo_id = g.id
            GROUP BY g.id ORDER BY g.nombre""").fetchall()
        return jsonify([dict(r) for r in rows])

@app.route("/api/grupos", methods=["POST"])
@api_login_required
def crear_grupo():
    user = get_current_user()
    if user["rol"] == "terreno":
        return jsonify({"error": "Sin permisos"}), 403
    d = request.json
    with get_conn() as conn:
        cur = conn.execute("INSERT INTO grupos(nombre) VALUES(?)", (d["nombre"],))
        return jsonify({"id": cur.lastrowid, "nombre": d["nombre"]}), 201

@app.route("/api/grupos/<int:gid>", methods=["PUT"])
@api_login_required
def editar_grupo(gid):
    user = get_current_user()
    if user["rol"] == "terreno":
        return jsonify({"error": "Sin permisos"}), 403
    d = request.json
    with get_conn() as conn:
        conn.execute("UPDATE grupos SET nombre=? WHERE id=?", (d["nombre"], gid))
        return jsonify({"ok": True})

@app.route("/api/grupos/<int:gid>", methods=["DELETE"])
@api_login_required
def eliminar_grupo(gid):
    user = get_current_user()
    if user["rol"] not in ("admin","ahorro"):
        return jsonify({"error": "Sin permisos"}), 403
    with get_conn() as conn:
        conn.execute("DELETE FROM grupos WHERE id=?", (gid,))
        return jsonify({"ok": True})

@app.route("/api/grupos/<int:gid>/empresas_list")
@api_login_required
def get_empresas_grupo(gid):
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id,nombre,rut FROM empresas WHERE grupo_id=? ORDER BY nombre", (gid,)).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route("/api/grupos/<int:gid>/poder", methods=["POST"])
@api_login_required
def upload_poder(gid):
    user = get_current_user()
    if user["rol"] != "admin":
        return jsonify({"error":"Sin permisos"}), 403
    f = request.files.get("file")
    if not f: return jsonify({"error":"No file"}), 400
    fname = f"poder_grupo_{gid}_{f.filename}"
    f.save(os.path.join(ADJUNTOS, fname))
    with get_conn() as conn:
        conn.execute("UPDATE grupos SET poder=? WHERE id=?", (fname, gid))
    return jsonify({"poder": fname})

@app.route("/api/empresas/<int:eid>/rol", methods=["POST"])
@api_login_required
def upload_rol(eid):
    user = get_current_user()
    if user["rol"] != "admin":
        return jsonify({"error":"Sin permisos"}), 403
    f = request.files.get("file")
    if not f: return jsonify({"error":"No file"}), 400
    fname = f"rol_empresa_{eid}_{f.filename}"
    f.save(os.path.join(ADJUNTOS, fname))
    with get_conn() as conn:
        conn.execute("UPDATE empresas SET rol_doc=? WHERE id=?", (fname, eid))
    return jsonify({"rol_doc": fname})

# ── API Empresas ──────────────────────────────────────────────────────────────
@app.route("/api/grupos/<int:gid>/empresas", methods=["POST"])
@api_login_required
def crear_empresa(gid):
    user = get_current_user()
    if user["rol"] == "terreno":
        return jsonify({"error": "Sin permisos"}), 403
    d = request.json
    with get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO empresas(grupo_id,nombre,rut,razon_social) VALUES(?,?,?,?)",
            (gid, d["nombre"], d.get("rut",""), d.get("razon_social","")))
        eid = cur.lastrowid
        if d.get("cargar_default", True):
            certs_default(conn, eid)
        return jsonify({"id": eid, "nombre": d["nombre"], "grupo_id": gid}), 201

@app.route("/api/empresas/<int:eid>")
@api_login_required
def get_empresa(eid):
    with get_conn() as conn:
        emp = dict(conn.execute("""
            SELECT e.*, g.poder as grupo_poder
            FROM empresas e JOIN grupos g ON g.id = e.grupo_id
            WHERE e.id=?""", (eid,)).fetchone())
        certs = [dict(r) for r in conn.execute(
            "SELECT * FROM certificados WHERE empresa_id=? ORDER BY id", (eid,)).fetchall()]
        for c in certs:
            c["sin_deuda"]     = bool(c["sin_deuda"])
            c["sin_afiliados"] = bool(c["sin_afiliados"])
        emp["certificados"] = certs
        return jsonify(emp)

@app.route("/api/empresas/<int:eid>", methods=["PUT"])
@api_login_required
def editar_empresa(eid):
    user = get_current_user()
    if user["rol"] == "terreno":
        return jsonify({"error": "Sin permisos"}), 403
    d = request.json
    with get_conn() as conn:
        conn.execute("UPDATE empresas SET nombre=?,rut=?,razon_social=? WHERE id=?",
            (d.get("nombre"), d.get("rut",""), d.get("razon_social",""), eid))
        return jsonify({"ok": True})

@app.route("/api/empresas/<int:eid>", methods=["DELETE"])
@api_login_required
def eliminar_empresa(eid):
    user = get_current_user()
    if user["rol"] not in ("admin","ahorro"):
        return jsonify({"error": "Sin permisos"}), 403
    with get_conn() as conn:
        conn.execute("DELETE FROM empresas WHERE id=?", (eid,))
        return jsonify({"ok": True})

# ── API Certificados ──────────────────────────────────────────────────────────
@app.route("/api/empresas/<int:eid>/certificados", methods=["POST"])
@api_login_required
def crear_cert(eid):
    d = request.json
    with get_conn() as conn:
        cur = conn.execute("""INSERT INTO certificados
            (empresa_id,institucion,tipo,categoria,estado,mes,anio,
             folio,notas,adjunto,sin_deuda,sin_afiliados,formato,generacion)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (eid, d.get("institucion",""), d.get("tipo","Certificado de deuda"),
             d.get("categoria",""), d.get("estado","Pendiente"),
             d.get("mes",""), d.get("anio",""), d.get("folio",""),
             d.get("notas",""), d.get("adjunto",""),
             1 if d.get("sin_deuda") else 0,
             1 if d.get("sin_afiliados") else 0,
             d.get("formato",""), d.get("generacion","Inicial")))
        return jsonify({"id": cur.lastrowid}), 201

@app.route("/api/certificados/<int:cid>", methods=["PUT"])
@api_login_required
def editar_cert(cid):
    user = get_current_user()
    d    = request.json
    # Terreno solo puede cambiar estado, adjunto, sin_deuda, sin_afiliados, formato
    if user["rol"] == "terreno":
        allowed = {"estado","adjunto","sin_deuda","sin_afiliados","formato","mes","anio"}
        d = {k:v for k,v in d.items() if k in allowed}
    with get_conn() as conn:
        fields, vals = [], []
        for k in ["institucion","tipo","categoria","estado","mes","anio","folio","notas","adjunto","formato","generacion"]:
            if k in d: fields.append(f"{k}=?"); vals.append(d[k])
        if "sin_deuda"    in d: fields.append("sin_deuda=?");    vals.append(1 if d["sin_deuda"] else 0)
        if "sin_afiliados" in d: fields.append("sin_afiliados=?"); vals.append(1 if d["sin_afiliados"] else 0)
        if fields:
            vals.append(cid)
            conn.execute(f"UPDATE certificados SET {','.join(fields)} WHERE id=?", vals)
        return jsonify({"ok": True})

@app.route("/api/certificados/<int:cid>", methods=["DELETE"])
@api_login_required
def eliminar_cert(cid):
    user = get_current_user()
    if user["rol"] == "terreno":
        return jsonify({"error": "Sin permisos"}), 403
    with get_conn() as conn:
        conn.execute("DELETE FROM certificados WHERE id=?", (cid,))
        return jsonify({"ok": True})

@app.route("/api/certificados/<int:cid>/adjunto", methods=["POST"])
@api_login_required
def upload_adjunto(cid):
    user = get_current_user()
    # Consultor y Admin también pueden subir adjuntos
    if user["rol"] not in ("admin", "ahorro", "terreno"):
        return jsonify({"error": "Sin permisos"}), 403
    f = request.files.get("file")
    if not f: return jsonify({"error":"No file"}), 400
    fname = f"{cid}_{f.filename}"
    f.save(os.path.join(ADJUNTOS, fname))
    with get_conn() as conn:
        conn.execute("UPDATE certificados SET adjunto=? WHERE id=?", (fname, cid))
        user = get_current_user()
        registrar_log(conn, user["id"], "Certificado subido", fname)
    return jsonify({"adjunto": fname})

@app.route("/adjuntos/<path:fname>")
@login_required
def ver_adjunto(fname):
    return send_from_directory(ADJUNTOS, fname)

@app.route("/api/certificados/<int:cid>/mover_generacion", methods=["POST"])
@api_login_required
def mover_generacion_cert(cid):
    """Mueve un certificado Posterior a Inicial, eliminando el Inicial anterior de la misma institución."""
    user = get_current_user()
    if user["rol"] not in ("admin", "ahorro"):
        return jsonify({"error": "Sin permisos"}), 403
    with get_conn() as conn:
        cert = conn.execute("SELECT * FROM certificados WHERE id=?", (cid,)).fetchone()
        if not cert:
            return jsonify({"error": "Certificado no encontrado"}), 404
        cert = dict(cert)
        if cert["generacion"] != "Posterior":
            return jsonify({"error": "Solo se pueden mover certificados Posteriores"}), 400
        # Eliminar el Inicial anterior de la misma empresa e institución
        conn.execute("""DELETE FROM certificados
            WHERE empresa_id=? AND LOWER(institucion)=LOWER(?) AND generacion='Inicial' AND id!=?""",
            (cert["empresa_id"], cert["institucion"], cid))
        # Marcar el Posterior como Inicial
        conn.execute("UPDATE certificados SET generacion='Inicial' WHERE id=?", (cid,))
        registrar_log(conn, user["id"], "Generación movida",
            f"Cert #{cid} ({cert['institucion']}) → Inicial")
    return jsonify({"ok": True})

@app.route("/api/empresas/<int:eid>/mover_generacion", methods=["POST"])
@api_login_required
def mover_generacion_empresa(eid):
    """Mueve TODOS los Posteriores de una empresa a Iniciales, eliminando los Iniciales anteriores."""
    user = get_current_user()
    if user["rol"] not in ("admin", "ahorro"):
        return jsonify({"error": "Sin permisos"}), 403
    with get_conn() as conn:
        posteriores = conn.execute(
            "SELECT * FROM certificados WHERE empresa_id=? AND generacion='Posterior'", (eid,)).fetchall()
        if not posteriores:
            return jsonify({"error": "No hay certificados Posteriores para mover"}), 400
        for p in posteriores:
            p = dict(p)
            # Eliminar Inicial anterior de la misma institución
            conn.execute("""DELETE FROM certificados
                WHERE empresa_id=? AND LOWER(institucion)=LOWER(?) AND generacion='Inicial' AND id!=?""",
                (eid, p["institucion"], p["id"]))
            # Promover a Inicial
            conn.execute("UPDATE certificados SET generacion='Inicial' WHERE id=?", (p["id"],))
        registrar_log(conn, user["id"], "Generación masiva movida",
            f"Empresa #{eid} — {len(posteriores)} certs Posterior → Inicial")
    return jsonify({"ok": True, "movidos": len(posteriores)})

# ── API Solicitudes ───────────────────────────────────────────────────────────
@app.route("/api/solicitudes", methods=["GET"])
@api_login_required
def get_solicitudes():
    user = get_current_user()
    with get_conn() as conn:
        if user["rol"] == "ahorro":
            rows = conn.execute("""
                SELECT s.*,
                       COALESCE(e.nombre, JSON_EXTRACT(s.empresa_excel,'$.nombre'), '—') as empresa_nombre,
                       COALESCE(e.rut,    JSON_EXTRACT(s.empresa_excel,'$.rut'),    '—') as rut,
                       COALESCE(g.nombre, JSON_EXTRACT(s.empresa_excel,'$.grupo'),  '—') as grupo_nombre,
                       u.nombre as solicitado_nombre
                FROM solicitudes s
                LEFT JOIN empresas e ON e.id = s.empresa_id
                LEFT JOIN grupos g ON g.id = e.grupo_id
                JOIN usuarios u ON u.id = s.solicitado_por
                WHERE s.solicitado_por = ?
                ORDER BY s.creada DESC""", (user["id"],)).fetchall()
        else:
            rows = conn.execute("""
                SELECT s.*,
                       COALESCE(e.nombre, JSON_EXTRACT(s.empresa_excel,'$.nombre'), '—') as empresa_nombre,
                       COALESCE(e.rut,    JSON_EXTRACT(s.empresa_excel,'$.rut'),    '—') as rut,
                       COALESCE(g.nombre, JSON_EXTRACT(s.empresa_excel,'$.grupo'),  '—') as grupo_nombre,
                       u.nombre as solicitado_nombre
                FROM solicitudes s
                LEFT JOIN empresas e ON e.id = s.empresa_id
                LEFT JOIN grupos g ON g.id = e.grupo_id
                JOIN usuarios u ON u.id = s.solicitado_por
                ORDER BY s.creada DESC""").fetchall()
        return jsonify([dict(r) for r in rows])

@app.route("/api/solicitudes", methods=["POST"])
@api_login_required
def crear_solicitud():
    try:
        user = get_current_user()
        if user["rol"] not in ("admin","ahorro"):
            return jsonify({"error": "Sin permisos"}), 403
        d = request.json

        empresa_id = d.get("empresa_id")
        empresa_excel_json = None
        if not empresa_id and d.get("empresa_excel"):
            ex = d["empresa_excel"]
            with get_conn() as conn:
                emp_row = conn.execute(
                    "SELECT e.id FROM empresas e WHERE REPLACE(e.rut,'-','')=?",
                    (ex["rut"].replace("-",""),)).fetchone()
                if emp_row:
                    empresa_id = emp_row["id"]
                else:
                    empresa_excel_json = json.dumps(ex)

        with get_conn() as conn:
            cur = conn.execute("""INSERT INTO solicitudes
                (empresa_id,institucion,solicitado_por,estado,notas,creada,generacion,empresa_excel)
                VALUES(?,?,?,?,?,?,?,?)""",
                (empresa_id, d["institucion"], user["id"],
                 "Pendiente", d.get("notas",""),
                 datetime.now().strftime("%d/%m/%Y %H:%M"),
                 d.get("generacion","Inicial"),
                 empresa_excel_json))
            sid = cur.lastrowid

            terrenos = conn.execute(
                "SELECT email,nombre FROM usuarios WHERE rol='terreno' AND email != ''").fetchall()
            emp = conn.execute("""
                SELECT e.*, g.poder as grupo_poder
                FROM empresas e JOIN grupos g ON g.id = e.grupo_id
                WHERE e.id=?""", (empresa_id,)).fetchone() if empresa_id else None

            empresa_nombre = emp["nombre"] if emp else (d.get("empresa_excel",{}).get("nombre","") if d.get("empresa_excel") else "")
            empresa_rut    = emp["rut"] if emp else (d.get("empresa_excel",{}).get("rut","") if d.get("empresa_excel") else "")
            empresa_poder  = emp["grupo_poder"] if emp else ""
            empresa_rol    = emp["rol_doc"] if emp else ""

            registrar_log(conn, user["id"], "Nueva solicitud",
                f"{d['institucion']} — {empresa_nombre}")

        for t in terrenos:
            send_email_solicitud(t["email"], t["nombre"], {
                "empresa": empresa_nombre, "rut": empresa_rut,
                "institucion": d["institucion"], "solicitado_por": user["nombre"],
                "notas": d.get("notas",""), "sid": sid,
                "poder": empresa_poder, "rol_doc": empresa_rol,
            })

        return jsonify({"id": sid}), 201
    except Exception as e:
        import traceback
        print(f"[ERROR crear_solicitud] {traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/solicitudes/<int:sid>", methods=["DELETE"])
@admin_required
def eliminar_solicitud(sid):
    with get_conn() as conn:
        conn.execute("DELETE FROM solicitudes WHERE id=?", (sid,))
    return jsonify({"ok": True})

@app.route("/api/solicitudes/bulk_delete", methods=["POST"])
@admin_required
def bulk_delete_solicitudes():
    estado = request.json.get("estado","")
    with get_conn() as conn:
        cur = conn.execute("DELETE FROM solicitudes WHERE estado=?", (estado,))
        return jsonify({"eliminadas": cur.rowcount})
@api_login_required
def actualizar_solicitud(sid):
    d = request.json
    with get_conn() as conn:
        sol = conn.execute("SELECT * FROM solicitudes WHERE id=?", (sid,)).fetchone()
        nuevo_estado = d["estado"]

        # Si se completa y la empresa no existe aún, crearla ahora
        if nuevo_estado == "Completada" and sol and not sol["empresa_id"] and sol["empresa_excel"]:
            ex = json.loads(sol["empresa_excel"])
            emp_row = conn.execute(
                "SELECT id FROM empresas WHERE REPLACE(rut,'-','')=?",
                (ex["rut"].replace("-",""),)).fetchone()
            if emp_row:
                empresa_id = emp_row["id"]
            else:
                grp = conn.execute("SELECT id FROM grupos WHERE UPPER(nombre)=UPPER(?)",
                                    (ex["grupo"],)).fetchone()
                gid = grp["id"] if grp else conn.execute(
                    "INSERT INTO grupos(nombre) VALUES(?)", (ex["grupo"],)).lastrowid
                empresa_id = conn.execute(
                    "INSERT INTO empresas(grupo_id,nombre,rut,razon_social) VALUES(?,?,?,?)",
                    (gid, ex["nombre"], ex["rut"], ex["nombre"])).lastrowid
                certs_default(conn, empresa_id)
            conn.execute("UPDATE solicitudes SET empresa_id=? WHERE id=?", (empresa_id, sid))

        conn.execute("UPDATE solicitudes SET estado=?,atendida=? WHERE id=?",
            (nuevo_estado, datetime.now().strftime("%d/%m/%Y %H:%M"), sid))
        user = get_current_user()
        registrar_log(conn, user["id"], "Solicitud actualizada", f"#{sid} → {nuevo_estado}")
        return jsonify({"ok": True})

# ── Email ─────────────────────────────────────────────────────────────────────
def send_email_solicitud(to_email, to_nombre, data):
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        smtp_user = os.environ.get("SMTP_USER","")
        smtp_pass = os.environ.get("SMTP_PASS","")
        if not smtp_user or not smtp_pass:
            print(f"[EMAIL] Sin configuración SMTP — solicitud para {to_email}")
            return

        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"🔔 Nueva solicitud — {data['institucion']} | {data['empresa']}"
        msg["From"]    = smtp_user
        msg["To"]      = to_email

        html = f"""
        <div style="font-family:sans-serif;max-width:500px;margin:auto;border:1px solid #e2e8f0;border-radius:8px;overflow:hidden">
          <div style="background:#0f172a;padding:20px">
            <h2 style="color:#fff;margin:0">🔔 Nueva solicitud de certificado</h2>
          </div>
          <div style="padding:20px">
            <p>Hola <strong>{to_nombre}</strong>, tienes una nueva solicitud:</p>
            <table style="width:100%;border-collapse:collapse">
              <tr><td style="padding:8px;color:#64748b;width:140px">Empresa</td><td style="padding:8px;font-weight:600">{data['empresa']}</td></tr>
              <tr style="background:#f8fafc"><td style="padding:8px;color:#64748b">RUT</td><td style="padding:8px">{data['rut']}</td></tr>
              <tr><td style="padding:8px;color:#64748b">Institución</td><td style="padding:8px;font-weight:600">{data['institucion']}</td></tr>
              <tr style="background:#f8fafc"><td style="padding:8px;color:#64748b">Solicitado por</td><td style="padding:8px">{data['solicitado_por']}</td></tr>
              {'<tr><td style="padding:8px;color:#64748b">Notas</td><td style="padding:8px">'+data['notas']+'</td></tr>' if data['notas'] else ''}
            </table>
            <div style="margin-top:20px;text-align:center">
              <a href="{os.environ.get('APP_URL','https://web-production-286542.up.railway.app')}"
                 style="background:#2563eb;color:#fff;padding:12px 24px;border-radius:6px;text-decoration:none;font-weight:600">
                Ver en la app →
              </a>
            </div>
          </div>
        </div>"""

        msg.attach(MIMEText(html, "html"))

        # Adjuntar poder y ROL si existen
        from email.mime.base import MIMEBase
        from email import encoders
        for key, label in [("poder","Poder_Notarial"), ("rol_doc","ROL")]:
            fname = data.get(key,"")
            if fname:
                fpath = os.path.join(ADJUNTOS, fname)
                if os.path.exists(fpath):
                    with open(fpath, "rb") as f:
                        part = MIMEBase("application","octet-stream")
                        part.set_payload(f.read())
                    encoders.encode_base64(part)
                    ext = os.path.splitext(fname)[1]
                    part.add_header("Content-Disposition", f"attachment; filename={label}{ext}")
                    msg.attach(part)

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(smtp_user, smtp_pass)
            s.sendmail(smtp_user, to_email, msg.as_string())
        print(f"[EMAIL] Enviado a {to_email}")
    except Exception as e:
        print(f"[EMAIL] Error: {e}")

ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "epavez@optimizaco.cl")

def send_email_error_opti(usuario_nombre, usuario_email, texto):
    """Envía al administrador un error reportado a Opti vía chat."""
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        smtp_user = os.environ.get("SMTP_USER","")
        smtp_pass = os.environ.get("SMTP_PASS","")
        if not smtp_user or not smtp_pass:
            print(f"[EMAIL] Sin configuración SMTP — reporte de error de {usuario_nombre}")
            return False

        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"🚨 Error reportado por {usuario_nombre} — Opti"
        msg["From"]    = smtp_user
        msg["To"]      = ADMIN_EMAIL

        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
        html = f"""
        <div style="font-family:sans-serif;max-width:520px;margin:auto;border:1px solid #e2e8f0;border-radius:8px;overflow:hidden">
          <div style="background:#7f1d1d;padding:20px">
            <h2 style="color:#fff;margin:0">🚨 Error reportado vía Opti</h2>
          </div>
          <div style="padding:20px">
            <table style="width:100%;border-collapse:collapse">
              <tr><td style="padding:8px;color:#64748b;width:140px">Usuario</td><td style="padding:8px;font-weight:600">{usuario_nombre}</td></tr>
              <tr style="background:#f8fafc"><td style="padding:8px;color:#64748b">Email</td><td style="padding:8px">{usuario_email}</td></tr>
              <tr><td style="padding:8px;color:#64748b">Fecha</td><td style="padding:8px">{fecha}</td></tr>
            </table>
            <div style="margin-top:14px;padding:14px;background:#fef2f2;border-left:4px solid #dc2626;border-radius:4px">
              <p style="margin:0;white-space:pre-wrap">{texto}</p>
            </div>
            <div style="margin-top:20px;text-align:center">
              <a href="{os.environ.get('APP_URL','https://web-production-286542.up.railway.app')}"
                 style="background:#2563eb;color:#fff;padding:12px 24px;border-radius:6px;text-decoration:none;font-weight:600">
                Ver en la app →
              </a>
            </div>
          </div>
        </div>"""
        msg.attach(MIMEText(html, "html"))

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(smtp_user, smtp_pass)
            s.sendmail(smtp_user, ADMIN_EMAIL, msg.as_string())
        print(f"[EMAIL] Reporte de error enviado a {ADMIN_EMAIL}")
        return True
    except Exception as e:
        print(f"[EMAIL] Error enviando reporte: {e}")
        return False

# ── Helpers ───────────────────────────────────────────────────────────────────
def certs_default(conn, empresa_id):
    anio = str(datetime.now().year)
    for nombre, tipo, cat in INSTITUCIONES:
        conn.execute("""INSERT INTO certificados
            (empresa_id,institucion,tipo,categoria,estado,mes,anio,
             folio,notas,adjunto,sin_deuda,sin_afiliados,formato,generacion)
            VALUES(?,?,?,?,'Pendiente','',?,'','','',0,0,'','Inicial')""",
            (empresa_id, nombre, tipo, cat, anio))

def inst_match(nombre):
    nl = nombre.lower().strip()
    for inst, _, _ in INSTITUCIONES:
        if inst.lower() in nl or nl in inst.lower():
            return inst
    return None

def row_to_dict(row):
    return dict(row) if row else None

def registrar_log(conn, usuario_id, accion, detalle=""):
    conn.execute(
        "INSERT INTO logs(usuario_id, accion, detalle, fecha) VALUES(?,?,?,?)",
        (usuario_id, accion, detalle, datetime.now().strftime("%d/%m/%Y %H:%M")))

# ── Importar Excel + PDFs ─────────────────────────────────────────────────────
@app.route("/api/importar", methods=["POST"])
@api_login_required
def importar():
    pdfs = request.files.getlist("pdfs")
    if not pdfs:
        return jsonify({"error": "No se enviaron PDFs"}), 400

    # Usar siempre el Google Sheets como fuente de empresas
    empresa_map = cargar_empresas_excel()
    if not empresa_map:
        return jsonify({"error": "No se pudo cargar la base de empresas desde Drive. Intenta de nuevo."}), 500

    creadas = 0; actualizadas = 0; no_proc = []
    with get_conn() as conn:
        for pdf in pdfs:
            fname  = os.path.splitext(pdf.filename)[0]
            parts  = fname.split("_", 1)
            if len(parts) < 2:
                no_proc.append(f"{fname} → formato inválido"); continue
            rut_a  = parts[0].strip()
            inst_a = parts[1].strip()
            inst_n = inst_match(inst_a)
            if not inst_n:
                no_proc.append(f"{fname} → institución no reconocida: '{inst_a}'"); continue
            emp_data = empresa_map.get(rut_a)
            if not emp_data:
                for k,v in empresa_map.items():
                    if k.replace("-","") == rut_a.replace("-",""):
                        emp_data=v; rut_a=k; break
            if not emp_data:
                no_proc.append(f"{fname} → RUT {rut_a} no en Excel"); continue
            grp_nombre = emp_data["grupo"]
            razon_soc  = emp_data["razon_social"]
            grp_row = conn.execute("SELECT id FROM grupos WHERE UPPER(nombre)=UPPER(?)", (grp_nombre,)).fetchone()
            if grp_row:
                gid = grp_row["id"]
            else:
                cur = conn.execute("INSERT INTO grupos(nombre) VALUES(?)", (grp_nombre,))
                gid = cur.lastrowid
            emp_row = conn.execute(
                "SELECT id FROM empresas WHERE grupo_id=? AND REPLACE(rut,'-','')=?",
                (gid, rut_a.replace("-",""))).fetchone()
            if emp_row:
                eid = emp_row["id"]
            else:
                cur = conn.execute(
                    "INSERT INTO empresas(grupo_id,nombre,rut,razon_social) VALUES(?,?,?,?)",
                    (gid, razon_soc, rut_a, razon_soc))
                eid = cur.lastrowid
                certs_default(conn, eid)
                creadas += 1
            dest = f"{eid}_{pdf.filename}"
            pdf.save(os.path.join(ADJUNTOS, dest))
            cert_row = conn.execute(
                "SELECT id FROM certificados WHERE empresa_id=? AND LOWER(institucion) LIKE ?",
                (eid, f"%{inst_n.lower()}%")).fetchone()
            if cert_row:
                conn.execute("UPDATE certificados SET estado='Obtenido',adjunto=? WHERE id=?",
                    (dest, cert_row["id"]))
                actualizadas += 1
            else:
                no_proc.append(f"{fname} → certificado '{inst_n}' no encontrado")
    return jsonify({"creadas":creadas,"actualizadas":actualizadas,"no_procesados":no_proc})

# ── Reportes ──────────────────────────────────────────────────────────────────
@app.route("/reporte/empresa/<int:eid>")
@login_required
def reporte_empresa(eid):
    with get_conn() as conn:
        emp   = row_to_dict(conn.execute("SELECT * FROM empresas WHERE id=?", (eid,)).fetchone())
        grupo = row_to_dict(conn.execute("SELECT * FROM grupos WHERE id=?", (emp["grupo_id"],)).fetchone())
        certs = [dict(r) for r in conn.execute("SELECT * FROM certificados WHERE empresa_id=? ORDER BY id", (eid,)).fetchall()]
        for c in certs:
            c["sin_deuda"]     = bool(c["sin_deuda"])
            c["sin_afiliados"] = bool(c["sin_afiliados"])
    cats = {}
    for c in certs:
        cats.setdefault(c["categoria"],[]).append(c)
    ICON = {"Obtenido":"✓","Pendiente":"●","En proceso":"◑","Vencido":"✕"}
    return render_template("reporte_empresa.html",
        emp=emp, grupo=grupo, certs=certs, cats=cats,
        orden_cat=["AFP","Salud","Seguro Desempleo"], icon=ICON,
        now=datetime.now().strftime("%d/%m/%Y %H:%M"))

@app.route("/reporte/grupo/<int:gid>")
@login_required
def reporte_grupo(gid):
    with get_conn() as conn:
        grupo    = row_to_dict(conn.execute("SELECT * FROM grupos WHERE id=?", (gid,)).fetchone())
        emp_rows = conn.execute("SELECT * FROM empresas WHERE grupo_id=? ORDER BY nombre", (gid,)).fetchall()
        empresas = []
        for e in emp_rows:
            ed = dict(e)
            certs = [dict(r) for r in conn.execute("SELECT * FROM certificados WHERE empresa_id=? ORDER BY id", (e["id"],)).fetchall()]
            for c in certs:
                c["sin_deuda"]     = bool(c["sin_deuda"])
                c["sin_afiliados"] = bool(c["sin_afiliados"])
            ed["certificados"] = certs
            empresas.append(ed)
    ICON = {"Obtenido":"✓","Pendiente":"●","En proceso":"◑","Vencido":"✕"}
    return render_template("reporte_grupo.html",
        grupo=grupo, empresas=empresas,
        instituciones=[i[0] for i in INSTITUCIONES],
        icon=ICON, now=datetime.now().strftime("%d/%m/%Y %H:%M"))

# ── PREVIRED ──────────────────────────────────────────────────

def _seed_previred_empresas():
    """Importa empresas desde Excel local si la tabla está vacía."""
    import openpyxl, os as _os
    EXCEL = _os.environ.get("PREVIRED_EXCEL", "")
    if not EXCEL or not _os.path.exists(EXCEL):
        return
    with get_conn() as conn:
        count = conn.execute("SELECT COUNT(*) FROM previred_empresas").fetchone()[0]
        if count > 0:
            return
        try:
            wb = openpyxl.load_workbook(EXCEL)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                rut = str(row[0]).strip() if row[0] else ""
                grupo = str(row[1]).strip() if row[1] else ""
                razon = str(row[2]).strip() if row[2] else ""
                if rut and rut != "None":
                    conn.execute(
                        "INSERT OR IGNORE INTO previred_empresas(rut,grupo,razon_social) VALUES(?,?,?)",
                        (rut, grupo, razon))
        except Exception as e:
            print(f"[previred] Error importando Excel: {e}")

_seed_previred_empresas()

@app.route("/empresas")
@login_required
def empresas_page():
    return render_template("empresas.html")

@app.route("/previred")
@login_required
def previred():
    return render_template("previred.html")

@app.route("/api/previred/empresas")
@api_login_required
def previred_empresas_list():
    q = request.args.get("q", "").strip().lower()
    grupo = request.args.get("grupo", "").strip()
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 20))
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, rut, grupo, razon_social FROM previred_empresas WHERE activa=1 ORDER BY grupo, razon_social"
        ).fetchall()
    items = [dict(r) for r in rows]
    if q:
        items = [i for i in items if q in i["rut"].lower() or q in i["grupo"].lower() or q in i["razon_social"].lower()]
    if grupo:
        items = [i for i in items if i["grupo"] == grupo]
    total = len(items)
    grupos = sorted(set(i["grupo"] for i in [dict(r) for r in rows] if i["grupo"]))
    start = (page - 1) * per_page
    return jsonify({"items": items[start:start+per_page], "total": total, "grupos": grupos, "page": page})

@app.route("/api/previred/empresas", methods=["POST"])
@api_login_required
def previred_empresa_crear():
    d = request.json or {}
    rut = d.get("rut", "").strip()
    if not rut:
        return jsonify({"error": "RUT requerido"}), 400
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO previred_empresas(rut, grupo, razon_social) VALUES(?,?,?)",
            (rut, d.get("grupo","").strip(), d.get("razon_social","").strip()))
    _exportar_previred_excel()
    return jsonify({"ok": True})

@app.route("/api/previred/empresas/<int:eid>", methods=["PUT"])
@api_login_required
def previred_empresa_editar(eid):
    d = request.json or {}
    with get_conn() as conn:
        conn.execute(
            "UPDATE previred_empresas SET rut=?, grupo=?, razon_social=? WHERE id=?",
            (d.get("rut","").strip(), d.get("grupo","").strip(), d.get("razon_social","").strip(), eid))
    _exportar_previred_excel()
    return jsonify({"ok": True})

@app.route("/api/previred/empresas/<int:eid>", methods=["DELETE"])
@api_login_required
def previred_empresa_eliminar(eid):
    with get_conn() as conn:
        conn.execute("UPDATE previred_empresas SET activa=0 WHERE id=?", (eid,))
    _exportar_previred_excel()
    return jsonify({"ok": True})

@app.route("/api/previred/empresas/export")
@api_login_required
def previred_export():
    import openpyxl, io
    from flask import send_file
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT rut, grupo, razon_social FROM previred_empresas WHERE activa=1 ORDER BY grupo, razon_social"
        ).fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empresas"
    ws.append(["1 RUT", "2 GRUPO", "5 RAZON SOCIAL"])
    for r in rows:
        ws.append([r["rut"], r["grupo"], r["razon_social"]])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="DATOS_RAZONES_SOCIALES.xlsx")

@app.route("/api/previred/empresas/import", methods=["POST"])
@api_login_required
def previred_import():
    import openpyxl, io
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "Sin archivo"}), 400
    reemplazar = request.form.get("reemplazar", "0") == "1"
    try:
        data = f.read()
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            rut = str(r[0]).strip() if r[0] is not None else ""
            grupo = str(r[1]).strip() if r[1] is not None else ""
            razon = str(r[2]).strip() if r[2] is not None else ""
            if rut and rut.lower() != "none":
                rows.append((rut, grupo, razon))
    except Exception as e:
        return jsonify({"error": f"Error leyendo Excel: {str(e)}"}), 400
    if not rows:
        return jsonify({"error": "El archivo no contiene datos válidos"}), 400
    try:
        conn = get_conn()
        if reemplazar:
            conn.execute("UPDATE previred_empresas SET activa=0")
        for rut, grupo, razon in rows:
            existing = conn.execute("SELECT id FROM previred_empresas WHERE rut=?", (rut,)).fetchone()
            if existing:
                conn.execute("UPDATE previred_empresas SET grupo=?, razon_social=?, activa=1 WHERE rut=?",
                             (grupo, razon, rut))
            else:
                conn.execute("INSERT INTO previred_empresas(rut, grupo, razon_social) VALUES(?,?,?)",
                             (rut, grupo, razon))
        conn.commit()
        conn.close()
    except Exception as e:
        return jsonify({"error": f"Error guardando: {str(e)}"}), 500
    _exportar_previred_excel()
    return jsonify({"ok": True, "importadas": len(rows)})

def _exportar_previred_excel():
    """Escribe el Excel en el volumen persistente del servidor."""
    import openpyxl, os as _os
    DEST = _os.path.join(_os.path.dirname(DB_PATH), "DATOS_RAZONES_SOCIALES.xlsx")
    try:
        with get_conn() as conn:
            rows = conn.execute(
                "SELECT rut, grupo, razon_social FROM previred_empresas WHERE activa=1 ORDER BY grupo, razon_social"
            ).fetchall()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["1 RUT", "2 GRUPO", "5 RAZON SOCIAL"])
        for r in rows:
            ws.append([r["rut"], r["grupo"], r["razon_social"]])
        wb.save(DEST)
    except Exception as e:
        print(f"[previred] Error exportando Excel: {e}")

# ============================================================
#  BASE DEUDAS AFP
# ============================================================

@app.route("/base-deudas")
@login_required
def base_deudas_page():
    return render_template("base_deudas.html")

@app.route("/api/base-deudas/unificar", methods=["POST"])
@api_login_required
def base_deudas_unificar():
    from pypdf import PdfWriter
    import io

    archivos = request.files.getlist("archivos")
    pdfs = [f for f in archivos if f.filename.lower().endswith(".pdf")]
    no_pdf = [f.filename for f in archivos if not f.filename.lower().endswith(".pdf")]
    if no_pdf:
        return jsonify({"error": f"El paso 1 solo acepta PDFs. Archivo incorrecto: {', '.join(no_pdf)}"}), 400
    if not pdfs:
        return jsonify({"error": "No se recibieron PDFs"}), 400

    from pypdf import PdfReader, PdfWriter as _W
    writer = _W()
    errores = []
    for f in pdfs:
        data = f.read()
        try:
            reader = PdfReader(io.BytesIO(data), strict=False)
            # intentar página por página para aislar PDFs problemáticos
            ok = False
            for page in reader.pages:
                try:
                    writer.add_page(page)
                    ok = True
                except Exception:
                    pass
            if not ok:
                errores.append(f.filename)
        except Exception as e:
            errores.append(f.filename)

    if not writer.pages:
        return jsonify({"error": f"No se pudo leer ningún PDF. Verifica que los archivos sean PDFs válidos. Errores: {', '.join(errores)}"}), 400

    try:
        buf = io.BytesIO()
        writer.write(buf)
        buf.seek(0)
    except Exception as e:
        return jsonify({"error": f"Error al combinar PDFs: {str(e)}"}), 500

    resp = send_file(buf, mimetype="application/pdf",
                     as_attachment=True, download_name="certificados_unificados.pdf")
    if errores:
        resp.headers["X-Errores"] = ", ".join(errores)
    return resp


@app.route("/api/base-deudas/procesar", methods=["POST"])
@api_login_required
def base_deudas_procesar():
    from base_deudas_logic import procesar_lote
    import base64, io

    archivos = request.files.getlist("archivos")
    if not archivos:
        return jsonify({"error": "No se recibieron archivos"}), 400

    pdfs   = {f.filename: f.read() for f in archivos if f.filename.lower().endswith(".pdf")}
    excels = {f.filename: f.read() for f in archivos if f.filename.lower().endswith(".xlsx")}

    if not excels:
        return jsonify({"error": "Debes subir el Excel generado por Adobe"}), 400

    # Armar pares: cada Excel se empareja con un PDF si hay (o con bytes vacíos)
    any_pdf_bytes = next(iter(pdfs.values()), b"")
    any_pdf_name  = next(iter(pdfs.keys()), "")
    pares = []
    for excel_nombre, excel_bytes in excels.items():
        stem = excel_nombre.replace("_ADOBE","").replace("_adobe","")
        pdf_nombre = stem if stem.lower().endswith(".pdf") else stem.rsplit(".",1)[0] + ".pdf"
        pdf_bytes  = pdfs.get(pdf_nombre, any_pdf_bytes)
        pdf_nombre = pdf_nombre if pdf_nombre in pdfs else any_pdf_name
        pares.append({
            "pdf_bytes": pdf_bytes, "pdf_nombre": pdf_nombre,
            "excel_bytes": excel_bytes, "excel_nombre": excel_nombre,
        })

    logs = []
    def _log(msg, tipo="info"):
        logs.append({"msg": msg, "tipo": tipo})

    try:
        resultado_bytes = procesar_lote(pares, log=_log)
    except Exception as e:
        return jsonify({"error": str(e), "logs": logs}), 500

    excel_b64 = base64.b64encode(resultado_bytes).decode()

    # Determinar nombre del archivo desde los datos procesados
    def _nombre_archivo():
        import openpyxl as _opx
        try:
            _wb = _opx.load_workbook(io.BytesIO(resultado_bytes))
            ruts_unicos   = set()
            razones_unicas = set()
            for _hoja in ["Base AFP", "Base Isapre"]:
                if _hoja not in _wb.sheetnames: continue
                _ws = _wb[_hoja]
                for _r in range(2, _ws.max_row + 1):
                    _rut = str(_ws.cell(_r, 1).value or "").strip()
                    _raz = str(_ws.cell(_r, 2).value or "").strip()
                    if _rut: ruts_unicos.add(_rut)
                    if _raz: razones_unicas.add(_raz)
            if not razones_unicas:
                return f"Base de deuda_{_time.strftime('%Y%m%d')}.xlsx"
            if len(razones_unicas) == 1:
                nombre = list(razones_unicas)[0][:60]
                return f"Base de deuda_{nombre}.xlsx"
            # Múltiples razones: buscar nombre del grupo en la DB
            conn = get_conn()
            for _rut in ruts_unicos:
                row = conn.execute(
                    "SELECT g.nombre FROM empresas e JOIN grupos g ON e.grupo_id=g.id WHERE e.rut=?",
                    (_rut,)
                ).fetchone()
                if row:
                    conn.close()
                    return f"Base de deuda_{row['nombre'][:60]}.xlsx"
            conn.close()
            # Fallback: usar la razón más larga como representativa
            nombre = max(razones_unicas, key=len)[:60]
            return f"Base de deuda_{nombre}.xlsx"
        except Exception:
            return f"Base de deuda_{_time.strftime('%Y%m%d')}.xlsx"

    nombre_archivo = _nombre_archivo()
    return jsonify({"ok": True, "excel_b64": excel_b64, "logs": logs,
                    "nombre": nombre_archivo})

# ============================================================
#  PREVIRED — AUTOMATIZACIÓN (descarga + conversión)
# ============================================================

DB_PATH = os.environ.get("DB_PATH",
          os.path.join(os.path.dirname(os.path.abspath(__file__)), "certificados.db"))
_DATA_ROOT = os.path.dirname(DB_PATH)
_PLANILLAS_DIR = os.path.join(_DATA_ROOT, "planillas")
_TEMP_DIR      = os.path.join(_DATA_ROOT, "temp_previred")
_EXCELS_DIR    = os.path.join(_DATA_ROOT, "excels")
for _d in [_PLANILLAS_DIR, _TEMP_DIR, _EXCELS_DIR]:
    os.makedirs(_d, exist_ok=True)

def _limpiar_archivos_antiguos(max_horas: int = 4):
    """Elimina archivos y carpetas de descargas más antiguos que max_horas."""
    limite = _time.time() - max_horas * 3600
    for directorio in [_PLANILLAS_DIR, _TEMP_DIR, _EXCELS_DIR]:
        if not os.path.isdir(directorio):
            continue
        for nombre in os.listdir(directorio):
            ruta = os.path.join(directorio, nombre)
            try:
                if os.path.getmtime(ruta) < limite:
                    if os.path.isdir(ruta):
                        shutil.rmtree(ruta, ignore_errors=True)
                    else:
                        os.remove(ruta)
            except Exception:
                pass

_limpiar_archivos_antiguos()

_tareas: dict = {}

# ── Configuración Previred (guardada en SQLite) ───────────────
def _get_previred_config() -> dict:
    conn = get_conn()
    rows = conn.execute("SELECT clave, valor FROM previred_config").fetchall()
    conn.close()
    return {r["clave"]: r["valor"] for r in rows}

def _set_previred_config(clave: str, valor: str):
    conn = get_conn()
    conn.execute("INSERT OR REPLACE INTO previred_config(clave, valor) VALUES(?,?)", (clave, valor))
    conn.commit()
    conn.close()

@app.route("/api/previred/config", methods=["GET"])
@api_login_required
def previred_config_get():
    cfg = _get_previred_config()
    return jsonify({
        "rut":          cfg.get("rut", ""),
        "pass_guardado": bool(cfg.get("pass", "")),
        "output_path":  cfg.get("output_path", ""),
    })

@app.route("/api/previred/config", methods=["POST"])
@api_login_required
def previred_config_set():
    d = request.json or {}
    rut         = d.get("rut", "").strip()
    pwd         = d.get("pwd", "").strip()
    output_path = d.get("output_path", "").strip()
    if rut:
        _set_previred_config("rut", rut)
    if pwd:
        _set_previred_config("pass", pwd)
    _set_previred_config("output_path", output_path)
    return jsonify({"ok": True})

def _nueva_tarea() -> str:
    tid = uuid.uuid4().hex[:10]
    _tareas[tid] = {"logs": [], "done": False, "error": False, "archivo": None, "zip": None}
    return tid

def _log(tid: str, msg: str, tipo: str = "info"):
    if tid in _tareas:
        _tareas[tid]["logs"].append({
            "msg": msg, "tipo": tipo,
            "t": _time.strftime("%H:%M:%S")
        })

@app.route("/api/previred/tarea/<tid>")
@api_login_required
def previred_tarea(tid):
    t = _tareas.get(tid)
    if not t:
        return jsonify({"error": "Tarea no encontrada"}), 404
    since = int(request.args.get("since", 0))
    return jsonify({
        "logs":    t["logs"][since:],
        "done":    t["done"],
        "error":   t["error"],
        "archivo": t["archivo"],
        "zip":     bool(t.get("zip")),
    })

@app.route("/api/previred/descargar-zip/<tid>")
@api_login_required
def previred_descargar_zip(tid):
    t = _tareas.get(tid)
    if not t or not t.get("zip") or not os.path.exists(t["zip"]):
        return jsonify({"error": "ZIP no disponible"}), 404
    return send_file(t["zip"], as_attachment=True,
                     download_name=os.path.basename(t["zip"]),
                     mimetype="application/zip")

@app.route("/api/previred/descargar-excel/<tid>")
@api_login_required
def previred_descargar_excel(tid):
    t = _tareas.get(tid)
    if not t or not t.get("archivo") or not os.path.exists(t["archivo"]):
        return jsonify({"error": "Archivo no disponible"}), 404
    return send_file(t["archivo"], as_attachment=True,
                     download_name="Planillas_Unificadas.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/previred/planillas")
@api_login_required
def previred_listar_planillas():
    archivos = sorted([
        f for f in os.listdir(_PLANILLAS_DIR) if f.endswith(".pdf")
    ])
    return jsonify({"archivos": archivos, "total": len(archivos)})

@app.route("/api/previred/iniciar", methods=["POST"])
@api_login_required
def previred_iniciar():
    d = request.json or {}
    tipo = d.get("tipo")         # descargar | convertir | ambos
    periodos_raw = d.get("periodos", [])  # [{"mes":1,"anio":2024}, ...]

    # Soporte multi-empresa: acepta {"empresas":[{rut,razon},...]} o legado {rut_empresa, razon_social}
    empresas = d.get("empresas", [])
    if not empresas:
        rut = d.get("rut_empresa", "").strip()
        rs  = d.get("razon_social", "").strip()
        if rut:
            empresas = [{"rut": rut, "razon": rs}]

    if tipo not in ("descargar", "convertir", "ambos"):
        return jsonify({"error": "Tipo inválido"}), 400

    periodos = [(int(p["mes"]), int(p["anio"])) for p in periodos_raw]

    tid = _nueva_tarea()

    def run():
        import traceback
        _oi = {"usar": False, "base": _PLANILLAS_DIR, "path": ""}
        try:
            _log(tid, f"Tarea iniciada — tipo: {tipo}", "info")
            _cfg_all = _get_previred_config()
            _op = (_cfg_all.get("output_path") or "").strip()
            _oi["usar"] = bool(_op and os.path.isdir(_op))
            _oi["base"] = _op if _oi["usar"] else _PLANILLAS_DIR
            _oi["path"] = _op
            if _oi["usar"]:
                _log(tid, f"Carpeta de destino: {_op}", "ok")

            if tipo in ("descargar", "ambos"):
                if not empresas:
                    _log(tid, "RUT de empresa requerido para descargar", "err")
                    _tareas[tid]["error"] = True
                    _tareas[tid]["done"]  = True
                    return
                if not periodos:
                    _log(tid, "Selecciona al menos un período", "err")
                    _tareas[tid]["error"] = True
                    _tareas[tid]["done"]  = True
                    return
                rut_usr  = os.environ.get("PREVIRED_RUT", "") or _cfg_all.get("rut", "")
                cont_usr = os.environ.get("PREVIRED_PASS", "") or _cfg_all.get("pass", "")
                if not rut_usr or not cont_usr:
                    _log(tid, "Credenciales Previred no configuradas", "err")
                    _log(tid, "Abre Configuración (⚙) en la página de Previred e ingresa tu RUT y contraseña", "warn")
                    _tareas[tid]["error"] = True
                    _tareas[tid]["done"]  = True
                    return
                _log(tid, "Verificando Chrome instalado...", "info")
                try:
                    import shutil as _sh, subprocess
                    chrome = _sh.which("chromium") or _sh.which("chromium-browser") or _sh.which("google-chrome")
                    if not chrome:
                        raise RuntimeError("No se encontró chromium ni google-chrome en el PATH")
                    result = subprocess.run([chrome, "--version"], capture_output=True, text=True, timeout=10)
                    _log(tid, f"Chrome listo: {result.stdout.strip()}", "ok")
                except Exception as ce:
                    _log(tid, f"Chrome no disponible: {ce}", "err")
                    _log(tid, "Railway necesita nixpacks.toml con chromium — verifica que el archivo existe en el repo", "warn")
                    _tareas[tid]["error"] = True
                    _tareas[tid]["done"]  = True
                    return
                from previred_logic import descargar
                todas_rutas_pdf = []
                for emp in empresas:
                    rut_empresa  = (emp.get("rut") or "").strip()
                    razon_social = (emp.get("razon") or "").strip()
                    if not rut_empresa:
                        continue
                    _log(tid, f"── Empresa: {rut_empresa} {('— ' + razon_social) if razon_social else ''}", "info")
                    carpeta_emp = os.path.join(_oi["base"], rut_empresa.replace(".", "").replace("-", ""))
                    os.makedirs(carpeta_emp, exist_ok=True)
                    descargar(rut_usr, cont_usr, rut_empresa, periodos,
                              carpeta_emp, _TEMP_DIR, lambda m, t: _log(tid, m, t),
                              razon_social=razon_social)
                    pdfs_emp = [os.path.join(carpeta_emp, f) for f in os.listdir(carpeta_emp) if f.endswith(".pdf")]
                    todas_rutas_pdf.extend(pdfs_emp)
                if todas_rutas_pdf:
                    tag = empresas[0].get("rut","").replace(".","").replace("-","")
                    if len(empresas) > 1:
                        tag += f"_y{len(empresas)-1}mas"
                    nombre_zip = f"Planillas_{tag}_{_time.strftime('%Y%m%d_%H%M%S')}.zip"
                    ruta_zip = os.path.join(_EXCELS_DIR, nombre_zip)
                    with zipfile.ZipFile(ruta_zip, "w", zipfile.ZIP_DEFLATED) as zf:
                        for rp in todas_rutas_pdf:
                            zf.write(rp, os.path.basename(rp))
                    _tareas[tid]["zip"] = ruta_zip
                    _log(tid, f"ZIP listo con {len(todas_rutas_pdf)} PDF(s)", "ok")
                    if _oi["usar"]:
                        shutil.copy2(ruta_zip, os.path.join(_oi["path"], nombre_zip))
                        _log(tid, f"ZIP guardado en: {_oi['path']}", "ok")

            if tipo in ("convertir", "ambos"):
                from pdf_excel_logic import generar_excel_bytes
                rut_empresa  = (empresas[0].get("rut")   if empresas else "") or ""
                razon_social = (empresas[0].get("razon") if empresas else "") or ""
                if tipo == "ambos" and empresas:
                    rutas = []
                    for emp in empresas:
                        rut_e = (emp.get("rut") or "").replace(".", "").replace("-", "")
                        carpeta_src = os.path.join(_oi["base"], rut_e)
                        if os.path.isdir(carpeta_src):
                            rutas += sorted([os.path.join(carpeta_src, f)
                                             for f in os.listdir(carpeta_src) if f.endswith(".pdf")])
                else:
                    carpeta_src = _oi["base"]
                    if not os.path.isdir(carpeta_src):
                        _log(tid, f"Carpeta no existe: {carpeta_src}", "err")
                        _tareas[tid]["error"] = True
                        _tareas[tid]["done"]  = True
                        return
                    rutas = sorted([os.path.join(carpeta_src, f)
                                    for f in os.listdir(carpeta_src) if f.endswith(".pdf")])
                if not rutas:
                    _log(tid, "No hay PDFs en la carpeta para convertir", "warn")
                    _tareas[tid]["done"] = True
                    return
                _log(tid, f"Convirtiendo {len(rutas)} PDF(s)...", "info")
                xls_bytes = generar_excel_bytes(
                    rutas, rut_empresa, razon_social,
                    log=lambda m, t: _log(tid, m, t)
                )
                nombre_archivo = f"Planillas_{_time.strftime('%Y%m%d_%H%M%S')}.xlsx"
                ruta_excel = os.path.join(_EXCELS_DIR, nombre_archivo)
                with open(ruta_excel, "wb") as f:
                    f.write(xls_bytes)
                _tareas[tid]["archivo"] = ruta_excel
                _log(tid, f"Excel listo: {nombre_archivo}", "ok")
                if _oi["usar"]:
                    shutil.copy2(ruta_excel, os.path.join(_oi["base"], nombre_archivo))
                    _log(tid, f"Excel guardado en: {_oi['base']}", "ok")

            _tareas[tid]["done"] = True
            _log(tid, "Proceso finalizado", "ok")
        except Exception as e:
            tb = traceback.format_exc()
            _log(tid, f"Error inesperado: {e}", "err")
            _log(tid, tb[:400], "err")
            _tareas[tid]["error"] = True
            _tareas[tid]["done"]  = True
        finally:
            if tipo in ("descargar", "ambos") and not _oi["usar"]:
                for emp in empresas:
                    rut_e = (emp.get("rut") or "").replace(".", "").replace("-", "")
                    shutil.rmtree(os.path.join(_PLANILLAS_DIR, rut_e), ignore_errors=True)
                for fn in os.listdir(_TEMP_DIR):
                    try:
                        os.remove(os.path.join(_TEMP_DIR, fn))
                    except Exception:
                        pass

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"task_id": tid})


# ============================================================
#  NOTAS
# ============================================================

PALABRAS_TAREA_DEFAULT = [
    'llamar','crear','realizar','revisar','enviar','reunión','reunion',
    'pagar','gestionar','solicitar','confirmar','contactar','hacer',
    'subir','mandar','completar','preparar','coordinar','agendar','verificar'
]

def _detectar_tareas(texto, palabras):
    tl = texto.lower()
    return any(p.lower() in tl for p in palabras)

def _pulir_texto_nota(texto):
    """Corrige ortografía/redacción del texto dictado a Opti, sin cambiar el sentido."""
    api_key = os.environ.get("OPENAI_API_KEY","")
    if not api_key or not texto.strip():
        return texto
    try:
        import urllib.request, json as jsonlib
        data = jsonlib.dumps({
            "model": "gpt-3.5-turbo",
            "messages": [
                {"role": "system", "content": "Corrige ortografía, tildes y puntuación del texto del usuario, sin cambiar su significado ni agregar contenido. Responde solo con el texto corregido, sin comillas ni explicaciones."},
                {"role": "user", "content": texto}
            ],
            "max_tokens": 300,
            "temperature": 0.2
        }).encode('utf-8')
        req = urllib.request.Request(
            "https://api.openai.com/v1/chat/completions",
            data=data,
            headers={"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"})
        with urllib.request.urlopen(req, timeout=10) as r:
            resp = jsonlib.loads(r.read())
        return resp["choices"][0]["message"]["content"].strip().strip('"')
    except Exception as e:
        print(f"[OPTI] Error puliendo texto: {e}")
        return texto

def _crear_nota_desde_texto(user_id, texto, etiqueta="info", color="amarillo"):
    """Crea una nota (y tarea si corresponde) a partir de texto libre. Usado por Opti."""
    titulo = texto.strip()
    if len(titulo) > 70:
        titulo = titulo[:67] + "..."
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO notas(usuario_id,titulo,contenido,etiqueta,color,fijada,creada,actualizada) VALUES(?,?,?,?,?,?,?,?)",
            (user_id, titulo, texto.strip(), etiqueta, color, 0, now, now))
        nid = cur.lastrowid
        cfg = conn.execute("SELECT palabras_clave FROM config_recordatorios WHERE usuario_id=?", (user_id,)).fetchone()
        palabras = (cfg["palabras_clave"].split(",") if cfg and cfg["palabras_clave"] else PALABRAS_TAREA_DEFAULT)
        tarea_id = None
        if _detectar_tareas(texto, palabras):
            prioridad = "alta" if etiqueta == "urgente" else "media"
            cur2 = conn.execute(
                "INSERT INTO tareas(usuario_id,titulo,descripcion,prioridad,estado,nota_id,creada) VALUES(?,?,?,?,?,?,?)",
                (user_id, titulo, texto.strip(), prioridad, "pendiente", nid, now))
            tarea_id = cur2.lastrowid
    return nid, tarea_id

@app.route("/notas")
@login_required
def notas_page():
    return render_template("notas.html")

@app.route("/tareas")
@login_required
def tareas_page():
    return render_template("tareas.html")

@app.route("/api/notas")
@api_login_required
def get_notas():
    user = get_current_user()
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM notas WHERE usuario_id=? ORDER BY fijada DESC, actualizada DESC",
            (user["id"],)).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/notas", methods=["POST"])
@api_login_required
def crear_nota():
    user = get_current_user()
    d = request.json or {}
    titulo = d.get("titulo","").strip()
    if not titulo:
        return jsonify({"error":"Título requerido"}), 400
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO notas(usuario_id,titulo,contenido,etiqueta,color,fijada,creada,actualizada) VALUES(?,?,?,?,?,?,?,?)",
            (user["id"], titulo, d.get("contenido",""), d.get("etiqueta","info"),
             d.get("color","amarillo"), 1 if d.get("fijada") else 0, now, now))
        nid = cur.lastrowid
        # Auto-detectar tarea
        cfg = conn.execute("SELECT palabras_clave FROM config_recordatorios WHERE usuario_id=?", (user["id"],)).fetchone()
        palabras = (cfg["palabras_clave"].split(",") if cfg and cfg["palabras_clave"] else PALABRAS_TAREA_DEFAULT)
        texto_completo = titulo + " " + d.get("contenido","")
        tarea_id = None
        if _detectar_tareas(texto_completo, palabras):
            prioridad = "alta" if d.get("etiqueta") == "urgente" else "media"
            cur2 = conn.execute(
                "INSERT INTO tareas(usuario_id,titulo,descripcion,prioridad,estado,nota_id,creada) VALUES(?,?,?,?,?,?,?)",
                (user["id"], titulo, d.get("contenido",""), prioridad, "pendiente", nid, now))
            tarea_id = cur2.lastrowid
    return jsonify({"id": nid, "tarea_creada": tarea_id is not None, "tarea_id": tarea_id}), 201

@app.route("/api/notas/<int:nid>", methods=["PUT"])
@api_login_required
def editar_nota(nid):
    user = get_current_user()
    d = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        nota = conn.execute("SELECT * FROM notas WHERE id=? AND usuario_id=?", (nid, user["id"])).fetchone()
        if not nota:
            return jsonify({"error":"No encontrada"}), 404
        conn.execute(
            "UPDATE notas SET titulo=?,contenido=?,etiqueta=?,color=?,fijada=?,actualizada=? WHERE id=?",
            (d.get("titulo", nota["titulo"]), d.get("contenido", nota["contenido"]),
             d.get("etiqueta", nota["etiqueta"]), d.get("color", nota["color"]),
             1 if d.get("fijada") else 0, now, nid))
    return jsonify({"ok": True})

@app.route("/api/notas/<int:nid>", methods=["DELETE"])
@api_login_required
def eliminar_nota(nid):
    user = get_current_user()
    with get_conn() as conn:
        conn.execute("DELETE FROM notas WHERE id=? AND usuario_id=?", (nid, user["id"]))
    return jsonify({"ok": True})

# ============================================================
#  TAREAS
# ============================================================

@app.route("/api/tareas")
@api_login_required
def get_tareas():
    user = get_current_user()
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT t.*, n.titulo as nota_titulo
            FROM tareas t LEFT JOIN notas n ON n.id = t.nota_id
            WHERE t.usuario_id=?
            ORDER BY CASE t.prioridad WHEN 'alta' THEN 0 WHEN 'media' THEN 1 ELSE 2 END,
                     t.estado, t.creada DESC""",
            (user["id"],)).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/tareas", methods=["POST"])
@api_login_required
def crear_tarea():
    user = get_current_user()
    d = request.json or {}
    titulo = d.get("titulo","").strip()
    if not titulo:
        return jsonify({"error":"Título requerido"}), 400
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO tareas(usuario_id,titulo,descripcion,prioridad,estado,fecha_limite,nota_id,creada) VALUES(?,?,?,?,?,?,?,?)",
            (user["id"], titulo, d.get("descripcion",""), d.get("prioridad","media"),
             "pendiente", d.get("fecha_limite",""), d.get("nota_id"), now))
    return jsonify({"id": cur.lastrowid}), 201

@app.route("/api/tareas/<int:tid>", methods=["PUT"])
@api_login_required
def editar_tarea(tid):
    user = get_current_user()
    d = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        t = conn.execute("SELECT * FROM tareas WHERE id=? AND usuario_id=?", (tid, user["id"])).fetchone()
        if not t:
            return jsonify({"error":"No encontrada"}), 404
        completada_en = now if d.get("estado") == "completada" and t["estado"] != "completada" else t["completada_en"]
        conn.execute(
            "UPDATE tareas SET titulo=?,descripcion=?,prioridad=?,estado=?,fecha_limite=?,completada_en=? WHERE id=?",
            (d.get("titulo", t["titulo"]), d.get("descripcion", t["descripcion"]),
             d.get("prioridad", t["prioridad"]), d.get("estado", t["estado"]),
             d.get("fecha_limite", t["fecha_limite"]), completada_en, tid))
    return jsonify({"ok": True})

@app.route("/api/tareas/<int:tid>", methods=["DELETE"])
@api_login_required
def eliminar_tarea(tid):
    user = get_current_user()
    with get_conn() as conn:
        conn.execute("DELETE FROM tareas WHERE id=? AND usuario_id=?", (tid, user["id"]))
    return jsonify({"ok": True})

# ── Config recordatorios ──────────────────────────────────────

@app.route("/api/recordatorios/config")
@api_login_required
def get_config_recordatorios():
    user = get_current_user()
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM config_recordatorios WHERE usuario_id=?", (user["id"],)).fetchone()
    if row:
        return jsonify(dict(row))
    return jsonify({
        "usuario_id": user["id"], "notas_activo": 1, "tareas_activo": 1,
        "hora_1": "09:00", "hora_2": "13:00", "hora_3": "17:00",
        "palabras_clave": ",".join(PALABRAS_TAREA_DEFAULT)
    })

@app.route("/api/recordatorios/config", methods=["POST"])
@api_login_required
def set_config_recordatorios():
    user = get_current_user()
    d = request.json or {}
    with get_conn() as conn:
        conn.execute("""INSERT INTO config_recordatorios(usuario_id,notas_activo,tareas_activo,hora_1,hora_2,hora_3,palabras_clave)
            VALUES(?,?,?,?,?,?,?)
            ON CONFLICT(usuario_id) DO UPDATE SET
                notas_activo=excluded.notas_activo, tareas_activo=excluded.tareas_activo,
                hora_1=excluded.hora_1, hora_2=excluded.hora_2, hora_3=excluded.hora_3,
                palabras_clave=excluded.palabras_clave""",
            (user["id"], 1 if d.get("notas_activo",True) else 0,
             1 if d.get("tareas_activo",True) else 0,
             d.get("hora_1","09:00"), d.get("hora_2","13:00"), d.get("hora_3","17:00"),
             d.get("palabras_clave", ",".join(PALABRAS_TAREA_DEFAULT))))
    return jsonify({"ok": True})

@app.route("/api/recordatorios/pendientes")
@api_login_required
def get_pendientes_resumen():
    user = get_current_user()
    with get_conn() as conn:
        notas_count = conn.execute(
            "SELECT COUNT(*) FROM notas WHERE usuario_id=? AND etiqueta IN ('urgente','pendiente')", (user["id"],)).fetchone()[0]
        tareas_count = conn.execute(
            "SELECT COUNT(*) FROM tareas WHERE usuario_id=? AND estado='pendiente'", (user["id"],)).fetchone()[0]
        tareas_vencidas = conn.execute(
            "SELECT COUNT(*) FROM tareas WHERE usuario_id=? AND estado='pendiente' AND fecha_limite != '' AND fecha_limite < ?",
            (user["id"], datetime.now().strftime("%Y-%m-%d"))).fetchone()[0]
    return jsonify({"notas_urgentes": notas_count, "tareas_pendientes": tareas_count, "tareas_vencidas": tareas_vencidas})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
