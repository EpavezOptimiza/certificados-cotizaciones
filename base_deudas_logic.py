"""
Procesa pares PDF + Excel Adobe de certificados AFP.
Portado de limpiar_excel_adobe.py para funcionar con bytes en memoria.
"""
import io, re, os
from datetime import date
from collections import defaultdict
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "static", "template_base_deudas.xlsx")

# ── Constantes ────────────────────────────────────────────────────────────────
INSTITUCION_DEFAULT = "DESCONOCIDA"
AFP_NOMBRES = ["PLANVITAL", "CAPITAL", "CUPRUM", "HABITAT", "PROVIDA", "MODELO", "UNO"]

# Mapa de keywords → nombre canónico de institución (orden importa: más específico primero)
INSTITUCIONES_MAP = [
    ("PLANVITAL",     "AFP PLANVITAL"),
    ("CAPITAL",       "AFP CAPITAL"),
    ("CUPRUM",        "AFP CUPRUM"),
    ("HABITAT",       "AFP HABITAT"),
    ("PROVIDA",       "AFP PROVIDA"),
    ("MODELO",        "AFP MODELO"),
    ("UNO",           "AFP UNO"),
    ("CRUZ BLANCA",   "ISAPRE CRUZ BLANCA"),
    ("CRUZBLANCA",    "ISAPRE CRUZ BLANCA"),
    ("CONSALUD",      "ISAPRE CONSALUD"),
    ("NUEVA MASVIDA", "ISAPRE NUEVA MASVIDA"),
    ("MASVIDA",       "ISAPRE NUEVA MASVIDA"),
    ("AFC",           "AFC"),
]

def _detectar_inst_texto(*textos: str) -> str:
    """Detecta institución buscando keywords en cualquiera de los textos dados."""
    upper = " ".join(t.upper() for t in textos)
    for keyword, nombre in INSTITUCIONES_MAP:
        if keyword in upper:
            return nombre
    return INSTITUCION_DEFAULT

RE_INSTITUCION   = re.compile(r"^(AFP\s+\w+(?:\s+\w+)?)\s+S\.?\s*A\.?,", re.IGNORECASE)
RE_RUT_EMPRESA   = re.compile(r"R\.?U\.?T\.?\s*:?\s*(\d{1,2}\.\d{3}\.\d{3}\s*-\s*[\dKk])", re.IGNORECASE)
RE_RAZON_SOCIAL  = re.compile(
    r"(?:(?:el|al|del)\s+empleador\s+(?:cuya\s+raz[o6ó]n\s+social\s+es\s+)?|empleador\s+cuya\s+raz[o6ó]n\s+social\s+es\s+)(.+?)(?:,?\s+R\.?U\.?T\.?|,?\s+Rut\s*:)",
    re.IGNORECASE)
RE_RAZON_PREFIJO = re.compile(r"^(?:cuya\s+raz[o6ó]n\s+social\s+es\s+)", re.IGNORECASE)
RE_RAZON_HABITAT = re.compile(r"(?:Se\w{1,5}res?\s*[\n\r\s]*)(.+?)\s+Rut\s*:", re.IGNORECASE | re.DOTALL)
RE_NO_DEUDA      = re.compile(r"CERTIFICADO\s+DE\s+NO\s+DEUDA|REGISTRO\s+DE\s+NO\s+DEUDA", re.IGNORECASE)
RE_RUT_SIN_PUNTOS = re.compile(r"[Rr]ut\s+(\d{7,8}-[\dKk])\b")
RE_RUT_EMPLEADOR  = re.compile(r"R\.?U\.?T\.?\s+[Ee]mpleador\s+(\d{1,2}\.\d{3}\.\d{3}\s*-\s*[\dKk])", re.IGNORECASE)
RE_RAZON_CERTIFICA = re.compile(r"certifica que[:\s]+([A-ZÁÉÍÓÚÑ][A-Za-záéíóúñÁÉÍÓÚÑ\s\.]+?),?\s+[Rr]ut\s", re.IGNORECASE)
RE_PERIODO_STR   = re.compile(r"^(\d{2})/(\d{4})")

def _normalizar_rut_sin_puntos(rut_raw: str) -> str:
    """'76003477-0' → '76.003.477-0', '7003477-0' → '7.003.477-0'"""
    parts = rut_raw.split("-")
    if len(parts) != 2:
        return rut_raw
    digits, dv = parts[0].strip(), parts[1].strip()
    if len(digits) == 8:
        return f"{digits[0:2]}.{digits[2:5]}.{digits[5:]}-{dv}"
    if len(digits) == 7:
        return f"{digits[0]}.{digits[1:4]}.{digits[4:]}-{dv}"
    return rut_raw

def _extraer_rut(texto: str) -> str:
    """Intenta múltiples formatos de RUT; devuelve RUT normalizado sin espacios."""
    m = RE_RUT_EMPRESA.search(texto)
    if m:
        return re.sub(r"\s", "", m.group(1))
    m = RE_RUT_EMPLEADOR.search(texto)
    if m:
        return re.sub(r"\s", "", m.group(1))
    m = RE_RUT_SIN_PUNTOS.search(texto)
    if m:
        return re.sub(r"\s", "", _normalizar_rut_sin_puntos(m.group(1)))
    return ""

def _extraer_razon(texto: str) -> str:
    """Intenta múltiples patrones de razón social."""
    m = RE_RAZON_SOCIAL.search(texto) or RE_RAZON_HABITAT.search(texto)
    if m:
        return _limpiar_razon(m.group(1))
    m = RE_RAZON_CERTIFICA.search(texto)
    if m:
        return _limpiar_razon(m.group(1))
    return ""
RE_GRUPO_PDF     = re.compile(
    r"^(PLANILLAS COMPLEMENTARIAS|DECL\.?\s*Y NO PAGO\s*AUTOM\.?\s*\(?DNPA\)?)\s+"
    r"\d+\s+(\d{2}/\d{4})\s+([\d\.]+)\s+([\d\.]+)\s+", re.IGNORECASE)
RE_DETALLE_PDF   = re.compile(r"^(\d{1,2}\.\d{3}\.\d{3}-[\dKk])\s+.+?\s+([\d\.]+)\s+([\d\.]+)\s*$")
RUT_RE           = re.compile(r"^\d{1,2}\.\d{3}\.\d{3}-[\dKk]$")
RUT_CON_NOMBRE   = re.compile(r"^(\d{1,2}\.\d{3}\.\d{3}-[\dKk])\s+(.+)")
RE_NUD_HDR       = re.compile(r'n[uú]mero\s+\w+(?:\s+de)?\s+deuda[^0-9]*(\d{6,})', re.IGNORECASE)
RE_RUT_HDR       = re.compile(r"^r\.u\.t\.", re.IGNORECASE)
RE_RUT_EMP       = re.compile(r"^\d{1,2}\.\d{3}\.\d{3}-[\dKk]$")

NOISE_TEXTS = {"este certificado", "certificado de deudas", "afp   planvital", "santiago,"}
STOP_TEXTS  = {"resumen", "total resumen", "total general", "administradora", "r.u.t."}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _limpiar_razon(razon: str) -> str:
    razon = RE_RAZON_PREFIJO.sub("", razon.strip())
    return re.sub(r"\s+", " ", razon).strip().rstrip(",")

def _periodo(s):
    import datetime as _dt
    if s is None or s == "":
        return s
    # Ya es date/datetime (Adobe exporta fechas como objetos)
    if isinstance(s, (_dt.date, _dt.datetime)):
        return date(s.year, s.month, 1)
    s = str(s).strip().lstrip("'")
    if not s:
        return s
    # Formato estándar MM/YYYY
    m = RE_PERIODO_STR.match(s)
    if m:
        mes, anio = int(m.group(1)), int(m.group(2))
        if 1 <= mes <= 12 and 1900 <= anio <= 2100:
            return date(anio, mes, 1)
    # Formato YYYY-MM-DD (date serializado como string)
    m3 = re.match(r'^(\d{4})-(\d{2})-\d{2}', s)
    if m3:
        anio, mes = int(m3.group(1)), int(m3.group(2))
        if 1 <= mes <= 12 and 1900 <= anio <= 2100:
            return date(anio, mes, 1)
    # Formato AFP Habitat: M1YYYY o MM1YYYY (el "1" es el día, ej: 9/1/2019 → 912019)
    m2 = re.match(r'^(\d{1,2})1(\d{4})$', s)
    if m2:
        mes, anio = int(m2.group(1)), int(m2.group(2))
        if 1 <= mes <= 12 and 1990 <= anio <= 2100:
            return date(anio, mes, 1)
    return s

def _limpiar_num_pdf(s: str) -> int:
    return int(s.replace(".", ""))

def _convertir_monto(val) -> int:
    if val is None:
        return 0
    if isinstance(val, str):
        limpio = val.replace("$","").replace(" ","").replace(",","").strip("'\" ")
        if "/" in limpio or not limpio.replace(".","").isdigit():
            return 0
        return int(limpio.replace(".",""))
    if isinstance(val, float):
        return round(val * 1000)
    return int(val)

def _es_monto(val) -> bool:
    if val is None: return False
    if isinstance(val, (int, float)): return True
    if isinstance(val, str):
        limpio = val.replace("$","").replace(" ","").replace(",","").strip("'\" ")
        return bool(limpio) and "/" not in limpio and limpio.replace(".","").isdigit()
    return False

def _es_ruido(val: str) -> bool:
    lower = val.lower()
    return any(t in lower for t in NOISE_TEXTS)

def _es_stop(val: str) -> bool:
    return any(val.lower().strip().startswith(t) for t in STOP_TEXTS)

def _es_grupo(b_val) -> bool:
    if not isinstance(b_val, str): return False
    return any(x in b_val.upper() for x in ["PLANILLAS","PAGO AUTOM","ECL.","DECL.","DNPA"])

def _norm_estado(estado: str) -> str:
    upper = estado.upper()
    if any(x in upper for x in ["INGRESADA","TRIBUNAL","JUICIO"]): return "JUICIO"
    if "PREJUDICIAL" in upper: return "PREJUDICIAL"
    return estado

def _norm_origen(b_val: str) -> str:
    return "PLANILLAS COMPLEMENTARIAS" if "PLANILLAS" in b_val.upper() else "DECL. Y NO PAGO AUTOM. (DNPA)"


# ── Detección ─────────────────────────────────────────────────────────────────

def detectar_institucion(pdf_bytes: bytes, nombre_archivo: str = "") -> str:
    if pdf_bytes and len(pdf_bytes) > 256:
        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                texto = pdf.pages[0].extract_text() or ""
            for linea in texto.splitlines():
                m = RE_INSTITUCION.match(linea.strip())
                if m:
                    return m.group(1).upper().strip()
        except Exception:
            pass
    return _detectar_inst_texto(nombre_archivo)

RE_RAZON_AFC = re.compile(
    r"(?:empleador\s+)([A-ZÁÉÍÓÚÑ][A-Za-záéíóúñÁÉÍÓÚÑ0-9\s\.,]+?),?\s+RUT\s+(\d{1,2}\.\d{3}\.\d{3}-[\dKk])",
    re.IGNORECASE)

def detectar_no_deuda(pdf_bytes: bytes) -> tuple:
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        texto = " ".join((p.extract_text() or "") for p in pdf.pages[:2])
    if not RE_NO_DEUDA.search(texto):
        return False, "", ""
    rut   = _extraer_rut(texto)
    razon = _extraer_razon(texto)
    # Fallback AFC: "empleador NOMBRE, RUT XX.XXX.XXX-X"
    if not razon or not rut:
        m_afc = RE_RAZON_AFC.search(texto)
        if m_afc:
            if not razon: razon = m_afc.group(1).strip().rstrip(",")
            if not rut:   rut   = m_afc.group(2)
    return True, rut, razon

def extraer_datos_empresa(ws) -> tuple:
    for r in range(1, min(12, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if not isinstance(val, str):
                continue
            rut = _extraer_rut(val)
            if rut:
                razon = _extraer_razon(val)
                return rut, razon
    return "", ""


# ── Lectura PDF (fuente de verdad) ────────────────────────────────────────────

def leer_totales_pdf(pdf_bytes: bytes) -> dict:
    totales = {}
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            for linea in (page.extract_text() or "").splitlines():
                m = RE_GRUPO_PDF.match(linea.strip())
                if m:
                    origen  = _norm_origen(m.group(1))
                    periodo = _periodo(m.group(2))
                    act     = _limpiar_num_pdf(m.group(4))
                    clave   = (origen, periodo)
                    totales[clave] = totales.get(clave, 0) + act
    return totales

def leer_montos_pdf(pdf_bytes: bytes) -> dict:
    lookup = {}
    periodo_actual = ""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            for linea in (page.extract_text() or "").splitlines():
                linea = linea.strip()
                if not linea: continue
                m_g = RE_GRUPO_PDF.match(linea)
                if m_g:
                    periodo_actual = m_g.group(2)
                    continue
                m_d = RE_DETALLE_PDF.match(linea)
                if m_d and periodo_actual:
                    rut  = m_d.group(1)
                    nom  = _limpiar_num_pdf(m_d.group(2))
                    act  = _limpiar_num_pdf(m_d.group(3))
                    clave = (rut, _periodo(periodo_actual))
                    lookup.setdefault(clave, []).append((nom, act))
    return lookup


# ── Corrección de montos ──────────────────────────────────────────────────────

def _corregir_con_pdf(monto_excel: int, candidatos: list, campo: str) -> tuple:
    idx = 0 if campo == "nom" else 1
    for cand in candidatos:
        pdf_val = cand[idx]
        if pdf_val == monto_excel:
            return monto_excel, False
        for factor in (10, 100, 1000):
            if monto_excel * factor == pdf_val:
                return pdf_val, True
    return monto_excel, False

def autocorregir(filas: list, ws, totales_pdf: dict) -> tuple:
    correcciones = []
    grupos = defaultdict(list)
    for i, f in enumerate(filas):
        grupos[(f.get("origen",""), f.get("periodo",""))].append(i)

    for clave, pdf_total in totales_pdf.items():
        indices = grupos.get(clave, [])
        if not indices: continue
        total_actual = sum(filas[i]["monto_act"] for i in indices)
        if total_actual == pdf_total: continue

        filas_excel = [filas[i]["_fila_excel"] for i in indices]
        mejor_col   = None
        mejor_diff  = abs(pdf_total - total_actual)

        for col in range(2, ws.max_column + 1):
            vals = []
            for fr in filas_excel:
                v = ws.cell(fr, col).value
                vals.append(_convertir_monto(v) if _es_monto(v) else None)
            validos = [v for v in vals if v is not None]
            if validos:
                total_col = sum(validos)
                if total_col == pdf_total and len(validos) == len(indices):
                    mejor_col  = col
                    mejor_diff = 0
                    break
                elif abs(pdf_total - total_col) < mejor_diff:
                    mejor_diff = abs(pdf_total - total_col)

        if mejor_col is not None:
            for idx, fr in zip(indices, filas_excel):
                v = ws.cell(fr, mejor_col).value
                if _es_monto(v):
                    filas[idx]["monto_act"] = _convertir_monto(v)
            correcciones.append(f"[AUTO] {clave[0]} {clave[1]}: total {total_actual:,} → {pdf_total:,}")
        else:
            correcciones.append(f"[REVISAR] {clave[0]} {clave[1]}: extraido={total_actual:,} PDF={pdf_total:,}")

    return filas, correcciones


# ── Parseo Excel Habitat ──────────────────────────────────────────────────────

def _es_habitat(ws) -> bool:
    for r in range(1, 8):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if isinstance(val, str) and val.strip().strip("'").upper() == "NUD":
                return True
    return False

def _parsear_habitat(wb, pdf_lookup: dict) -> tuple:
    ws    = wb.active
    filas = []
    _EXCLUIR = {"origen deuda","nud","totales",""}

    COL_ORIGEN_P1=None; COL_PERIODO_P1=None; COL_NOM_P1=None; COL_TOT_P1=None; COL_EST_P1=None
    for _r in range(1, min(10, ws.max_row+1)):
        if ws.cell(_r,1).value == "NUD" or (isinstance(ws.cell(_r,1).value,str) and "NUD" in str(ws.cell(_r,1).value).upper().split()):
            for _c in range(1, ws.max_column+1):
                _v = str(ws.cell(_r,_c).value or "").lower()
                if ("origen" in _v or "rtgen" in _v or "orl" in _v) and "deuda" in _v:
                    COL_ORIGEN_P1=_c
                elif ("peri" in _v or "perl" in _v or "cotiz" in _v) and COL_PERIODO_P1 is None: COL_PERIODO_P1=_c
                elif ("nominal" in _v or "noml" in _v) and "ondo" in _v: COL_NOM_P1=_c
                elif "total" in _v and ("pag" in _v or "peger" in _v): COL_TOT_P1=_c
                elif "jur" in _v or "estudio" in _v: COL_EST_P1=_c
            break

    # Fallback: detectar columnas desde la primera fila de datos NUD real
    if None in (COL_ORIGEN_P1, COL_PERIODO_P1, COL_NOM_P1, COL_TOT_P1):
        for r in range(1, min(15, ws.max_row+1)):
            nud_v = ws.cell(r, 1).value
            if not isinstance(nud_v, int) or nud_v < 100000:
                continue
            nums, strs = [], []
            for c in range(2, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip().lower() not in _EXCLUIR:
                    strs.append(c)
                elif isinstance(v, (int, float)) and v > 0:
                    nums.append(c)
            if strs and COL_ORIGEN_P1 is None:
                COL_ORIGEN_P1 = strs[0]
            if COL_PERIODO_P1 is None:
                # Buscar período por valor: string "MM/YYYY" o fecha Excel
                import datetime as _dt
                for _c2 in range(2, ws.max_column + 1):
                    _pv = ws.cell(r, _c2).value
                    if isinstance(_pv, str) and re.match(r'^\d{2}/\d{4}$', _pv.strip()):
                        COL_PERIODO_P1 = _c2; break
                    if isinstance(_pv, (_dt.date, _dt.datetime)):
                        COL_PERIODO_P1 = _c2; break
            if len(nums) >= 2 and COL_NOM_P1 is None:
                COL_NOM_P1 = nums[0]
            if len(nums) >= 2 and COL_TOT_P1 is None:
                COL_TOT_P1 = nums[-1]
            if COL_EST_P1 is None:
                # buscar string en los últimos cols
                for c in range(ws.max_column, 1, -1):
                    v = ws.cell(r, c).value
                    if isinstance(v, str) and v.strip():
                        COL_EST_P1 = c; break
            break

    # Valores por defecto si aún sin detectar
    if COL_ORIGEN_P1 is None:  COL_ORIGEN_P1 = 3
    if COL_PERIODO_P1 is None: COL_PERIODO_P1 = 4
    if COL_NOM_P1 is None:     COL_NOM_P1 = 6
    if COL_TOT_P1 is None:     COL_TOT_P1 = 21
    if COL_EST_P1 is None:     COL_EST_P1 = 23

    nud_info = {}
    for r in range(1, ws.max_row+1):
        nud = ws.cell(r,1).value
        v_or = ws.cell(r,COL_ORIGEN_P1).value
        if not isinstance(nud,int): continue
        if nud < 100000: continue
        if not (isinstance(v_or,str) and v_or.strip().lower() not in _EXCLUIR): continue
        nud_info[nud] = {
            "periodo": _periodo(str(ws.cell(r,COL_PERIODO_P1).value or "").strip()),
            "monto_nom": _convertir_monto(ws.cell(r,COL_NOM_P1).value),
            "total_pagar": _convertir_monto(ws.cell(r,COL_TOT_P1).value),
            "estado": _norm_estado(str(ws.cell(r,COL_EST_P1).value or "").upper().strip()),
            "origen": _norm_origen(v_or.strip()),
        }

    grupos = {}
    nuds_ord  = list(nud_info.keys())
    curr_nud  = nuds_ord[0] if nuds_ord else None
    in_part2  = False
    nombre_col = 3   # columna del nombre en Part 2; se detecta desde el header R.U.T.
    total_col  = 24  # columna "Total" en la tabla de detalle por empleado

    for r in range(1, ws.max_row+1):
        v1 = ws.cell(r,1).value
        # Entero pequeño = fragmento de RUT (ej: 13 de "13.133.756-6"), tratar como string
        if isinstance(v1, int) and v1 < 100000:
            v1 = str(v1)
        if not isinstance(v1,str): continue
        v1s = v1.strip()
        if not in_part2 and RE_RUT_HDR.match(v1s):
            in_part2 = True
            # Detectar columnas nombre y Total desde el header "R.U.T. | ... | Nombre | ... | Total"
            for _c in range(2, ws.max_column + 1):
                _hv = str(ws.cell(r, _c).value or "").strip().lower()
                if "nomb" in _hv:
                    nombre_col = _c
                elif _hv == "total":
                    total_col = _c
            continue
        if not in_part2: continue
        m_nud = RE_NUD_HDR.search(v1s)
        if m_nud:
            curr_nud = int(m_nud.group(1)); continue
        if RE_RUT_HDR.match(v1s):
            # Nuevo bloque de detalle: re-detectar columnas
            for _c in range(2, ws.max_column + 1):
                _hv = str(ws.cell(r, _c).value or "").strip().lower()
                if "nomb" in _hv:
                    nombre_col = _c
                elif _hv == "total":
                    total_col = _c
            continue
        if any(x in v1s.lower() for x in ["totales","se extiende","saluda"]): break
        # RUT a veces dividido en varias celdas por OCR
        # Casos: "13.133.756-6" (completo), "13"|"133.756-6", "13"|"133"|"756-6"
        rut_candidato = v1s
        if re.match(r'^\d{1,2}$', v1s):
            c2_val = ws.cell(r, 2).value
            c2 = str(c2_val).strip() if c2_val is not None else ""
            if re.match(r'^\d{3}$', c2):
                # Tres partes: C1=RR, C2=DDD, C3=DDD-K
                c3 = str(ws.cell(r, 3).value or "").strip()
                if re.match(r'^\d{3}-[\dKk]$', c3):
                    rut_candidato = f"{v1s}.{c2}.{c3}"
                else:
                    rut_candidato = f"{v1s}.{c2}"
            elif c2:
                rut_candidato = f"{v1s}.{c2}"
        if curr_nud and RE_RUT_EMP.match(rut_candidato):
            nombre_raw = str(ws.cell(r, nombre_col).value or "").strip().strip("'")
            nombre = re.sub(r'\s+', ' ', nombre_raw)
            total_raw = ws.cell(r, total_col).value
            total_emp = _convertir_monto(total_raw) if total_raw is not None else 0
            grupos.setdefault(curr_nud, []).append((rut_candidato, nombre, total_emp))

    for nud, info in nud_info.items():
        emps = grupos.get(nud,[])
        if not emps:
            filas.append({"rut":"","nombre":"","monto_nom":info["monto_nom"],
                          "monto_act":info["total_pagar"],"_fila_excel":0,
                          "origen":info["origen"],"adm":None,
                          "periodo":info["periodo"],"estado":info["estado"],"abogado":""})
        else:
            # Dividir total_pagar solo entre empleados con total > 0
            emps_con_deuda = [e for e in emps if e[2] > 0]
            n_div = len(emps_con_deuda) if emps_con_deuda else len(emps)
            act_e = round(info["total_pagar"] / n_div)
            nom_e = round(info["monto_nom"] / n_div)
            for rut, nombre, total_emp in emps:
                if total_emp == 0:
                    continue  # no registrar empleados sin deuda en el período
                filas.append({"rut":rut,"nombre":nombre,
                              "monto_nom": nom_e,
                              "monto_act": act_e,
                              "_fila_excel":0,
                              "origen":info["origen"],"adm":None,
                              "periodo":info["periodo"],"estado":info["estado"],"abogado":""})
    return filas, [], ws


# ── Parseo Excel Adobe general ────────────────────────────────────────────────

def _parsear_excel(wb, pdf_lookup: dict) -> tuple:
    ws    = wb["Table 1"] if "Table 1" in wb.sheetnames else wb.active
    filas = []
    advertencias = []
    grupo = {"origen":"","adm":"","periodo":"","estado":"","abogado":""}
    split_pendiente = None
    grupo_fila_pendiente = None
    filas_antes_grupo = 0
    uso_pdf = {}

    def v(col): return ws.cell(r, col).value

    def agregar_fila(rut, nombre, nom_raw, act_raw):
        nom = _convertir_monto(nom_raw)
        act = _convertir_monto(act_raw)
        periodo = grupo.get("periodo","")
        clave   = (rut, periodo)
        uso_idx = uso_pdf.get(clave, 0)
        candidatos = pdf_lookup.get(clave, [])
        if candidatos:
            lista = candidatos[uso_idx:] if uso_idx < len(candidatos) else candidatos
            nom_f, corr_nom = _corregir_con_pdf(nom, lista, "nom")
            act_f, corr_act = _corregir_con_pdf(act, lista, "act")
            uso_pdf[clave] = uso_idx + 1
            if corr_nom:
                advertencias.append(f"[CORREGIDO] {rut} {periodo}: monto_nom {nom}→{nom_f}")
                nom = nom_f
            if corr_act:
                advertencias.append(f"[CORREGIDO] {rut} {periodo}: monto_act {act}→{act_f}")
                act = act_f
        filas.append({"rut":rut,"nombre":nombre,"monto_nom":nom,"monto_act":act,
                      "_fila_excel":r,**grupo})

    def fnone(*vals):
        for x in vals:
            if _es_monto(x): return x
        return None

    # Detectar columnas desde encabezado
    COL_ORIGEN=None; COL_ADM=None; COL_PERIODO=None
    COL_NOM_GRP=None; COL_ACT_GRP=None; COL_ESTADO=None; COL_ABOGADO=None
    COL_NOM_EMP=[]; COL_ACT_EMP=[]

    for r in range(1, min(8, ws.max_row+1)):
        for c in range(1, ws.max_column+1):
            val = ws.cell(r,c).value
            if not isinstance(val,str): continue
            vl = val.strip().lower()
            if "origen" in vl and "deuda" in vl: COL_ORIGEN=c
            elif "adm" in vl and "origen" in vl: COL_ADM=c
            elif "periodo" in vl: COL_PERIODO=c
            elif "nominal" in vl:
                COL_NOM_GRP=c; COL_NOM_EMP.append(c)
                if c+1 not in COL_NOM_EMP: COL_NOM_EMP.append(c+1)
            elif "actualizada" in vl or "actualizado" in vl:
                COL_ACT_GRP=c; COL_ACT_EMP.append(c)
            elif ("estado" in vl and "deuda" in vl or "cobranza" in vl or "juridico" in vl) and not COL_ESTADO:
                COL_ESTADO=c
            elif "abogado" in vl: COL_ABOGADO=c
        if COL_ORIGEN and COL_PERIODO and COL_NOM_GRP: break

    COL_ORIGEN  = COL_ORIGEN  or 2
    COL_ADM     = COL_ADM     or 8
    COL_PERIODO = COL_PERIODO or 10
    COL_NOM_EMP = COL_NOM_EMP or [14,15,16,19,20]
    COL_ACT_EMP = COL_ACT_EMP or [17,18,19,20,21,22,24,25]
    COL_ESTADO  = COL_ESTADO  or 22
    COL_ABOGADO = COL_ABOGADO or 26

    if COL_NOM_GRP:
        for off in range(4):
            c = COL_NOM_GRP+off
            if c not in COL_NOM_EMP: COL_NOM_EMP.append(c)
    if COL_ACT_GRP:
        for off in range(5):
            c = COL_ACT_GRP+off
            if c not in COL_ACT_EMP: COL_ACT_EMP.append(c)

    for r in range(1, ws.max_row+1):
        a = v(1)
        origen_val  = v(COL_ORIGEN)
        periodo_val = v(COL_PERIODO)
        adm_val     = v(COL_ADM)
        estado_val  = v(COL_ESTADO)
        abogado_val = v(COL_ABOGADO)
        nom_cands   = [v(c) for c in COL_NOM_EMP]
        act_cands   = [v(c) for c in COL_ACT_EMP]

        if isinstance(origen_val,str) and _es_stop(origen_val): break
        if isinstance(a,str) and _es_stop(a): break
        if isinstance(a,str) and _es_ruido(a): split_pendiente=None; continue
        if isinstance(a,str) and len(a)>60 and not RUT_RE.match(str(a)): split_pendiente=None; continue

        if _es_grupo(origen_val):
            if grupo_fila_pendiente is not None and len(filas)==filas_antes_grupo:
                agregar_fila("","",grupo_fila_pendiente["nom"],grupo_fila_pendiente["act"])
            split_pendiente = None
            # Si la columna asignada no tiene período, escanear toda la fila
            periodo_raw = periodo_val
            if not periodo_raw:
                import datetime as _dt2
                for _cc in range(1, ws.max_column+1):
                    _cv = ws.cell(r, _cc).value
                    if isinstance(_cv, (_dt2.date, _dt2.datetime)):
                        periodo_raw = _cv; break
                    if isinstance(_cv, str) and re.match(r'^\d{2}/\d{4}$', _cv.strip()):
                        periodo_raw = _cv.strip(); break
            grupo = {
                "origen":  _norm_origen(origen_val),
                "adm":     adm_val,
                "periodo": _periodo(periodo_raw),
                "estado":  _norm_estado(str(estado_val).upper().strip()) if estado_val else "",
                "abogado": str(abogado_val).strip() if abogado_val else "",
            }
            nom_g = fnone(*[v(c) for c in COL_NOM_EMP]) if COL_NOM_EMP else None
            act_g = fnone(*[v(c) for c in COL_ACT_EMP]) if COL_ACT_EMP else None
            if nom_g is not None and act_g is not None:
                grupo_fila_pendiente = {"nom":nom_g,"act":act_g}
                filas_antes_grupo    = len(filas)
            else:
                grupo_fila_pendiente = None
            continue

        if not _es_grupo(origen_val) and isinstance(a,str) and _es_grupo(a) and not RUT_RE.match(a):
            split_pendiente = None
            periodo_g = ""; estado_g = ""
            # Primero buscar período incrustado en el texto de col 1 (ej: Provida comprime todo en col 1)
            _m_per = re.search(r'(\d{2}/\d{4})', a)
            if _m_per:
                periodo_g = _m_per.group(1)
            # También buscar estado en col 1
            _a_upper = a.upper()
            if not estado_g:
                for _kw in ["SIN GESTION","PREJUDICIAL","JUICIO","RESOLUCION","INGRESADA"]:
                    if _kw in _a_upper:
                        estado_g = _norm_estado(_kw); break
            for cc in range(1, ws.max_column+1):
                val_g = ws.cell(r,cc).value
                if isinstance(val_g,str):
                    if not periodo_g and re.match(r'^\d{2}/\d{4}$',val_g.strip()):
                        periodo_g = val_g.strip()
                    elif cc!=1 and val_g.strip():
                        upper_g = val_g.upper().strip()
                        if not estado_g and any(x in upper_g for x in ["SIN GESTION","PREJUDICIAL","JUICIO","RESOLUCION","INGRESADA"]):
                            estado_g = _norm_estado(upper_g)
            grupo = {"origen":_norm_origen(a),"adm":None,
                     "periodo":_periodo(periodo_g) if periodo_g else "",
                     "estado":estado_g,"abogado":""}
            continue

        if split_pendiente and a is None and (v(COL_ORIGEN) is None or not _es_grupo(v(COL_ORIGEN))):
            nombre_split = split_pendiente.get("nombre") or ""
            if not nombre_split:
                for cc in range(2,10):
                    val_n = ws.cell(r,cc).value
                    if isinstance(val_n,str) and val_n.strip() and not _es_monto(val_n):
                        nombre_split=val_n.strip(); break
            vals_num = [(c,ws.cell(r,c).value) for c in range(1,ws.max_column+1) if _es_monto(ws.cell(r,c).value)]
            if len(vals_num)>=2:
                agregar_fila(split_pendiente["rut"],nombre_split,vals_num[0][1],vals_num[1][1])
                split_pendiente=None; continue
            elif len(vals_num)==1:
                agregar_fila(split_pendiente["rut"],nombre_split,fnone(*nom_cands),vals_num[0][1])
                split_pendiente=None; continue

        rut_det=None; nombre_en_a=""
        if isinstance(a,str):
            if RUT_RE.match(a): rut_det=a
            else:
                m2 = RUT_CON_NOMBRE.match(a)
                if m2: rut_det=m2.group(1); nombre_en_a=m2.group(2).strip()

        if rut_det:
            nombre = nombre_en_a or ""
            if not nombre:
                for col_val in [v(c) for c in range(2,10)]:
                    if col_val and str(col_val).strip(): nombre=str(col_val).strip(); break
            nom_f = fnone(*nom_cands)
            act_f = fnone(*act_cands)
            if nom_f is not None and act_f is not None:
                split_pendiente=None
                agregar_fila(rut_det,nombre,nom_f,act_f)
            else:
                split_pendiente={"rut":rut_det,"nombre":nombre}
            continue

        split_pendiente=None

    if grupo_fila_pendiente is not None and len(filas)==filas_antes_grupo:
        agregar_fila("","",grupo_fila_pendiente["nom"],grupo_fila_pendiente["act"])

    return filas, advertencias, ws


# ── Escritura resultado ───────────────────────────────────────────────────────

def _agregar_al_resultado(wb_res, filas: list, rut_empresa: str,
                          razon_social: str, institucion: str):
    es_isapre = any(k in institucion.upper() for k in ("ISAPRE", "CRUZ BLANCA", "CONSALUD", "MASVIDA"))
    hoja = "Base Isapre" if es_isapre else "Base AFP"
    ws = wb_res[hoja]
    razon_social = razon_social.upper() if razon_social else ""
    rut_fmt = rut_empresa.replace(".", "")

    # Borrar solo filas de esta empresa + esta institución (no borrar otras AFPs)
    for r in range(ws.max_row, 1, -1):
        mismo_rut  = str(ws.cell(r,1).value or "").replace(".","") == rut_fmt
        misma_inst = str(ws.cell(r,6).value or "").upper() == institucion.upper()
        if mismo_rut and misma_inst:
            ws.delete_rows(r)

    primera_libre = ws.max_row + 1
    if primera_libre == 2 and ws.cell(2,1).value is None:
        primera_libre = 2

    for i, f in enumerate(filas, start=primera_libre):
        ws.cell(i, 1, rut_fmt)
        ws.cell(i, 2, razon_social)
        ws.cell(i, 3, f["rut"].replace(".","") if f.get("rut") else "")
        ws.cell(i, 4, f.get("nombre",""))
        ws.cell(i, 5, f.get("origen",""))
        ws.cell(i, 6, institucion)
        celda = ws.cell(i, 7, f.get("periodo",""))
        celda.number_format = "mmm-yy"
        ws.cell(i, 8, f.get("estado",""))
        c9 = ws.cell(i, 9, f.get("monto_nom",0))
        c9.number_format  = '"$"#,##0'
        c10 = ws.cell(i,10, f.get("monto_act",0))
        c10.number_format = '"$"#,##0'
        for col in range(11, 18):
            ws.cell(i, col, "")


# ── Extracción directa desde PDF (sin Excel Adobe) ───────────────────────────

def _extraer_filas_pdf(texto: str) -> list:
    """Extrae filas de deuda directamente del texto del PDF AFP."""
    grupos = []
    grupo_actual = None
    detalles_actuales = []

    for linea in texto.splitlines():
        linea = linea.strip()
        if not linea:
            continue

        m_g = RE_GRUPO_PDF.match(linea)
        if m_g:
            if grupo_actual:
                grupos.append((grupo_actual, detalles_actuales[:]))
                detalles_actuales = []
            grupo_actual = {
                "rut": "", "nombre": "",
                "monto_nom": _limpiar_num_pdf(m_g.group(3)),
                "monto_act": _limpiar_num_pdf(m_g.group(4)),
                "origen": _norm_origen(m_g.group(1)), "adm": None,
                "periodo": _periodo(m_g.group(2)),
                "estado": "", "abogado": "",
            }
            continue

        m_d = RE_DETALLE_PDF.match(linea)
        if m_d and grupo_actual:
            detalles_actuales.append({
                "rut": m_d.group(1), "nombre": "",
                "monto_nom": _limpiar_num_pdf(m_d.group(2)),
                "monto_act": _limpiar_num_pdf(m_d.group(3)),
                "origen": grupo_actual["origen"], "adm": None,
                "periodo": grupo_actual["periodo"],
                "estado": "", "abogado": "",
            })

    if grupo_actual:
        grupos.append((grupo_actual, detalles_actuales[:]))

    filas = []
    for grupo, detalles in grupos:
        if detalles:
            filas.extend(detalles)
        else:
            filas.append(grupo)
    return filas


def procesar_lote_solo_pdf(pdfs: list, log=None) -> bytes:
    """
    pdfs: lista de dicts {"pdf_bytes": bytes, "pdf_nombre": str}
    Extrae datos del PDF AFP sin necesidad de Excel Adobe.
    Devuelve bytes del Excel consolidado.
    """
    def _log(msg, tipo="info"):
        if log: log(msg, tipo)

    wb_res = openpyxl.load_workbook(TEMPLATE_PATH)

    for item in pdfs:
        pdf_bytes  = item["pdf_bytes"]
        pdf_nombre = item["pdf_nombre"]

        _log(f"── {pdf_nombre}", "info")

        # Detección NO DEUDA
        try:
            es_no_deuda, rut_nd, razon_nd = detectar_no_deuda(pdf_bytes)
        except Exception as e:
            _log(f"  Error leyendo PDF: {e}", "error")
            continue

        if es_no_deuda:
            _log(f"  Sin deuda — {razon_nd} ({rut_nd})", "ok")
            _agregar_al_resultado(wb_res, [{
                "rut": "", "nombre": "", "monto_nom": 0, "monto_act": 0,
                "origen": "", "adm": "", "periodo": "", "estado": "SIN DEUDA", "abogado": "",
            }], rut_nd, razon_nd, detectar_institucion(pdf_bytes, pdf_nombre))
            continue

        # Extraer texto completo
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            texto_completo = "\n".join((p.extract_text() or "") for p in pdf.pages)

        # Empresa
        m_rut = RE_RUT_EMPRESA.search(texto_completo)
        rut_empresa = re.sub(r"\s", "", m_rut.group(1)) if m_rut else ""
        m_rs = RE_RAZON_SOCIAL.search(texto_completo) or RE_RAZON_HABITAT.search(texto_completo)
        razon_social = _limpiar_razon(m_rs.group(1)) if m_rs else ""
        institucion  = detectar_institucion(pdf_bytes, pdf_nombre)
        _log(f"  {razon_social} ({rut_empresa}) — {institucion}", "info")

        # Extraer filas de deuda
        filas = _extraer_filas_pdf(texto_completo)

        if not filas:
            _log("  No se encontraron filas de deuda estructuradas en el PDF", "warn")
            _log("  Prueba subiendo también el Excel de Adobe para este archivo", "warn")
            continue

        _log(f"  {len(filas)} filas extraídas", "info")
        _agregar_al_resultado(wb_res, filas, rut_empresa, razon_social, institucion)
        _log(f"  Listo: {len(filas)} filas agregadas", "ok")

    buf = io.BytesIO()
    wb_res.save(buf)
    buf.seek(0)
    return buf.read()


# ── API pública ───────────────────────────────────────────────────────────────

def procesar_lote(pares: list, log=None) -> bytes:
    """
    pares: lista de dicts {"pdf_bytes": bytes, "pdf_nombre": str,
                           "excel_bytes": bytes, "excel_nombre": str}
    Devuelve bytes del Excel consolidado.
    """
    def _log(msg, tipo="info"):
        if log: log(msg, tipo)

    wb_res = openpyxl.load_workbook(TEMPLATE_PATH)

    for par in pares:
        pdf_bytes    = par.get("pdf_bytes") or b""
        pdf_nombre   = par.get("pdf_nombre", "")
        excel_bytes  = par["excel_bytes"]
        excel_nombre = par["excel_nombre"]
        tiene_pdf    = len(pdf_bytes) > 256  # bytes vacíos o casi vacíos = sin PDF

        _log(f"── {excel_nombre}", "info")

        # Abrir Excel y procesar CADA HOJA como un certificado independiente
        wb_adobe = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        hojas = wb_adobe.sheetnames
        _log(f"  {len(hojas)} hoja(s) encontradas en el Excel", "info")

        for nombre_hoja in hojas:
            ws = wb_adobe[nombre_hoja]

            # Texto de las primeras filas para detectar no-deuda y empresa
            _txt = " ".join(
                str(ws.cell(r, c).value or "")
                for r in range(1, min(8, ws.max_row + 1))
                for c in range(1, ws.max_column + 1)
            )
            _txt_norm = re.sub(r'\s+', ' ', _txt).lower()

            # Detección NO DEUDA
            es_no_deuda = False
            rut_nd = razon_nd = ""
            if RE_NO_DEUDA.search(_txt) or "no registra" in _txt_norm:
                rut_nd   = _extraer_rut(_txt)
                razon_nd = _extraer_razon(_txt)
                es_no_deuda = True

            # Institución: busca en nombre de hoja, texto de celdas y nombre del archivo Excel
            inst = _detectar_inst_texto(nombre_hoja, _txt, excel_nombre)

            if es_no_deuda:
                _log(f"  [{nombre_hoja}] Sin deuda — {razon_nd} ({rut_nd})", "ok")
                _agregar_al_resultado(wb_res, [{
                    "rut": "", "nombre": "", "monto_nom": 0, "monto_act": 0,
                    "origen": "", "adm": "", "periodo": "", "estado": "SIN DEUDA", "abogado": "",
                }], rut_nd, razon_nd, inst)
                continue

            # Extraer empresa y filas de deuda
            rut_empresa, razon_social = extraer_datos_empresa(ws)
            if not rut_empresa and not razon_social:
                _log(f"  [{nombre_hoja}] Sin datos de empresa — omitida", "warn")
                continue

            _log(f"  [{nombre_hoja}] {razon_social} ({rut_empresa}) — {inst}", "info")

            # Apuntar el workbook a esta hoja para que los parsers la lean
            wb_adobe.active = ws
            if _es_habitat(ws):
                filas, advertencias, ws_adobe2 = _parsear_habitat(wb_adobe, {})
            else:
                filas, advertencias, ws_adobe2 = _parsear_excel(wb_adobe, {})

            if not filas:
                # Tabla vacía con empresa reconocida = certificado sin deuda
                _log(f"  [{nombre_hoja}] Sin filas de deuda — marcando SIN DEUDA", "ok")
                _agregar_al_resultado(wb_res, [{
                    "rut": "", "nombre": "", "monto_nom": 0, "monto_act": 0,
                    "origen": "", "adm": "", "periodo": "", "estado": "SIN DEUDA", "abogado": "",
                }], rut_empresa, razon_social, inst)
                continue

            _log(f"  [{nombre_hoja}] {len(filas)} filas extraídas", "info")
            for a in advertencias:
                _log(f"  {a}", "warn")

            _agregar_al_resultado(wb_res, filas, rut_empresa, razon_social, inst)
            _log(f"  [{nombre_hoja}] Listo: {len(filas)} filas agregadas", "ok")

        wb_adobe.close()

    # Post-proceso: para cada RUT usar el nombre más largo disponible en ambas hojas
    _normalizar_nombres(wb_res)

    buf = io.BytesIO()
    wb_res.save(buf)
    buf.seek(0)
    return buf.read()


def _normalizar_nombres(wb_res):
    """Para cada RUT, propaga el nombre más largo a todas sus filas en ambas hojas."""
    mejor = {}  # rut_fmt → nombre más largo
    hojas = [h for h in ["Base AFP", "Base Isapre"] if h in wb_res.sheetnames]
    for hoja in hojas:
        ws = wb_res[hoja]
        for r in range(2, ws.max_row + 1):
            rut = str(ws.cell(r, 1).value or "").strip()
            nombre = str(ws.cell(r, 2).value or "").strip()
            if rut and len(nombre) > len(mejor.get(rut, "")):
                mejor[rut] = nombre
    for hoja in hojas:
        ws = wb_res[hoja]
        for r in range(2, ws.max_row + 1):
            rut = str(ws.cell(r, 1).value or "").strip()
            if rut and mejor.get(rut):
                ws.cell(r, 2, mejor[rut])
