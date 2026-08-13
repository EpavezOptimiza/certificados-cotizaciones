"""Conexión a BASE MADRE.xlsx (SharePoint) — lectura automática de la hoja CLIENTES.

El enlace compartido del Excel se configura en la variable de entorno
BASE_MADRE_URL (NO va en el código: el repositorio es público).
Acepta el share-link tal cual lo entrega SharePoint y lo convierte
internamente al formato de descarga directa.
"""

import io
import os
import time
import threading
import http.cookiejar
import urllib.request

import openpyxl

# Cache en memoria: se refresca solo si pasaron REFRESCO_SEG segundos
REFRESCO_SEG = 600  # 10 minutos
_CACHE = {"filas": None, "columnas": None, "ts": 0, "error": None}
_LOCK = threading.Lock()


def url_guardada():
    """Enlace del Excel: variable de entorno o, si no existe, la base de datos
    (tabla app_config — se pega desde la página /base_madre)."""
    url = os.environ.get("BASE_MADRE_URL", "").strip()
    if url:
        return url
    try:
        from database import get_conn
        with get_conn() as conn:
            row = conn.execute(
                "SELECT valor FROM app_config WHERE clave='base_madre_url'").fetchone()
            return (row["valor"] or "").strip() if row else ""
    except Exception:
        return ""


def guardar_url(url):
    """Guarda el enlace en la base de datos y limpia el cache."""
    from database import get_conn
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO app_config(clave, valor) VALUES('base_madre_url', ?) "
            "ON CONFLICT(clave) DO UPDATE SET valor=excluded.valor", (url.strip(),))
    with _LOCK:
        _CACHE.update({"filas": None, "columnas": None, "ts": 0, "error": None})


def _url_descarga():
    """Convierte el share-link de SharePoint al endpoint de descarga directa.

    https://<tenant>-my.sharepoint.com/:x:/g/personal/<usuario>/<TOKEN>?e=...
      → https://<tenant>-my.sharepoint.com/personal/<usuario>/_layouts/15/download.aspx?share=<TOKEN>
    """
    url = url_guardada()
    if not url:
        return None
    if "download.aspx" in url:
        return url
    try:
        if "/:x:/g/personal/" in url:
            dominio = url.split("/:x:/")[0]
            resto = url.split("/:x:/g/personal/")[1]
            usuario, token = resto.split("/", 1)
            token = token.split("?")[0]
            return f"{dominio}/personal/{usuario}/_layouts/15/download.aspx?share={token}"
    except Exception:
        pass
    return url


def _descargar():
    url = _url_descarga()
    if not url:
        raise Exception("Falta pegar el enlace del Excel (usa el recuadro de configuración)")
    # SharePoint exige conservar cookies entre las redirecciones del enlace anónimo
    cj = http.cookiejar.CookieJar()
    op = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    op.addheaders = [("User-Agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64)")]
    with op.open(url, timeout=90) as r:
        data = r.read()
    if data[:2] != b"PK":
        raise Exception("SharePoint no devolvió el Excel — revisa que el enlace siga "
                        "vigente y compartido como 'Cualquier persona puede ver'")
    return data


def _parsear(data):
    """Lee la hoja CLIENTES (o la primera) → (columnas, filas como dicts)."""
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        # Preferir un nombre EXACTO ("Clientes") antes que una coincidencia
        # parcial: el Excel real agrego una hoja "RESUMEN DE CLIENTES" (una
        # tabla dinamica, no el listado de empresas) que tambien contiene
        # la palabra "cliente" y aparece ANTES que "CLIENTES" en el libro,
        # asi que una busqueda de solo substring la agarraba por error
        # (confirmado: hacia que devolviera 36 filas basura en vez de las
        # filas reales, y que TODAS las empresas del analisis DICOM
        # cayeran en "SIN GRUPO").
        ws = None
        for nombre in wb.sheetnames:
            if nombre.strip().lower() == "clientes":
                ws = wb[nombre]
                break
        if ws is None:
            for nombre in wb.sheetnames:
                if "cliente" in nombre.lower():
                    ws = wb[nombre]
                    break
        if ws is None:
            ws = wb.active
        try:
            ws.reset_dimensions()
        except Exception:
            pass

        import datetime as _dt

        def _fmt(v):
            if v is None:
                return ""
            if isinstance(v, (_dt.datetime, _dt.date)):
                return v.strftime("%d/%m/%Y")
            if isinstance(v, float) and v.is_integer():
                return str(int(v))
            return str(v).strip()

        gen = ws.iter_rows(values_only=True)
        headers = []
        for row in gen:
            headers = [str(h).strip() if h else "" for h in row]
            break
        cols = [i for i, h in enumerate(headers) if h]

        filas = []
        for row in gen:
            if not any(v is not None and str(v).strip() for v in row):
                continue
            d = {headers[i]: (_fmt(row[i]) if i < len(row) else "") for i in cols}
            if any(d.values()):
                filas.append(d)
        return [headers[i] for i in cols], filas
    finally:
        wb.close()


_ESTATUS_PERMITIDOS = ("vigente", "estudio inicial")


def _norm_estatus(s):
    import unicodedata
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return s.strip().lower()


def filtrar_permitidos(columnas, filas):
    """Devuelve sólo las filas con Estatus cliente VIGENTE o ESTUDIO INICIAL.
    Usado por la página BASE MADRE para mostrar únicamente esos estatus."""
    if not filas:
        return filas

    def col(*terms):
        for c in (columnas or []):
            cn = _norm_estatus(c)
            if all(t in cn for t in terms):
                return c
        return None

    c = col("estatus", "cliente") or col("estatus")
    if not c:
        return filas

    def ok(v):
        n = _norm_estatus(v)
        return any(n == p or n.startswith(p) for p in _ESTATUS_PERMITIDOS)

    return [f for f in filas if ok(f.get(c))]


def obtener_clientes(forzar=False):
    """Devuelve (columnas, filas, ts_ultima_lectura, error).

    Sirve desde cache si la última lectura tiene menos de REFRESCO_SEG.
    Si la descarga falla pero hay datos previos en cache, sigue sirviendo
    los datos antiguos y reporta el error.
    """
    with _LOCK:
        fresco = _CACHE["filas"] is not None and (time.time() - _CACHE["ts"]) < REFRESCO_SEG
        if fresco and not forzar:
            return _CACHE["columnas"], _CACHE["filas"], _CACHE["ts"], _CACHE["error"]
        try:
            data = _descargar()
            columnas, filas = _parsear(data)
            _CACHE.update({"filas": filas, "columnas": columnas,
                           "ts": time.time(), "error": None})
        except Exception as e:
            _CACHE["error"] = str(e)
        return _CACHE["columnas"], _CACHE["filas"], _CACHE["ts"], _CACHE["error"]


def obtener_datos_dicom():
    """Obtiene datos para DICOM: RUT vigentes o en estudio inicial,
    agrupados por GRUPO, con su consultor de deuda asignado."""
    columnas, filas, _, error = obtener_clientes()
    if error or not filas:
        return {"error": error or "No hay datos", "grupos": {}, "todos_ruts": [],
                "razones_sociales": [], "consultores_deuda": {}}

    def encontrar_col(terms):
        for col in columnas:
            cn = _norm_estatus(col)
            if all(t in cn for t in terms):
                return col
        return None

    col_rut = encontrar_col(["rut"])
    col_grupo = encontrar_col(["grupo"])
    col_razon = encontrar_col(["razon", "social"])
    col_estatus = encontrar_col(["estatus", "cliente"])
    col_consultor_deuda = encontrar_col(["consultor", "deuda"])

    grupos = {}
    todos_ruts = set()
    todas_razones = set()
    consultores_deuda = {}

    for fila in filas:
        if col_estatus:
            estatus = _norm_estatus(fila.get(col_estatus, ""))
            if not any(estatus == p or estatus.startswith(p) for p in _ESTATUS_PERMITIDOS):
                continue

        rut = (fila.get(col_rut) or "").strip()
        grupo = (fila.get(col_grupo) or "").strip()
        razon = (fila.get(col_razon) or "").strip()

        if not rut:
            continue

        todos_ruts.add(rut)
        if razon:
            todas_razones.add(razon)

        rut_norm = rut.replace(".", "").replace(" ", "")
        consultores_deuda[rut_norm] = (fila.get(col_consultor_deuda) or "").strip() if col_consultor_deuda else ""

        if grupo:
            if grupo not in grupos:
                grupos[grupo] = {"razones_sociales": set(), "ruts": set()}
            grupos[grupo]["ruts"].add(rut)
            if razon:
                grupos[grupo]["razones_sociales"].add(razon)

    resultado = {
        "grupos": {
            g: {"razones_sociales": sorted(list(d["razones_sociales"])),
                "ruts": sorted(list(d["ruts"]))}
            for g, d in grupos.items()
        },
        "todos_ruts": sorted(list(todos_ruts)),
        "razones_sociales": sorted(list(todas_razones)),
        "consultores_deuda": consultores_deuda,
    }

    return resultado
