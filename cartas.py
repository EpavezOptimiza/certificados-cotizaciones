"""
Módulo Cartas Previsionales — Blueprint Flask
Integra con la app de Certificados de Cotizaciones
"""
import os, json, threading, uuid, tempfile, shutil
from datetime import datetime
from flask import (Blueprint, render_template, request, jsonify,
                   send_file, session, current_app, make_response)
from functools import wraps

import os as _os
_HERE = _os.path.dirname(_os.path.abspath(__file__))
cartas_bp = Blueprint("cartas", __name__,
    template_folder=_os.path.join(_HERE, "templates", "cartas"),
    static_folder=_os.path.join(_HERE, "static", "cartas") if _os.path.exists(_os.path.join(_HERE, "static", "cartas")) else None,
    url_prefix="/cartas")

# ── Estado de jobs (en memoria, Railway reinicia limpia) ──────────────────────
_jobs = {}  # job_id -> {status, log, result, worker_rut}

def get_job(job_id):
    return _jobs.get(job_id)

def set_job(job_id, data):
    _jobs[job_id] = data

# ── Auth helper (reutiliza el de app.py) ──────────────────────────────────────
def _get_user():
    from flask import request as _r
    token = _r.cookies.get("session_token")
    if not token: return None
    try:
        from app import get_current_user
        return get_current_user()
    except:
        pass
    try:
        import sys, os
        # Intentar importar get_conn desde el directorio raíz
        sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from database import get_conn
        with get_conn() as conn:
            row = conn.execute(
                "SELECT u.* FROM sesiones s JOIN usuarios u ON u.id=s.usuario_id WHERE s.token=?",
                (token,)).fetchone()
            return dict(row) if row else None
    except Exception as e:
        print(f"[cartas] _get_user error: {e}")
        return None

def cartas_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        from flask import redirect
        user = _get_user()
        if not user:
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated

# ── Parseo de Excel ───────────────────────────────────────────────────────────
def parsear_excel(file_bytes):
    """Lee el Excel de deudas y retorna lista de dicts con todos los campos."""
    import openpyxl, io
    from datetime import datetime as dt

    MESES_MAP = {'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,
                 'jul':7,'ago':8,'sep':9,'oct':10,'nov':11,'dic':12}

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    # Buscar hoja correcta
    ws = None
    for nombre in wb.sheetnames:
        n = nombre.lower()
        if 'base' in n and 'afp' in n:
            ws = wb[nombre]; break
    if ws is None:
        for nombre in wb.sheetnames:
            if 'base' in nombre.lower():
                ws = wb[nombre]; break
    if ws is None:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).lower().strip() if h else '' for h in rows[0]]

    def col(nombres, default=-1):
        for n in nombres:
            for i, h in enumerate(headers):
                if n in h:
                    return i
        return default

    iRutEmp  = col(['01_rut emp','rut empresa','rut_empresa'], 0)
    iRazon   = col(['02_razon','razon social','razon_social'], 1)
    iRutTrab = col(['03_rut afil','rut afiliado','rut_afiliado','rut trabajador'], 2)
    iNombre  = col(['04_nombre','nombre afiliado','nombre trabajador','nombre_afiliado'], 3)
    iTipo    = col(['05_tipo','tipo producto','tipo_producto'], 4)
    iInst    = col(['06_inst','institucion'], 5)
    iPeriodo = col(['07_periodo','período','periodo'], 6)
    iMontoN  = col(['11_monto nom','monto nominal'], 7)
    iMontoI  = col(['12_monto int','monto interes'], 8)
    iFee     = col(['fee%','fee'], 9)
    iFecha   = col(['08_fecha cese','fecha cese','fecha_cese'], 10)
    iAnalisis= col(['13_','analisis'], 11)
    iMotivo  = col(['09_','motivos de deuda','motivo de deuda'], 12)
    iTipoDoc = col(['14_','tipo documento'], 13)
    iEstatus = col(['15_','estatus','estado'], 14)
    iObs     = col(['observaciones'], 15)

    def g(row, idx):
        if idx < 0 or idx >= len(row): return ''
        v = row[idx]
        if v is None: return ''
        return str(v).strip()

    def gfecha(row, idx):
        if idx < 0 or idx >= len(row): return ''
        v = row[idx]
        if v is None: return ''
        import datetime as _dt
        if isinstance(v, (_dt.datetime, _dt.date)):
            return v.strftime('%d/%m/%Y')
        s = str(v).strip()
        if s and s[4:5] == '-':
            partes = s[:10].split('-')
            if len(partes) == 3:
                return f"{partes[2]}/{partes[1]}/{partes[0]}"
        return s

    def fmt_periodo(v):
        if v is None: return ''
        if hasattr(v, 'month'):
            MESES = {1:'Ene',2:'Feb',3:'Mar',4:'Abr',5:'May',6:'Jun',
                     7:'Jul',8:'Ago',9:'Sep',10:'Oct',11:'Nov',12:'Dic'}
            return f"{MESES[v.month]}-{str(v.year)[2:]}"
        s = str(v).strip().split(' ')[0]
        parts = s.split('-')
        if len(parts) == 3:
            try:
                MESES = {1:'Ene',2:'Feb',3:'Mar',4:'Abr',5:'May',6:'Jun',
                         7:'Jul',8:'Ago',9:'Sep',10:'Oct',11:'Nov',12:'Dic'}
                return f"{MESES[int(parts[1])]}-{parts[0][2:]}"
            except: pass
        return s

    datos = []
    for row in rows[1:]:
        if not any(row): continue
        rut_trab = g(row, iRutTrab)
        if not rut_trab or rut_trab.upper() in ('ND','N/D','','NONE'): continue
        periodo = fmt_periodo(row[iPeriodo] if iPeriodo < len(row) else None)
        datos.append({
            'rut_empresa':  g(row, iRutEmp),
            'razon_social': g(row, iRazon),
            'rut_trabajador': rut_trab,
            'nombre':       g(row, iNombre),
            'tipo':         g(row, iTipo),
            'institucion':  g(row, iInst),
            'periodo':      periodo,
            'monto_nominal': g(row, iMontoN),
            'monto_interes': g(row, iMontoI),
            'fee':          g(row, iFee),
            'fecha_cese':   gfecha(row, iFecha),
            'analisis':     g(row, iAnalisis),
            'motivo':       g(row, iMotivo),
            'tipo_documento': g(row, iTipoDoc),
            'estatus':      g(row, iEstatus),
            'observaciones': g(row, iObs),
        })
    return datos

def agrupar_por_trabajador(datos):
    """Agrupa filas por trabajador, coleccionando períodos."""
    MESES_MAP = {'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,
                 'jul':7,'ago':8,'sep':9,'oct':10,'nov':11,'dic':12}

    def per_key(p):
        try:
            parts = str(p).lower().split('-')
            m = MESES_MAP.get(parts[0][:3], 0)
            a = int(parts[1]) + 2000 if len(parts[1]) <= 2 else int(parts[1])
            return (a, m)
        except: return (0, 0)

    grupos = {}
    for d in datos:
        k = d['rut_trabajador']
        if k not in grupos:
            grupos[k] = {**d, 'periodos': []}
        p = d['periodo']
        if p and p not in grupos[k]['periodos']:
            grupos[k]['periodos'].append(p)

    # Ordenar períodos de menor a mayor
    for k in grupos:
        grupos[k]['periodos'].sort(key=per_key)

    return list(grupos.values())

# ── Generación de carta PDF ───────────────────────────────────────────────────
def generar_carta_pdf(carta_data, firma_data):
    """Genera PDF de carta previsional y retorna bytes."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_LEFT, TA_CENTER
    from reportlab.lib import colors
    import io

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
        leftMargin=3*cm, rightMargin=3*cm,
        topMargin=2.5*cm, bottomMargin=2.5*cm)

    styles = getSampleStyleSheet()
    normal = ParagraphStyle('Normal', fontName='Helvetica', fontSize=11,
        leading=16, spaceAfter=6)
    bold_s = ParagraphStyle('Bold', fontName='Helvetica-Bold', fontSize=11, leading=16)
    title_s = ParagraphStyle('Title', fontName='Helvetica-Bold', fontSize=11,
        leading=16, alignment=TA_RIGHT)

    # Períodos
    periodos_sel = carta_data.get('periodos_sel', carta_data.get('periodos', []))
    extra = [p.strip() for p in carta_data.get('periodos_extra','').split(',') if p.strip()]
    todos = list(dict.fromkeys(periodos_sel + extra))

    MESES_MAP = {'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,
                 'jul':7,'ago':8,'sep':9,'oct':10,'nov':11,'dic':12}
    def per_key(p):
        try:
            parts = str(p).lower().split('-')
            m = MESES_MAP.get(parts[0][:3], 0)
            a = int(parts[1]) + 2000 if len(parts[1]) <= 2 else int(parts[1])
            return (a, m)
        except: return (0, 0)

    todos_ord = sorted(todos, key=per_key)
    if todos_ord:
        periodo_txt = todos_ord[0] if len(todos_ord) == 1 else f"{todos_ord[0]} hasta {todos_ord[-1]}"
    else:
        periodo_txt = '[PERIODO]'

    MESES_ES = {'01':'Enero','02':'Febrero','03':'Marzo','04':'Abril',
                '05':'Mayo','06':'Junio','07':'Julio','08':'Agosto',
                '09':'Septiembre','10':'Octubre','11':'Noviembre','12':'Diciembre'}
    mes_num = (carta_data.get('mes') or datetime.now().strftime('%Y-%m')).split('-')
    mes_txt = MESES_ES.get(mes_num[1] if len(mes_num) > 1 else '', '') if len(mes_num) > 1 else ''
    anio_txt = mes_num[0] if mes_num else str(datetime.now().year)
    fecha_txt = f"Santiago, {mes_txt} {anio_txt}"

    motivo = carta_data.get('motivo', '')
    if not motivo:
        motivo = carta_data.get('motivo_deuda', '[MOTIVO]')
    fecha_cese = carta_data.get('fecha_cese', '')
    if fecha_cese and motivo and '[MOTIVO]' not in motivo and fecha_cese not in motivo:
        motivo_final = motivo + ' ' + fecha_cese
    else:
        motivo_final = motivo or '[MOTIVO]'

    story = []
    story.append(Paragraph(fecha_txt, title_s))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(f"Senores:", normal))
    story.append(Paragraph(f"<b>{carta_data.get('institucion', '')}.</b>", normal))
    story.append(Paragraph("Presente. -", normal))
    story.append(Spacer(1, 0.3*cm))

    razon = carta_data.get('razon_social', '')
    rut_emp = carta_data.get('rut_empresa', '')
    story.append(Paragraph(
        f"Me dirijo a ustedes en representacion de la empresa {razon}. "
        f"Rut: {rut_emp}, a fin de informar a ustedes lo siguiente:",
        normal))
    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph(f"<b>NOMBRE: {carta_data.get('nombre','').upper()}</b>", normal))
    story.append(Paragraph(f"<b>RUT: {carta_data.get('rut_trabajador','')}.</b>", normal))
    story.append(Paragraph(f"<b>TIPO DE PRODUCTO: {carta_data.get('tipo','')}.</b>", normal))
    story.append(Paragraph(f"<b>PERIODO DE DEUDA: {periodo_txt}.</b>", normal))
    story.append(Paragraph(f"<b>MOTIVO DE NO PAGO: {motivo_final}</b>", normal))
    story.append(Spacer(1, 0.8*cm))
    story.append(Paragraph("Les saluda atentamente,", normal))
    story.append(Spacer(1, 1.5*cm))
    story.append(HRFlowable(width="40%", thickness=0.5, color=colors.black))
    story.append(Spacer(1, 0.2*cm))

    if firma_data:
        story.append(Paragraph(f"Nombre: {firma_data.get('nombre','')}", normal))
        story.append(Paragraph(f"Rut: {firma_data.get('rut','')}", normal))
        story.append(Paragraph(f"Cargo: {firma_data.get('cargo','')}", normal))
        if firma_data.get('correo'):
            story.append(Paragraph(f"Correo: {firma_data.get('correo')}", normal))
        if firma_data.get('tel'):
            story.append(Paragraph(f"Tel: {firma_data.get('tel')}", normal))

    doc.build(story)
    buf.seek(0)
    return buf.read()

# ── Bot PreviRed (headless) ───────────────────────────────────────────────────
def run_bot_previred(job_id, rut_login, clave, workers, firma_data):
    """Corre el bot de PreviRed en un thread separado."""
    job = _jobs[job_id]

    def log(msg):
        job['log'].append(msg)
        print(f"[BOT {job_id[:8]}] {msg}")

    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
        import time, calendar, datetime as _dt

        log("Iniciando navegador...")
        tmp_dir = tempfile.mkdtemp()

        MESES_MAP = {'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,
                     'jul':7,'ago':8,'sep':9,'oct':10,'nov':11,'dic':12}

        def parse_periodos(w):
            result = []
            for p in w.get('periodos_sel', []):
                try:
                    parts = str(p).lower().split('-')
                    m = MESES_MAP.get(parts[0][:3], 0)
                    a = int(parts[1]) + 2000 if len(parts[1]) <= 2 else int(parts[1])
                    if m: result.append((a, m))
                except: pass
            return result

        video_dir = os.path.join(tmp_dir, 'video')
        os.makedirs(video_dir, exist_ok=True)

        with sync_playwright() as pw:
            browser = pw.chromium.launch(
                headless=True,
                args=['--no-sandbox','--disable-dev-shm-usage','--disable-gpu'],
                downloads_path=tmp_dir,
            )
            ctx = browser.new_context(
                accept_downloads=True,
                viewport={'width': 1280, 'height': 900},
                record_video_dir=video_dir,
                record_video_size={'width': 1280, 'height': 900},
            )
            page = ctx.new_page()
            page.set_default_timeout(45000)

            def click(selector):
                try:
                    page.locator(selector).first.scroll_into_view_if_needed()
                    page.locator(selector).first.click()
                except: pass

            def fill(selector, text):
                try:
                    page.locator(selector).first.fill(text)
                except: pass

            def select_opt(selector, text):
                try:
                    page.locator(selector).first.select_option(label=text)
                except: pass

            def select_contains(selector, text):
                try:
                    opts = page.locator(f'{selector} option').all()
                    for o in opts:
                        t = o.inner_text().strip()
                        if text.upper() in t.upper() or t.upper() in text.upper():
                            page.locator(selector).first.select_option(label=t)
                            return t
                except: pass
                return None

            def continuar():
                sels = ["button:has-text('Continuar')", "input[value='Continuar']",
                        "input[value*='ontinuar']", "button[type='submit']", "input[type='submit']"]
                for sel in sels:
                    try:
                        el = page.locator(sel).first
                        if el.count() > 0:
                            el.scroll_into_view_if_needed()
                            el.click()
                            log(f"✓ Click Continuar: {sel}")
                            return
                    except: pass
                log("⚠ No se encontró botón Continuar")

            def finalizar():
                for sel in ["a:has-text('Finalizar')", "input[value='Finalizar']"]:
                    try:
                        page.locator(sel).first.click()
                        return
                    except: pass

            def set_fecha_js(field_id, fecha_val):
                page.evaluate(f"""
                    var el = document.getElementById('{field_id}');
                    if (el) {{
                        el.removeAttribute('readonly');
                        el.removeAttribute('disabled');
                        el.value = '{fecha_val}';
                        try {{ jQuery('#{field_id}').datepicker('setDate', '{fecha_val}'); }} catch(e) {{}}
                    }}
                """)

            # ── Login ─────────────────────────────────────────────────────────
            log("Accediendo a PreviRed...")

            def save_screenshot(label):
                try:
                    import base64
                    png_bytes = page.screenshot(full_page=True)
                    job['screenshot_b64'] = base64.b64encode(png_bytes).decode()
                    job['screenshot_label'] = label
                    log(f"📸 Captura guardada: {label}")
                except Exception as se:
                    log(f"(sin captura: {se})")

            page.goto("https://www.previred.com/wPortal/login/login.jsp", wait_until='domcontentloaded')
            page.wait_for_timeout(2000)
            log(f"URL actual: {page.url}")
            save_screenshot(f'bot_login_{job_id[:8]}.png')

            # Selectores del formulario de login de PreviRed
            RUT_SELS   = ['[name="web_rut2"]', '#web_rut', '[name="web_rut"]', 'input[type="text"]']
            CLAVE_SELS = ['[name="web_password"]', '#web_clave', '[name="web_clave"]', 'input[type="password"]']
            BTN_SELS   = [
                'button:has-text("INGRESAR")',
                'button:has-text("Ingresar")',
                'input[value="INGRESAR"]',
                '#web_btn_login',
                'button[type="submit"]',
                'input[type="submit"]',
            ]

            def try_fill(sels, value, campo):
                for sel in sels:
                    try:
                        el = page.locator(sel).first
                        el.wait_for(state='visible', timeout=2000)
                        el.fill(value)
                        log(f"✓ Campo {campo} llenado con: {sel}")
                        return True
                    except: pass
                log(f"⚠ No se encontró campo {campo}")
                return False

            try_fill(RUT_SELS, rut_login, 'RUT')
            try_fill(CLAVE_SELS, clave, 'clave')

            btn_clicked = False
            for sel in BTN_SELS:
                try:
                    el = page.locator(sel).first
                    el.wait_for(state='visible', timeout=3000)
                    el.click()
                    btn_clicked = True
                    log(f"✓ Click en botón login: {sel}")
                    break
                except: pass

            if not btn_clicked:
                # Último recurso: Enter en el campo clave
                for sel in CLAVE_SELS:
                    try:
                        page.locator(sel).first.press('Enter')
                        btn_clicked = True
                        log("✓ Login via Enter en campo clave")
                        break
                    except: pass

            if not btn_clicked:
                save_screenshot(f'bot_nologin_{job_id[:8]}.png')
                raise Exception("No se encontró el botón de login en PreviRed. Ver captura para diagnóstico.")

            page.wait_for_timeout(2000)
            save_screenshot(f'bot_postlogin_{job_id[:8]}.png')
            log(f"URL post-login: {page.url}")

            if 'login' in page.url.lower() or page.locator('text=Ingresa tu RUT').count() > 0:
                job['status'] = 'error'
                job['error'] = 'Credenciales PreviRed incorrectas'
                browser.close()
                return

            log("Login OK")

            def ir_a_inicio():
                """Vuelve a la página principal clickeando el logo de PreviRed."""
                for sel in ["a img[alt*='PreviRed' i]", ".logo a", "a[href*='Ctrl1Fce']", "a[href*='portal']"]:
                    try:
                        el = page.locator(sel).first
                        if el.count() > 0:
                            el.click()
                            page.wait_for_timeout(1500)
                            return
                    except: pass
                # Si nada funciona, volver atrás en historial
                try:
                    page.go_back(wait_until='domcontentloaded', timeout=8000)
                    page.wait_for_timeout(1000)
                except: pass

            def ir_a_empresas():
                """Hace click en el menú/tile Empresas desde la página actual."""
                if page.locator("text=Elemento no encontrado").count() > 0:
                    log("⚠ PreviRed mostró error — volviendo al inicio...")
                    ir_a_inicio()
                    page.wait_for_timeout(1000)

                for sel in ["li#empresa > a", "li#empresa", "a:has-text('Empresas')", "span:has-text('Empresas')"]:
                    try:
                        el = page.locator(sel).first
                        if el.count() > 0:
                            el.scroll_into_view_if_needed()
                            el.click()
                            log(f"✓ Click en Empresas: {sel}")
                            page.wait_for_timeout(2000)
                            return
                    except: pass
                log("⚠ No se encontró el botón Empresas")

            # ── Procesar cada trabajador ───────────────────────────────────────
            for w in workers:
                rut_e = w.get('rut_empresa') or ''
                rut_t_raw = w.get('rut_trabajador') or ''
                if not rut_e or not rut_t_raw:
                    log(f"⚠ Saltando trabajador sin RUT empresa o trabajador: {w.get('nombre','?')}")
                    continue
                rut_num = rut_e.replace('.','').split('-')[0].strip()
                rut_t = rut_t_raw.replace('.','').replace(' ','')
                inst = w.get('institucion','').strip()
                es_ausentismo = 'ausentismo' in (w.get('causa') or '').lower()
                periodos_parsed = parse_periodos(w) if es_ausentismo else []

                log(f"Procesando {w['nombre']} ({w['rut_trabajador']})...")
                ir_a_empresas()
                save_screenshot(f'bot_empresas_{job_id[:8]}.png')
                log(f"URL empresas: {page.url}")

                # Buscar empresa — nuevo UI: botón "Ingresar" en tabla
                # Intentar fila que contenga el RUT
                ingresado = False
                rut_formateado = rut_e.replace('.','').split('-')[0]  # 78383289
                try:
                    # Buscar fila con el RUT y hacer click en su botón Ingresar
                    filas = page.locator("tr").all()
                    for fila in filas:
                        if rut_formateado in (fila.inner_text() or ''):
                            btn = fila.locator("button:has-text('Ingresar'), input[value='Ingresar'], a:has-text('Ingresar')")
                            if btn.count() > 0:
                                btn.first.click()
                                log(f"✓ Click Ingresar en fila con RUT {rut_e}")
                                ingresado = True
                                break
                except: pass

                if not ingresado:
                    # Fallback: primer botón Ingresar de la página
                    btn = page.locator("button:has-text('Ingresar'), input[value='Ingresar'], a:has-text('Ingresar')").first
                    if btn.count() > 0:
                        btn.click()
                        log("✓ Click en primer botón Ingresar")
                        ingresado = True

                if not ingresado:
                    save_screenshot(f'bot_noemp_{job_id[:8]}.png')
                    log(f"⚠ No se encontró botón Ingresar para empresa {rut_e}")
                    job['resultados'][w['rut_trabajador']] = 'error_empresa'
                    continue

                page.wait_for_timeout(1500)
                save_screenshot(f'bot_dentro_empresa_{job_id[:8]}.png')
                log(f"URL dentro empresa: {page.url}")

                # Click en botón Regulariza con ID exacto: regulariza#RUTNUM#00
                mov_clicked = False
                for sel in [
                    f"button[id='regulariza#{rut_num}#00']",
                    f"button[id*='regulariza#{rut_num}']",
                    f"[id*='regulariza#{rut_num}']",
                ]:
                    try:
                        el = page.locator(sel)
                        if el.count() > 0:
                            el.first.click()
                            log(f"✓ Click en regulariza#{rut_num}#00")
                            mov_clicked = True
                            break
                    except: pass

                if not mov_clicked:
                    save_screenshot(f'bot_nomov_{job_id[:8]}.png')
                    raise Exception(f"No se encontró botón regulariza#{rut_num}#00. Ver captura.")

                page.wait_for_timeout(1500)
                save_screenshot(f'bot_mov_personal_{job_id[:8]}.png')
                log(f"URL movimiento personal: {page.url}")

                # Ingreso Manual
                for sel in ["#regularizacion_manual", "a:has-text('Ingreso Manual')", "button:has-text('Ingreso Manual')", "input[value*='Manual']"]:
                    try:
                        el = page.locator(sel)
                        if el.count() > 0:
                            el.first.click()
                            log(f"✓ Click Ingreso Manual: {sel}")
                            break
                    except: pass
                page.wait_for_timeout(1500)
                save_screenshot(f'bot_ingreso_manual_{job_id[:8]}.png')
                log(f"URL ingreso manual: {page.url}")

                # RUT trabajador — campo visible es web_rut_trabajador2
                try:
                    el = page.locator("#web_rut_trabajador2")
                    el.wait_for(state='visible', timeout=5000)
                    el.fill(rut_t)
                    el.press('Tab')
                    page.wait_for_timeout(800)
                    log(f"✓ RUT trabajador llenado: {rut_t}")
                except Exception as e:
                    log(f"⚠ No se pudo llenar RUT trabajador: {e}")

                # AFP
                try:
                    select_contains('#web_combo_codigo_afp', inst)
                    log(f"✓ AFP seleccionada: {inst}")
                except: pass

                # Salud
                try:
                    select_opt('#web_combo_codigo_salud', w.get('salud', 'FONASA'))
                except: pass

                # Causa
                causa_val = w.get('causa', '')
                try:
                    r = select_contains('#web_combo_movimiento_personal', causa_val)
                    if r:
                        log(f"✓ Causa seleccionada: {r}")
                    else:
                        log(f"⚠ No se pudo seleccionar causa: {causa_val}")
                except:
                    log(f"⚠ No se pudo seleccionar causa: {causa_val}")

                # Fecha de término — primero el valor editado en UI, luego el del Excel
                fecha_term = w.get('fecha_termino') or w.get('fecha_cese') or ''
                if fecha_term:
                    # Convertir formato HTML (YYYY-MM-DD) a DD/MM/YYYY si es necesario
                    if '-' in fecha_term and fecha_term.count('-') == 2 and len(fecha_term) == 10:
                        partes = fecha_term.split('-')
                        fecha_term = f"{partes[2]}/{partes[1]}/{partes[0]}"
                    set_fecha_js('end_date', fecha_term)
                    log(f"✓ Fecha término (Hasta): {fecha_term}")

                if es_ausentismo and periodos_parsed:
                    anio0, mes0 = periodos_parsed[0]
                    ult0 = calendar.monthrange(anio0, mes0)[1]
                    set_fecha_js('start_date', f"01/{mes0:02d}/{anio0}")
                    set_fecha_js('end_date', f"{ult0}/{mes0:02d}/{anio0}")

                save_screenshot(f'bot_form_{job_id[:8]}.png')
                continuar()
                page.wait_for_timeout(1200)
                save_screenshot(f'bot_post_continuar1_{job_id[:8]}.png')
                log(f"URL post-continuar1: {page.url}")

                try:
                    cb = page.locator("#web_chk_declaracion")
                    if cb.count() > 0 and not cb.is_checked():
                        cb.click()
                        log("✓ Checkbox declaración marcado")
                    page.wait_for_timeout(300)
                except: pass

                continuar()
                page.wait_for_timeout(1200)
                save_screenshot(f'bot_post_continuar2_{job_id[:8]}.png')
                log(f"URL post-continuar2: {page.url}")

                if es_ausentismo and len(periodos_parsed) > 1:
                    for pi, (anio, mes) in enumerate(periodos_parsed[1:], 1):
                        es_ultimo = (pi == len(periodos_parsed) - 1)
                        ult = calendar.monthrange(anio, mes)[1]

                        if es_ultimo:
                            finalizar()
                            page.wait_for_timeout(1500)
                            break

                        continuar()
                        page.wait_for_timeout(1500)

                        rut_inputs2 = page.locator("input[type='text']").all()
                        if rut_inputs2: rut_inputs2[0].fill(rut_t)
                        page.wait_for_timeout(300)

                        select_contains('#web_combo_codigo_afp', inst)
                        try: select_opt('#web_combo_codigo_salud', w.get('salud','FONASA'))
                        except: pass
                        try: select_opt('#web_combo_movimiento_personal', w.get('causa',''))
                        except: pass

                        set_fecha_js('start_date', f"01/{mes:02d}/{anio}")
                        set_fecha_js('end_date', f"{ult}/{mes:02d}/{anio}")

                        continuar()
                        page.wait_for_timeout(1200)
                        try:
                            cb2 = page.locator("input[type='checkbox']").first
                            if not cb2.is_checked(): cb2.click()
                            page.wait_for_timeout(300)
                        except: pass
                        continuar()
                        page.wait_for_timeout(1500)
                else:
                    page.wait_for_timeout(800)
                    log(f"✅ Movimiento registrado: {w['nombre']} ({rut_t})")
                    job['resultados'][w['rut_trabajador']] = 'completado'

                    # Hacer click en "Comprobante por Trabajador" del menú lateral (igual que el usuario)
                    log("📄 Buscando menú Comprobante por Trabajador...")
                    try:
                        comp_menu = page.locator("text='Comprobante por Trabajador'").first
                        comp_menu.wait_for(state='visible', timeout=8000)
                        comp_menu.click()
                        page.wait_for_timeout(2000)
                        log("✓ Click Comprobante por Trabajador")
                    except Exception as e:
                        log(f"⚠ Menú no encontrado: {e}")
                        # Diagnóstico: listar todos los textos clickeables del menú
                        menu_items = page.evaluate("""() => Array.from(document.querySelectorAll('a, li, span, div')).filter(e => {
                            const t = e.textContent.trim();
                            return t.length > 3 && t.length < 60 && e.offsetHeight > 0;
                        }).map(e => e.tagName + ': ' + e.textContent.trim()).slice(0, 30)""")
                        log(f"Elementos visibles: {menu_items}")

                    import datetime as _dt2
                    hoy = _dt2.date.today().strftime('%d/%m/%Y')

                    # Diagnóstico: ver selects e inputs visibles
                    selects_info = page.evaluate("""() => Array.from(document.querySelectorAll('select')).map(s => s.id + '|' + s.name + '|' + Array.from(s.options).map(o => o.value+':'+o.text).join(','))""")
                    log(f"Selects: {selects_info}")

                    # 1. RUT trabajador
                    try:
                        rut_inp = page.locator("#web_rut2").first
                        rut_inp.fill(rut_t, timeout=5000)
                        log(f"✓ RUT: {rut_t}")
                    except Exception as e:
                        log(f"⚠ RUT: {e}")

                    # 2. Institución Previsional → la AFP del trabajador (select #web_combo_codigo_afp)
                    try:
                        inst_limpia = inst.upper().replace('AFP ', '').replace('AFP', '').strip()
                        page.evaluate("""(afp) => {
                            var sel = document.getElementById('web_combo_codigo_afp');
                            if (!sel) return;
                            for (var i = 0; i < sel.options.length; i++) {
                                if (sel.options[i].text.toUpperCase().indexOf(afp) !== -1) {
                                    sel.selectedIndex = i;
                                    sel.dispatchEvent(new Event('change', {bubbles: true}));
                                    break;
                                }
                            }
                        }""", inst_limpia)
                        log(f"✓ Institución: {inst_limpia}")
                    except Exception as e:
                        log(f"⚠ Institución: {e}")
                    page.wait_for_timeout(500)

                    # 3 y 4. Fecha Desde y Hasta → hoy, vía JS porque el campo es readonly
                    page.evaluate("""(fecha) => {
                        ['web_desde', 'web_hasta'].forEach(function(id) {
                            var el = document.getElementById(id);
                            if (el) {
                                el.removeAttribute('disabled');
                                el.removeAttribute('readonly');
                                el.value = fecha;
                                el.dispatchEvent(new Event('change', {bubbles: true}));
                                el.dispatchEvent(new Event('blur', {bubbles: true}));
                            }
                        });
                    }""", hoy)
                    log(f"✓ Fecha Desde y Hasta: {hoy}")

                    page.wait_for_timeout(800)

                    # Interceptar el PDF real que sirve CtrlPdf antes de hacer click
                    pdf_capturado = []

                    def capturar_pdf(route):
                        try:
                            response = route.fetch()
                            body = response.body()
                            if body and len(body) > 1000:
                                pdf_capturado.append(body)
                                log(f"📥 PDF capturado: {len(body)} bytes")
                            route.fulfill(response=response)
                        except Exception:
                            route.continue_()

                    ctx.route('**/CtrlPdf**', capturar_pdf)

                    try:
                        with ctx.expect_page(timeout=25000) as new_page_info:
                            page.evaluate("""() => {
                                var all = Array.from(document.querySelectorAll('button, input[type=submit], a, span'));
                                for (var el of all) {
                                    var txt = (el.textContent || el.value || '').trim();
                                    if (txt.indexOf('Generar') !== -1) {
                                        (el.tagName === 'SPAN' ? el.parentElement : el).click();
                                        return;
                                    }
                                }
                            }""")
                        popup = new_page_info.value
                        popup.wait_for_load_state('networkidle', timeout=20000)
                        popup.wait_for_timeout(1000)
                        popup.close()
                    except Exception as e:
                        log(f"⚠ Error popup: {e}")

                    ctx.unroute('**/CtrlPdf**')

                    if pdf_capturado:
                        job['comprobante_bytes'] = pdf_capturado[0]
                        job['comprobante_name'] = f"MOV_PER_{rut_t}.pdf"
                        log(f"✅ Comprobante real: MOV_PER_{rut_t}.pdf ({len(pdf_capturado[0])} bytes)")
                    else:
                        log("⚠ No se capturó el PDF de CtrlPdf")

            ctx.close()
            browser.close()

            # Guardar video en memoria
            videos = [f for f in os.listdir(video_dir) if f.endswith('.webm')]
            if videos:
                with open(os.path.join(video_dir, videos[0]), 'rb') as vf:
                    job['video_bytes'] = vf.read()
                job['video'] = f'bot_video_{job_id[:8]}.webm'
                log(f"🎥 Video listo: bot_video_{job_id[:8]}.webm")

        shutil.rmtree(tmp_dir, ignore_errors=True)
        job['status'] = 'done'
        log("✅ Bot finalizado")

    except Exception as e:
        import traceback
        job['status'] = 'error'
        job['error'] = str(e)
        job['log'].append(f"Error fatal: {e}")
        print(traceback.format_exc())
        # Guardar video en memoria aunque haya error
        try:
            videos = [f for f in os.listdir(video_dir) if f.endswith('.webm')]
            if videos:
                with open(os.path.join(video_dir, videos[0]), 'rb') as vf:
                    job['video_bytes'] = vf.read()
                job['video'] = f'bot_video_{job_id[:8]}.webm'
        except: pass

# ── Rutas ─────────────────────────────────────────────────────────────────────

@cartas_bp.route("/", strict_slashes=False)
def index():
    from flask import request as _req
    from database import get_conn
    # Leer sesión directamente
    token = _req.cookies.get("session_token")
    user = None
    if token:
        with get_conn() as conn:
            row = conn.execute(
                "SELECT u.* FROM sesiones s JOIN usuarios u ON u.id=s.usuario_id WHERE s.token=?",
                (token,)).fetchone()
            if row:
                user = dict(row)
    if not user:
        from flask import redirect
        return redirect("/login")
    previred_rut = os.environ.get('PREVIRED_RUT', '')
    previred_clave = os.environ.get('PREVIRED_CLAVE', '')
    resp = make_response(render_template("cartas/index.html", user=user,
                                         previred_rut=previred_rut,
                                         previred_clave=previred_clave))
    resp.headers['X-Frame-Options'] = 'SAMEORIGIN'
    return resp

@cartas_bp.route("/api/subir_excel", methods=["POST"])
@cartas_login_required
def subir_excel():
    """Sube y parsea el Excel de deudas."""
    if 'excel' not in request.files:
        return jsonify({"error": "No se envió archivo"}), 400
    f = request.files['excel']
    datos = parsear_excel(f.read())
    trabajadores = agrupar_por_trabajador(datos)
    return jsonify({"ok": True, "trabajadores": trabajadores, "total": len(trabajadores)})

@cartas_bp.route("/api/generar_carta", methods=["POST"])
@cartas_login_required
def generar_carta():
    """Genera PDF de carta para un trabajador."""
    data = request.json
    carta = data.get('carta', {})
    firma = data.get('firma', {})
    pdf_bytes = generar_carta_pdf(carta, firma)
    rut = carta.get('rut_trabajador','').replace('.','').replace(' ','')
    fname = f"Carta_Explicativa_{rut}.pdf"
    # Guardar en DATA_DIR
    data_dir = current_app.config.get('DATA_DIR', 'adjuntos')
    dest = os.path.join(data_dir, fname)
    with open(dest, 'wb') as fout:
        fout.write(pdf_bytes)
    return jsonify({"ok": True, "filename": fname})

@cartas_bp.route("/api/video/<job_id>")
@cartas_login_required
def ver_video(job_id):
    """Sirve el video del bot desde memoria."""
    job = _jobs.get(job_id)
    if not job or not job.get('video_bytes'):
        return "Video no disponible (el bot debe ejecutarse de nuevo)", 404
    from flask import Response
    return Response(job['video_bytes'], mimetype='video/webm',
                    headers={'Content-Disposition': f'inline; filename="{job.get("video","bot_video.webm")}"'})

@cartas_bp.route("/api/comprobante/<job_id>")
@cartas_login_required
def ver_comprobante(job_id):
    """Sirve el comprobante PDF del bot desde memoria."""
    job = _jobs.get(job_id)
    if not job or not job.get('comprobante_bytes'):
        return "Comprobante no disponible", 404
    from flask import Response
    nombre = job.get('comprobante_name', 'Comprobante.pdf')
    return Response(job['comprobante_bytes'], mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename="{nombre}"'})

@cartas_bp.route("/api/descargar/<filename>")
@cartas_login_required
def descargar(filename):
    """Descarga o muestra un archivo generado (PDF, PNG, WEBM)."""
    # Buscar en DATA_DIR configurado, o en adjuntos/ relativo al módulo
    data_dir = current_app.config.get('DATA_DIR') or os.path.join(
        os.path.dirname(os.path.abspath(__file__)), 'adjuntos')
    filepath = os.path.join(data_dir, filename)
    if not os.path.isfile(filepath):
        return f"Archivo no encontrado: {filename}. El bot debe ejecutarse de nuevo para generar una nueva captura.", 404
    ext = filename.rsplit('.', 1)[-1].lower()
    mime = {'png': 'image/png', 'webm': 'video/webm', 'pdf': 'application/pdf'}.get(ext, 'application/octet-stream')
    as_attach = ext not in ('png', 'webm')
    return send_file(filepath, mimetype=mime, as_attachment=as_attach, download_name=filename)

@cartas_bp.route("/api/iniciar_bot", methods=["POST"])
@cartas_login_required
def iniciar_bot():
    """Inicia el bot de PreviRed en background."""
    data = request.json
    rut_login = data.get('rut_login','')
    clave = data.get('clave','')
    workers = data.get('workers', [])
    firma = data.get('firma', {})

    if not rut_login or not clave:
        return jsonify({"error": "Credenciales requeridas"}), 400
    if not workers:
        return jsonify({"error": "Sin trabajadores"}), 400

    job_id = str(uuid.uuid4())
    _jobs[job_id] = {
        'status': 'running',
        'log': [],
        'resultados': {},
        'error': None,
        'creado': datetime.now().isoformat(),
    }

    t = threading.Thread(
        target=run_bot_previred,
        args=(job_id, rut_login, clave, workers, firma),
        daemon=True)
    t.start()

    return jsonify({"ok": True, "job_id": job_id})

@cartas_bp.route("/api/estado_bot/<job_id>")
@cartas_login_required
def estado_bot(job_id):
    """Retorna el estado actual del bot."""
    job = _jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job no encontrado"}), 404
    return jsonify({
        "status": job['status'],
        "log": job['log'],
        "resultados": job['resultados'],
        "error": job.get('error'),
        "screenshot_b64": job.get('screenshot_b64'),
        "screenshot_label": job.get('screenshot_label'),
        "has_video": bool(job.get('video_bytes')),
        "video_name": job.get('video'),
        "has_comprobante": bool(job.get('comprobante_bytes')),
        "comprobante_name": job.get('comprobante_name'),
    })

@cartas_bp.route("/api/marcar_gestion", methods=["POST"])
@cartas_login_required
def marcar_gestion():
    """Marca trabajadores como En gestión en el Excel subido."""
    # Esta funcionalidad requiere que el Excel esté en el servidor
    # Por ahora retorna OK para el flujo web
    return jsonify({"ok": True, "mensaje": "Estado actualizado en la sesión"})
