"""Conexión al Excel de Encargados DICOM (SharePoint) — hoja "Empresas".

El enlace compartido se configura en la variable de entorno
DICOM_EMPRESAS_URL (NO va en el código: el repositorio es público) o,
si no existe, en la base de datos (tabla app_config — se pega desde la
página /dicom).
"""

import io
import os
import time
import threading
import http.cookiejar
import urllib.request

import openpyxl

REFRESCO_SEG = 600  # 10 minutos
_CACHE = {"encargados": None, "ts": 0, "error": None}
_LOCK = threading.Lock()


def url_guardada():
    url = os.environ.get("DICOM_EMPRESAS_URL", "").strip()
    if url:
        return url
    try:
        from database import get_conn
        with get_conn() as conn:
            row = conn.execute(
                "SELECT valor FROM app_config WHERE clave='dicom_empresas_url'").fetchone()
            return (row["valor"] or "").strip() if row else ""
    except Exception:
        return ""


def guardar_url(url):
    from database import get_conn
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO app_config(clave, valor) VALUES('dicom_empresas_url', ?) "
            "ON CONFLICT(clave) DO UPDATE SET valor=excluded.valor", (url.strip(),))
    with _LOCK:
        _CACHE.update({"encargados": None, "ts": 0, "error": None})


def _url_descarga(url):
    """Convierte el share-link de SharePoint al endpoint de descarga directa."""
    if "download.aspx" in url:
        return url
    try:
        if "/:x:/g/personal/" in url:
            dominio = url.split("/:x:/")[0]
            resto = url.split("/:x:/g/personal/")[1]
            usuario, token = resto.split("/", 1)
            token = token.split("?")[0]
            return f"{dominio}/personal/{usuario}/_layouts/15/download.aspx?share={token}"
    except Exception:
        pass
    return url


def _descargar():
    url = url_guardada()
    if not url:
        raise Exception("Falta pegar el enlace del Excel de Encargados DICOM")
    durl = _url_descarga(url)
    cj = http.cookiejar.CookieJar()
    op = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    op.addheaders = [("User-Agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64)")]
    with op.open(durl, timeout=90) as r:
        data = r.read()
    if data[:2] != b"PK":
        raise Exception("SharePoint no devolvió el Excel — revisa que el enlace siga "
                        "vigente y compartido como 'Cualquier persona puede ver'")
    return data


def _norm_grupo(grupo):
    """Normaliza el nombre de GRUPO para cruzarlo entre BASE MADRE y este
    Excel (pueden venir con distinta may/min o espacios de mas)."""
    import unicodedata
    s = unicodedata.normalize("NFKD", str(grupo or "")).encode("ascii", "ignore").decode()
    return " ".join(s.strip().lower().split())


def obtener_encargados_dicom(forzar=False):
    """Devuelve ({grupo_normalizado: 'Responsable DICOM'}, error) leyendo la
    hoja "Empresas" del Excel. El cruce es por GRUPO (no por RUT): dentro de
    ese Excel el Responsable DICOM es el mismo para todas las empresas de un
    mismo grupo. Sirve desde cache si la última lectura tiene menos de
    REFRESCO_SEG. Si falla pero hay cache previo, sigue sirviendo los datos
    antiguos y reporta el error."""
    with _LOCK:
        fresco = _CACHE["encargados"] is not None and (time.time() - _CACHE["ts"]) < REFRESCO_SEG
        if fresco and not forzar:
            return _CACHE["encargados"], _CACHE["error"]

    try:
        data = _descargar()
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            ws = None
            for nombre in wb.sheetnames:
                if "empresa" in nombre.lower():
                    ws = wb[nombre]
                    break
            if ws is None:
                ws = wb.active

            gen = ws.iter_rows(values_only=True)
            headers = [str(h).strip() if h else "" for h in (next(gen, None) or ())]

            def encontrar_col(terms):
                for i, h in enumerate(headers):
                    hn = h.lower()
                    if all(t in hn for t in terms):
                        return i
                return None

            i_grupo = encontrar_col(["grupo"])
            i_resp = encontrar_col(["responsable", "dicom"])

            encargados = {}
            if i_grupo is not None and i_resp is not None:
                for row in gen:
                    if row is None or i_grupo >= len(row):
                        continue
                    grupo = _norm_grupo(row[i_grupo])
                    if not grupo:
                        continue
                    resp = str(row[i_resp]).strip() if i_resp < len(row) and row[i_resp] else ""
                    if resp:
                        encargados[grupo] = resp
        finally:
            wb.close()

        with _LOCK:
            _CACHE.update({"encargados": encargados, "ts": time.time(), "error": None})
        return encargados, None
    except Exception as e:
        with _LOCK:
            _CACHE["error"] = str(e)
            resultado = _CACHE["encargados"] or {}
        return resultado, str(e)
