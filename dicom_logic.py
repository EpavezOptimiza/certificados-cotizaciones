"""Bot para el portal DICOM / Equifax (sec.equifax.cl/compraonline).

Por ahora SOLO inicia sesión y comprueba que entró. No compra ni descarga
nada: el portal cobra por cada documento y esa parte se agrega aparte,
cuando esté confirmado cómo se paga.
"""

import os
import time

from playwright.sync_api import sync_playwright

URL_LOGIN = "https://business.equifax.ca/auth/login"

# Paso 1 — identificadores verificados en el portal real (Okta / Client Central)
_SEL_USUARIO = "#idp-discovery-username, input[name='username']"
_SEL_NEXT    = "#idp-discovery-submit, input[type='submit'][value='Next']"
# Paso 2 — la contraseña aparece después de "Next"
_SEL_CLAVE   = "input[type='password']"


class LoginFallido(Exception):
    """No se pudo iniciar sesión en DICOM/Equifax."""


def _texto(page, limite=400):
    try:
        return (page.evaluate(
            "() => (document.body ? document.body.innerText : '')"
            ".replace(/\\s+/g, ' ').trim()") or "")[:limite]
    except Exception:
        return ""


def _sigue_en_login(page):
    """True si el formulario de login sigue en pantalla (no entró)."""
    try:
        return page.locator(_SEL_CLAVE).count() > 0
    except Exception:
        return False


def _dump_campos(page, log):
    """Detalla todos los campos visibles: etiqueta, placeholder, si es
    obligatorio y si tiene valor. Sirve para identificar campos ocultos
    a la vista (captcha, código, etc.) que estén frenando el envío."""
    try:
        campos = page.evaluate("""() => {
            const out = [];
            for (const e of document.querySelectorAll('input, select, textarea')) {
                if (e.offsetParent === null || e.type === 'hidden') continue;
                let etiqueta = e.getAttribute('aria-label') || '';
                if (!etiqueta && e.id) {
                    const l = document.querySelector('label[for="' + e.id + '"]');
                    if (l) etiqueta = l.innerText.trim();
                }
                if (!etiqueta) {
                    const cont = e.closest('div, li, td');
                    if (cont) {
                        const l = cont.querySelector('label');
                        if (l) etiqueta = l.innerText.trim();
                    }
                }
                out.push({
                    tipo: e.type || e.tagName.toLowerCase(),
                    name: e.name || '', id: e.id || '',
                    etiqueta: (etiqueta || '').replace(/\\s+/g, ' ').slice(0, 40),
                    placeholder: e.placeholder || '',
                    obligatorio: !!e.required || e.getAttribute('aria-required') === 'true',
                    conValor: !!(e.value || '').trim()
                });
            }
            return out.slice(0, 10);
        }""")
        log("  [debug] campos de la pantalla:", "warn")
        for c in campos:
            log(f"    {c['tipo']} name={c['name']!r} id={c['id']!r} "
                f"etiqueta={c['etiqueta']!r} ph={c['placeholder']!r} "
                f"obligatorio={c['obligatorio']} conValor={c['conValor']}", "warn")
    except Exception:
        pass


def _hay(page, selector):
    try:
        return page.locator(selector).count() > 0
    except Exception:
        return False


def _tipear(page, selector, texto, log=None, timeout_click=8000):
    """Escribe TECLA POR TECLA (no page.fill).

    El portal está hecho con un framework JS: si el valor se asigna por
    programa, la página lo muestra pero internamente lo considera vacío y el
    botón de acceso no hace nada. Con pulsaciones reales sí lo registra.

    El click lleva timeout corto: si el campo está bloqueado o prellenado
    (Okta lo hace con el usuario), no se pierden 45 s esperándolo.
    """
    loc = page.locator(selector).first
    try:
        loc.click(timeout=timeout_click)
    except Exception:
        # Campo no clickeable (readonly/disabled/cubierto): enfocar por JS
        try:
            loc.evaluate("e => e.focus()")
        except Exception:
            if log:
                log(f"  no se pudo enfocar {selector[:30]}", "warn")
            return False
    try:
        loc.press("Control+a")
        loc.press("Backspace")
    except Exception:
        pass
    try:
        loc.press_sequentially(texto, delay=70)
    except AttributeError:          # Playwright antiguo
        loc.type(texto, delay=70)
    except Exception:
        if log:
            log(f"  no se pudo escribir en {selector[:30]}", "warn")
        return False
    try:
        loc.dispatch_event("input")
        loc.dispatch_event("change")
    except Exception:
        pass
    return True


VERSION = "v6 (corrige el falso error por la palabra 'credenciales')"


def _pide_verificacion(page):
    """True si el portal muestra la pantalla de verificación en dos pasos,
    ya sea "elige un método" (paso 2/3) o "ingresa el código" (paso 3/3)."""
    try:
        return bool(page.evaluate("""() => {
            const t = (document.body ? document.body.innerText : '').toLowerCase();
            return t.includes('verify your account') || t.includes('send me the code') ||
                   t.includes('verification code') || t.includes('security code') ||
                   t.includes('codigo de verificacion') ||
                   t.includes('código de verificación') ||
                   t.includes('verifique su cuenta') || t.includes('enviarme el codigo') ||
                   t.includes('enviarme el código') ||
                   t.includes('metodo de verificacion') ||
                   t.includes('método de verificación') ||
                   t.includes('codigo de seguridad') ||
                   t.includes('código de seguridad');
        }"""))
    except Exception:
        return False


def _pide_elegir_metodo(page):
    """True si el portal está en el paso "Selecciona un método de
    verificación" (aún no se pidió el envío del código)."""
    try:
        return bool(page.evaluate("""() => {
            const t = (document.body ? document.body.innerText : '').toLowerCase();
            return t.includes('metodo de verificacion') || t.includes('método de verificación') ||
                   t.includes('select a verification method') ||
                   t.includes('choose a verification method');
        }"""))
    except Exception:
        return False


def _resolver_verificacion(page, log, obtener_codigo):
    """Pide el envío del código, espera a que la persona lo escriba en la
    plataforma y lo ingresa en el portal (flujo semiautomático).

    El asistente nuevo de Equifax (3 pasos) agrega un paso intermedio que
    el flujo viejo no tenía: "Selecciona un método de verificación", con
    un radio button por método (acá solo aparece correo) y un botón
    "Enviar" separado. Hay que marcar el método ANTES de apretar Enviar,
    si no, el click no hace nada.
    """
    log("El portal pide un código de verificación", "warn")

    # 0. Paso "Selecciona un método de verificación": marcar el radio del
    # método (normalmente el único disponible, correo) antes de enviar.
    if _pide_elegir_metodo(page):
        log("Seleccionando método de verificación (correo)...", "info")
        try:
            radio = page.locator("input[type='radio']").first
            if radio.count() > 0 and not radio.is_checked():
                radio.click(timeout=4000)
        except Exception:
            # Algunos diseños hacen clickeable toda la fila, no solo el radio
            try:
                page.evaluate("""() => {
                    const fila = document.querySelector('[class*="method"], [class*="option"], label');
                    if (fila) fila.click();
                }""")
            except Exception:
                pass
        time.sleep(0.5)

    # 1. Botón "Enviar" / "SEND ME THE CODE" (si aún no se ha enviado)
    try:
        enviado = page.evaluate("""() => {
            for (const b of document.querySelectorAll('button, input[type=submit], a')) {
                const t = ((b.innerText || b.value || '')).trim().toLowerCase();
                if (t === 'enviar' || t.includes('send me the code') || t.includes('enviarme el') ||
                    t.includes('send code') || t.includes('enviar código') ||
                    t.includes('enviar codigo')) { b.click(); return true; }
            }
            return false;
        }""")
        if enviado:
            log("Código solicitado — revisa tu correo", "info")
            time.sleep(3)
    except Exception:
        pass

    # 2. Esperar a que la persona escriba el código en la plataforma
    codigo = obtener_codigo()
    if not codigo:
        raise LoginFallido(
            "No se recibió el código de verificación a tiempo. Vuelve a intentarlo "
            "y escribe el código apenas llegue a tu correo.")
    log("Código recibido — ingresándolo...", "info")

    # 3. Escribirlo en el campo correspondiente
    campo = None
    for sel in ("input[name='passCode']", "input[autocomplete='one-time-code']",
                "input[type='tel']", "input[name*='code' i]", "input[id*='code' i]",
                "input[type='text']"):
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=4000)
            campo = sel
            break
        except Exception:
            continue
    if not campo:
        raise LoginFallido("No se encontró dónde escribir el código de verificación")

    _tipear(page, campo, str(codigo).strip(), log)
    time.sleep(0.5)

    # 4. Enviar
    for descripcion, loc in (
        ("botón Verify", page.get_by_role("button", name="Verify")),
        ("botón Verificar", page.get_by_role("button", name="Verificar")),
        ("submit", page.locator("input[type='submit'], button[type='submit']")),
    ):
        try:
            el = loc.first
            el.wait_for(state="visible", timeout=3000)
            el.click(timeout=4000)
            log(f"Código enviado ({descripcion})", "info")
            break
        except Exception:
            continue
    else:
        try:
            page.locator(campo).first.press("Enter")
        except Exception:
            pass

    # 5. Esperar a que pase la verificación
    fin = time.time() + 60
    while time.time() < fin:
        if not _pide_verificacion(page):
            return True
        time.sleep(1)
    raise LoginFallido("El portal no aceptó el código de verificación")


def hacer_login(page, correo, clave, log, obtener_codigo=None):
    """Login en dos pasos: usuario → Next → contraseña → (código) → entrar."""
    log(f"DICOM {VERSION}", "info")
    log("Abriendo el portal de Equifax...", "info")
    page.goto(URL_LOGIN, wait_until="domcontentloaded", timeout=45000)

    # ── Paso 1: usuario (el correo) y botón Next ────────────────────────────
    try:
        page.wait_for_selector(_SEL_USUARIO, state="visible", timeout=25000)
    except Exception:
        raise LoginFallido(f"No apareció el campo Username. "
                           f"La página muestra: {_texto(page, 200)}")

    log("Escribiendo el usuario...", "info")
    _tipear(page, _SEL_USUARIO, correo, log)
    time.sleep(0.4)

    log("Presionando Next...", "info")
    try:
        page.click(_SEL_NEXT, timeout=8000)
    except Exception:
        try:
            page.press(_SEL_USUARIO, "Enter")
        except Exception:
            raise LoginFallido("No se pudo presionar Next")

    # ── Paso 2: pantalla de Okta (usuario + contraseña) ─────────────────────
    try:
        page.wait_for_selector(_SEL_CLAVE, state="visible", timeout=25000)
        # Dar tiempo a que el widget de Okta termine de montarse: si se escribe
        # antes, el texto queda en pantalla pero el widget no lo registra.
        time.sleep(2.5)
    except Exception:
        detalle = _texto(page, 250)
        raise LoginFallido(
            "Tras Next no apareció el campo de contraseña. "
            f"El portal muestra: {detalle}")

    # Tras "Next" el portal redirige a su pantalla de acceso. Equifax rediseñó
    # este paso como un asistente ("Paso de inicio de sesión 1 de 3"): ya no
    # es el widget clásico de Okta con usuario+contraseña juntos, sino una
    # pantalla dedicada SOLO a la contraseña (el usuario ya quedó fijado en
    # el paso anterior y se muestra como texto "Iniciando sesión como
    # correo@... Cambiar", sin campo editable). El campo real de contraseña
    # confirmado en el HTML es id="knowledgeFactor" (nombre en jerga de Okta
    # Identity Engine para el factor de autenticación por clave).
    _OKTA_USER = "#okta-signin-username, input[name='username']"
    _OKTA_PASS = "#knowledgeFactor, #okta-signin-password, input[name='password']"

    # Solo intentar reescribir el usuario si existe un campo EDITABLE de
    # usuario en esta pantalla (el asistente nuevo no lo tiene: es solo
    # texto informativo). Antes esto asumía que "no se pudo leer" el valor
    # equivalía a "hay que escribirlo", lo que generaba un intento de foco
    # fallido en un campo que ya no existe en el flujo nuevo.
    hay_campo_usuario = False
    try:
        hay_campo_usuario = page.locator(_OKTA_USER).count() > 0
    except Exception:
        pass

    if hay_campo_usuario:
        try:
            actual = page.evaluate(
                """(sel) => { const e = document.querySelector(sel.split(',')[0].trim())
                     || document.querySelector("input[name='username']");
                     return e ? (e.value || '') : null; }""", _OKTA_USER)
        except Exception:
            actual = None

        if actual and actual.strip().lower() == correo.strip().lower():
            log("El usuario ya viene puesto por el portal", "info")
        else:
            log("Escribiendo el usuario...", "info")
            try:
                _tipear(page, _OKTA_USER, correo, log)
            except Exception:
                pass
            time.sleep(0.6)
    else:
        log("Este paso no pide el usuario de nuevo (ya quedó fijado antes)", "info")

    log("Escribiendo la contraseña...", "info")
    if not _tipear(page, _OKTA_PASS, clave, log):
        _dump_campos(page, log)
    time.sleep(0.8)

    # ¿El botón quedó activo? Si Okta no registró lo escrito, sigue inhabilitado
    # y el click no hace nada (sin mostrar error). En ese caso se reescribe.
    def _estado_boton():
        try:
            return page.evaluate("""() => {
                for (const b of document.querySelectorAll(
                        "input[type=submit], button, .button-primary")) {
                    const t = ((b.innerText || b.value || '')).trim().toLowerCase();
                    if (t.includes('iniciar') || t.includes('sign in')) {
                        const cs = getComputedStyle(b);
                        return {texto: (b.innerText || b.value || '').trim(),
                                deshabilitado: !!b.disabled ||
                                    b.getAttribute('aria-disabled') === 'true' ||
                                    (b.className || '').includes('disabled'),
                                opacidad: cs.opacity};
                    }
                }
                return null;
            }""")
        except Exception:
            return None

    est = _estado_boton()
    log(f"  estado del botón: {est}", "info")

    if est and est.get("deshabilitado"):
        log("  el botón está inhabilitado — reescribiendo los campos...", "warn")
        # Reescribir despacio, campo por campo, dando tiempo a que el widget valide
        try:
            _tipear(page, "#okta-signin-username, input[name='username']", correo, log)
            time.sleep(0.8)
            _tipear(page, "#okta-signin-password, input[name='password']", clave, log)
            time.sleep(1.2)
            est = _estado_boton()
            log(f"  estado del botón tras reescribir: {est}", "info")
        except Exception as e:
            log(f"  no se pudo reescribir: {type(e).__name__}", "warn")

    # Qué botones hay realmente en esta pantalla (queda en el log para diagnóstico)
    try:
        botones = page.evaluate("""() => Array.from(
            document.querySelectorAll("button, input[type=submit], input[type=button], a"))
            .filter(e => e.offsetParent !== null)
            .map(e => (e.tagName + '[' + (e.type || '') + ']' +
                       (e.id ? '#' + e.id : '') + ' "' +
                       (e.innerText || e.value || '').trim().slice(0, 25) + '"'))
            .slice(0, 8)""")
        log(f"  botones en pantalla: {botones}", "info")
    except Exception:
        pass

    url_antes = page.url

    # Enviar con un click REAL (los clicks por programa tampoco los registra
    # el framework). Se prueba por texto del botón y luego por submit.
    enviado = ""
    for descripcion, localizador in (
        ("botón Iniciar Sesión", page.get_by_role("button", name="Iniciar Sesión")),
        ("botón Sign in",        page.get_by_role("button", name="Sign in")),
        ("texto Iniciar Sesión", page.get_by_text("Iniciar Sesión", exact=False)),
        ("submit",               page.locator("input[type='submit'], button[type='submit']")),
    ):
        try:
            el = localizador.first
            el.wait_for(state="visible", timeout=4000)
            el.click(timeout=5000)
            enviado = descripcion
            break
        except Exception:
            continue
    if not enviado:
        try:
            page.locator(_SEL_CLAVE).first.press("Enter")
            enviado = "Enter"
        except Exception:
            enviado = "no se pudo enviar"
    log(f"  enviado con: {enviado}", "info")

    log("Enviando...", "info")
    # Esperar hasta 90 s: el portal pasa por /auth/implicit/callback y ahí se
    # queda "cargando" un buen rato antes de entrar. Se informa el avance.
    fin = time.time() + 90
    error_texto = ""
    ultimo_aviso = time.time()
    url_previa = page.url
    estable = 0          # veces seguidas SIN formulario (evita el falso positivo)
    while time.time() < fin:
        # ¿Cambió de dirección? Señal de que el acceso avanzó
        try:
            if page.url != url_previa:
                log(f"  → {page.url[:90]}", "info")
                url_previa = page.url
                if "callback" in page.url.lower() or "login" not in page.url.lower():
                    log("  El portal está procesando el acceso...", "info")
        except Exception:
            pass

        # El formulario de Okta desaparece un instante al enviarse: hay que
        # confirmar que se fue DE VERDAD (varias comprobaciones seguidas).
        if not _sigue_en_login(page):
            estable += 1
            if estable >= 4:
                break
        else:
            estable = 0

        if time.time() - ultimo_aviso > 15:
            ultimo_aviso = time.time()
            restante = int(fin - time.time())
            log(f"  esperando respuesta del portal... ({restante}s restantes)", "info")
        # Errores REALES: solo desde los recuadros de error de Okta, nunca del
        # texto general de la página. (Antes se buscaba la palabra "credenciales"
        # en todo el cuerpo y coincidía con "Ingrese sus credenciales para
        # iniciar sesión", el texto de bienvenida: fallo instantáneo y falso.)
        try:
            error_texto = page.evaluate("""() => {
                const cajas = document.querySelectorAll(
                    '.okta-form-infobox-error, [data-se="o-form-error-container"], ' +
                    '.o-form-error-container, .infobox-error, [role="alert"]');
                for (const c of cajas) {
                    if (c.offsetParent === null) continue;
                    const t = (c.innerText || '').trim().replace(/\\s+/g, ' ');
                    if (t) return t.slice(0, 200);
                }
                return '';
            }""") or ""
        except Exception:
            pass
        if error_texto:
            break
        time.sleep(0.5)

    # ── Verificación en dos/tres pasos ──────────────────────────────────────
    # El asistente nuevo de Equifax (3 pasos) muestra un spinner de carga
    # justo después de "Continuar" mientras pasa al siguiente paso -- ese
    # siguiente paso puede ser la pantalla de código de verificación, pero
    # tarda un momento en aparecer. Revisar "_pide_verificacion" una sola
    # vez justo después del submit puede pillar el spinner de transición y
    # no ver el pedido de código todavía (se seguiría de largo sin pedirlo).
    # Por eso se reintenta ambos chequeos (código / seguir en login) varias
    # veces con margen, en vez de una sola foto instantánea.
    if not _pide_verificacion(page):
        for _ in range(6):
            time.sleep(1)
            if _pide_verificacion(page) or not _sigue_en_login(page):
                break

    if _pide_verificacion(page):
        if obtener_codigo is None:
            raise LoginFallido(
                "El portal pide un CÓDIGO DE VERIFICACIÓN y esta prueba no está "
                "preparada para pedírtelo.")
        _resolver_verificacion(page, log, obtener_codigo)

    if _sigue_en_login(page):
        # Diagnóstico completo para saber por qué rebotó
        estado_campos, alertas = "", ""
        try:
            estado_campos = page.evaluate("""() => {
                return Array.from(document.querySelectorAll('input'))
                    .filter(e => e.offsetParent !== null && e.type !== 'hidden')
                    .map(e => (e.type + ':' + (e.name || e.id || '?') + ':' +
                               ((e.value || '').trim() ? 'con dato' : 'VACIO')))
                    .join(' | ');
            }""") or ""
            alertas = page.evaluate("""() => {
                const sels = ['[role=alert]', '.error', '.alert', '.o-form-error-container',
                              '.infobox-error', '[class*=error]', '[class*=Error]'];
                const out = [];
                for (const s of sels) {
                    for (const e of document.querySelectorAll(s)) {
                        const t = (e.innerText || '').trim().replace(/\\s+/g, ' ');
                        if (t && t.length < 200 && !out.includes(t)) out.push(t);
                    }
                }
                return out.slice(0, 3).join(' / ');
            }""") or ""
        except Exception:
            pass
        cambio = "sí" if page.url != url_antes else "no"
        detalle = error_texto or alertas or _texto(page, 180)
        raise LoginFallido(
            f"No se pudo entrar. Portal: {detalle}"
            f"{(' [avisos: ' + alertas + ']') if alertas and alertas != detalle else ''}"
            f" [campos: {estado_campos}] [cambió de página: {cambio}]")

    # Salió del login. Ahora puede quedar en /auth/implicit/callback "cargando":
    # se espera a que termine de resolver antes de dar por buena la entrada.
    if "callback" in (page.url or "").lower():
        log("Procesando el acceso (pantalla de carga)...", "info")
        fin_cb = time.time() + 60
        while time.time() < fin_cb and "callback" in (page.url or "").lower():
            time.sleep(1)
        log(f"  → {page.url[:90]}", "info")

    try:
        page.wait_for_load_state("networkidle", timeout=20000)
    except Exception:
        pass

    log(f"Sesión iniciada — {page.url}", "ok")
    return True


def extraer_rut_de_nombre(nombre_archivo):
    """Extrae el RUT desde el nombre del archivo descargado de Equifax.

    Los Boletin Laboral vienen nombrados como:
      'Boletin Laboral_59061500-5.pdf'        (con guion)
      'Boletin Laboral_96400000K_11-06-2026.pdf'  (sin guion, con fecha)
      'Boletin Laboral_761965476_11-06-2026.pdf'  (sin guion, dv numérico)

    El dígito verificador es siempre el último carácter del bloque de
    RUT (dígito o K); el resto son el cuerpo. Devuelve 'cuerpo-DV' o ''.
    """
    import re
    base = os.path.splitext(nombre_archivo)[0]
    m = re.search(r'(\d{7,8})-?([0-9Kk])(?=_|$)', base)
    if not m:
        return ''
    return f"{m.group(1)}-{m.group(2).upper()}"


import re as _re

# Encabezado de la tabla "Motivo Institución Boletín Pág. Fecha Infracción
# Cotizaciones Meses Monto $ Fiscalizador Resolución" tal como lo entrega
# pypdf (los acentos suelen salir como U+FFFD, por eso se usa "." comodín).
_HEADER_TABLA_RE = _re.compile(
    r'Motivo\s*Instituci.n\s*Bolet.n\s*P.g\.\s*Fecha\s*Infracci.n'
    r'\s*Cotizaciones\s*Meses\s*Monto\s*\$\s*Fiscalizador\s*Resoluci.n'
)
_TRABAJADOR_RE = _re.compile(
    r'((?:\d{2}/\d{4}\s+\$[\d.,]+\s+\d{1,3},\d{2}\s*\n?)+)'
    r'(\d{1,2}\.\d{3}\.\d{3}-[\dKk])\s?([A-ZÑÁÉÍÓÚ \n]+?)\nTotal Trabajador:\s*(\d{1,3},\d{2})'
)
_PERIODO_RE = _re.compile(r'(\d{2}/\d{4})\s+\$([\d.,]+)\s+(\d{1,3},\d{2})')
# El texto extraído pega el motivo directo con el nombre de la institución
# (ej. "...DeclaradasAFP. CUPRUM"); se inserta un espacio antes de estas
# marcas conocidas para que quede legible en el Excel.
_SEPARAR_INSTITUCION_RE = _re.compile(
    r'(?<=\S)(AFP|CCAF|C\.C\.A\.F|DIRECCION DEL TRABAJO|ADM DE FONDOS|MUTUAL|ISL)'
)


def _extraer_bloques_institucion(texto_pag1):
    """Divide el texto de la página 1 en un bloque por cada institución con
    deuda o multa (el Boletin Laboral repite la tabla de encabezado una vez
    por cada institución cuando hay más de una)."""
    bloques = []
    for chunk in _HEADER_TABLA_RE.split(texto_pag1)[1:]:
        m = _re.match(r'\s*(.+?)\s*\d{3}\s*-\s*\d{2}/\d{2}/\d{4}', chunk, _re.S)
        institucion = ' '.join(m.group(1).split()) if m else 'Sin especificar'
        institucion = _SEPARAR_INSTITUCION_RE.sub(r' \1', institucion)

        trabajadores = []
        for tm in _TRABAJADOR_RE.finditer(chunk):
            periodos_raw, rut_t, nombre_t, _total_t = tm.groups()
            nombre_t = ' '.join(nombre_t.split())
            for fecha, monto_pesos, _utm in _PERIODO_RE.findall(periodos_raw):
                trabajadores.append({
                    'rut': rut_t,
                    'nombre': nombre_t,
                    'periodo': fecha,
                    'monto': int(_re.sub(r'[^\d]', '', monto_pesos.split(',')[0])),
                })

        tipo_correo = 'DT' if ('DIRECCION DEL TRABAJO' in institucion.upper()
                                or 'MULTA' in institucion.upper()) else 'Previsional'
        bloques.append({'institucion': institucion, 'tipo_correo': tipo_correo,
                         'trabajadores': trabajadores})
    return bloques


def extraer_datos_pdf(ruta_pdf, nombre_archivo=None):
    """Extrae datos de un PDF Boletin Laboral: RUT empresa, razón social,
    deudas previsionales, multas, monto UTM total e instituciones con deuda
    (cada una con su lista de trabajadores/periodos si aplica).

    El RUT se obtiene primero del NOMBRE DE ARCHIVO (formato fijo que
    trae el portal Equifax) porque es mucho más confiable que buscarlo
    en el texto del PDF. La razón social es la primera línea del PDF.

    Retorna dict con estructura:
    {
      'rut': 'xx.xxx.xxx-x',
      'razon_social': 'NOMBRE EMPRESA',
      'deudas': 0,       # cantidad de Deudas Previsionales (del resumen)
      'multas': 0,       # cantidad de Multas (del resumen)
      'monto_utm': 0.0,  # Monto Total UTM (del resumen)
      'instituciones': [
          {'institucion': str, 'tipo_correo': 'Previsional'|'DT',
           'trabajadores': [{'rut','nombre','periodo','monto'}, ...]},
          ...
      ]
    }
    """
    detalles = {
        'rut': '', 'razon_social': '', 'deudas': 0, 'multas': 0,
        'monto_utm': 0.0, 'instituciones': []
    }
    try:
        import pypdf

        if nombre_archivo:
            detalles['rut'] = extraer_rut_de_nombre(nombre_archivo)

        with open(ruta_pdf, 'rb') as f:
            reader = pypdf.PdfReader(f)
            texto_pag1 = reader.pages[0].extract_text() or ""
            texto_completo = texto_pag1
            for page in reader.pages[1:]:
                texto_completo += page.extract_text() or ""

        if not detalles['rut']:
            ruts = _re.findall(r'\b\d{1,2}\.\d{3}\.\d{3}-[\dKk]\b', texto_completo)
            if ruts:
                detalles['rut'] = ruts[0]

        primera_linea = texto_pag1.split('\n', 1)[0].strip()
        detalles['razon_social'] = primera_linea

        m = _re.search(r'([\d]+,\d+)\s*Monto Total UTM', texto_completo)
        if m:
            detalles['monto_utm'] = float(m.group(1).replace(',', '.'))

        m = _re.search(r'Deudas Previsionales\s*:\s*(\d+)', texto_completo)
        if m:
            detalles['deudas'] = int(m.group(1))

        m = _re.search(r'Multas\s*:\s*(\d+)', texto_completo)
        if m:
            detalles['multas'] = int(m.group(1))

        detalles['instituciones'] = _extraer_bloques_institucion(texto_pag1)

        return detalles
    except Exception:
        return detalles


def generar_excel_analisis(datos_por_pdf, grupos_base_madre):
    """Genera Excel con dos hojas a partir de datos de PDFs.

    datos_por_pdf: list de dicts retornados por extraer_datos_pdf()
    grupos_base_madre: dict {rut: {'grupo': '...', 'razon_social': '...'}}

    Retorna bytes del Excel o None si error.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        wb.remove(wb.active)

        # Sheet 1: Resumen — una fila por cada institución con deuda/multa;
        # si la empresa no tiene ninguna, una única fila "Sin Dicom".
        ws1 = wb.create_sheet('Resumen Analisis', 0)
        ws1.append(['RUT Empresa', 'GRUPO', 'Empresa', 'Tiene DICOM', 'Institucion',
                    'Tipo Correo', 'Tipo de Correo (Grupo)'])

        # Sheet 2: Deuda Previsional — una fila por cada trabajador/periodo
        ws2 = wb.create_sheet('Deuda Previsional', 1)
        ws2.append(['Grupo', 'Rut Emp', 'Empresa', 'Rut Trabj', 'Nombre Trabj',
                   'Institucion', 'Periodo', 'Monto nominal', 'Analisis',
                   'Solicitud documentos', 'Motivo', 'Gestion'])

        # "Tipo de Correo (Grupo)": a diferencia de "Tipo Correo" (que es por
        # institución/fila), esta se calcula sobre TODO el grupo y se repite
        # en todas sus filas. Prioridad de mayor a menor: si CUALQUIER razón
        # social del grupo tiene "Previsional", el grupo entero pasa a
        # "Previsional" (aunque las demás no); si ninguna tiene Previsional
        # pero alguna tiene "DT", el grupo pasa a "DT"; solo si el grupo
        # completo es "Sin Dicom" queda como "Sin Dicom".
        _PRIORIDAD_TIPO_CORREO = {'Sin Dicom': 1, 'DT': 2, 'Previsional': 3}
        filas_ws1 = []
        tipo_correo_grupo = {}

        for datos in datos_por_pdf:
            if not datos.get('rut'):
                continue

            rut_clean = datos['rut'].replace('.', '').replace(' ', '')
            grupo_info = grupos_base_madre.get(rut_clean, {})
            grupo = grupo_info.get('grupo', 'SIN GRUPO')
            empresa = datos.get('razon_social', '')
            instituciones = datos.get('instituciones', [])

            tiene_dicom = 'Si' if (datos.get('deudas', 0) > 0 or
                                   datos.get('multas', 0) > 0 or
                                   datos.get('monto_utm', 0) > 0) else 'No'

            if tiene_dicom == 'No' or not instituciones:
                filas_ws1.append((datos['rut'], grupo, empresa, 'No', 'Sin Dicom', 'Sin Dicom'))
                tipo_correo_grupo.setdefault(grupo, 'Sin Dicom')
                continue

            for bloque in instituciones:
                filas_ws1.append((datos['rut'], grupo, empresa, 'Si',
                                   bloque['institucion'], bloque['tipo_correo']))
                actual = tipo_correo_grupo.get(grupo, 'Sin Dicom')
                if _PRIORIDAD_TIPO_CORREO[bloque['tipo_correo']] > _PRIORIDAD_TIPO_CORREO[actual]:
                    tipo_correo_grupo[grupo] = bloque['tipo_correo']

                for trab in bloque.get('trabajadores', []):
                    ws2.append([
                        grupo, datos['rut'], empresa,
                        trab['rut'], trab['nombre'], bloque['institucion'],
                        trab['periodo'], trab['monto'],
                        '', '', '', ''  # Analisis/Solicitud/Motivo/Gestion: consultor
                    ])

        for rut, grupo, empresa, tiene_dicom, institucion, tipo_correo in filas_ws1:
            ws1.append([rut, grupo, empresa, tiene_dicom, institucion, tipo_correo,
                        tipo_correo_grupo.get(grupo, 'Sin Dicom')])

        # Auto-adjust column widths
        for ws in [ws1, ws2]:
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_len:
                            max_len = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        # Retornar bytes
        import io
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()
    except Exception as e:
        return None


def probar_login(correo, clave, log, ruta_captura=None, obtener_codigo=None, ruts=None):
    """Entra al portal, confirma el acceso y guarda una captura de pantalla.

    Equifax detecta navegadores automatizados: en modo invisible (headless)
    el acceso rebota sin dar error. Por eso se levanta una pantalla virtual
    (xvfb) y el navegador corre CON pantalla, igual que en el módulo Mi DT.
    """
    display = None
    headless = True
    try:
        from pyvirtualdisplay import Display
        display = Display(visible=0, size=(1440, 1000))
        display.start()
        headless = False
        log("Pantalla virtual iniciada — navegador con pantalla", "info")
    except Exception as e:
        log(f"Sin pantalla virtual ({str(e)[:40]}); se usará modo invisible", "warn")

    try:
      with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=headless,
            args=["--no-sandbox", "--disable-dev-shm-usage",
                  "--disable-blink-features=AutomationControlled",
                  "--disable-features=IsolateOrigins,site-per-process"],
        )
        ctx = browser.new_context(
            viewport={"width": 1400, "height": 900},
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                        "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"),
            locale="es-CL",
            timezone_id="America/Santiago",
        )
        # Ocultar las señales típicas de automatización
        ctx.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            Object.defineProperty(navigator, 'languages', {get: () => ['es-CL','es','en']});
            Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3,4,5]});
            window.chrome = window.chrome || {runtime: {}};
        """)
        page = ctx.new_page()
        page.set_default_timeout(45000)

        def _captura(etiqueta=""):
            if not ruta_captura:
                return
            try:
                os.makedirs(os.path.dirname(ruta_captura), exist_ok=True)
                page.screenshot(path=ruta_captura, full_page=False)
                log(f"Captura de pantalla guardada{etiqueta}", "ok")
            except Exception as e:
                log(f"No se pudo guardar la captura: {type(e).__name__}", "warn")

        try:
            try:
                hacer_login(page, correo, clave, log, obtener_codigo=obtener_codigo)
            except LoginFallido:
                # Guardar la captura del FALLO para poder ver qué muestra el portal
                _captura(" (pantalla del error)")
                _dump_campos(page, log)
                raise

            # Click en botón "Ingresar" para acceder a Reportes Interactivos
            try:
                btn = page.locator('[data-test-id="appTypeButton"], button:has-text("Ingresar")').first
                if btn.count() > 0:
                    btn.click()
                    page.wait_for_load_state('domcontentloaded', timeout=15000)
                    log("✓ Click en botón 'Ingresar'", "ok")
                else:
                    log("⚠ No se encontró botón 'Ingresar'", "warn")
            except Exception as e:
                log(f"⚠ Error al hacer click en 'Ingresar': {e}", "warn")

            # Dejar constancia de dónde quedó y qué se ve
            info = {"url": page.url, "titulo": ""}
            try:
                info["titulo"] = page.title()
            except Exception:
                pass
            log(f"Página tras el acceso: {info['titulo']}", "info")
            resumen = _texto(page, 300)
            if resumen:
                log(f"Contenido: {resumen}", "info")
            _captura()
            return info
        finally:
            try:
                ctx.close()
            except Exception:
                pass
            browser.close()
    finally:
        if display:
            try:
                display.stop()
            except Exception:
                pass


# ──────────────────────────────────────────────────────────────────────────
# Búsqueda y descarga real de Boletines Laborales (flujo completo, mapeado
# a mano navegando el portal real: Productos -> Ingresar (pestaña nueva
# "Interactive Reports") -> "+" -> Nuevo Reporte -> Producto=Boletin Laboral
# -> Rut -> Generar Reporte -> "+" -> Generar PDF (descarga directa) ->
# "+" -> Nuevo Reporte (siguiente RUT).
# ──────────────────────────────────────────────────────────────────────────

def _abrir_reportes_interactivos(page, log):
    """Click en 'Ingresar' (tarjeta Reportes Interactivos) — abre una
    PESTAÑA NUEVA (Interactive Reports); hay que capturarla con
    expect_page, igual que el mismo patrón ya resuelto en Previred.

    La pestaña nueva primero pasa por una URL de relay SSO
    (.../ir/sso/relay?accountId=...) que carga en blanco mientras hace la
    redirección real hacia /ir/report (la Angular SPA recién bootstrapea
    ahí). "domcontentloaded" se cumple con esa página en blanco, mucho
    antes de que el formulario real exista -- por eso hay que esperar
    explícitamente a que aparezca contenido real (el texto "Producto"),
    con un margen generoso, en vez de asumir que ya está listo."""
    try:
        with page.context.expect_page(timeout=15000) as popup_info:
            btn = page.locator(
                '[data-test-id="appTypeButton"], button:has-text("Ingresar")').first
            btn.click(timeout=8000)
        reportes = popup_info.value
        reportes.set_default_timeout(45000)
        log(f"Reportes Interactivos abriendo (relay SSO)...", "info")

        try:
            reportes.get_by_text("Producto", exact=False).first.wait_for(
                state="visible", timeout=60000)
        except Exception:
            # Respaldo: si no aparece "Producto" pero sí "Nuevo Reporte",
            # igual seguimos -- el formulario podría estar cerrado y haya
            # que abrirlo con el "+" más adelante.
            try:
                reportes.get_by_text("Nuevo Reporte", exact=False).first.wait_for(
                    state="visible", timeout=10000)
            except Exception:
                pass

        log(f"Reportes Interactivos abierto: {reportes.url}", "ok")
        return reportes
    except Exception as e:
        raise LoginFallido(
            f"No se pudo abrir 'Reportes Interactivos': {type(e).__name__}: {e}")


def _dump_reportes_ui(page, log, ruta_captura=None):
    """Diagnóstico cuando algo del portal 'Interactive Reports' no aparece
    donde se esperaba: vuelca los botones visibles + guarda una captura,
    en vez de seguir adivinando el selector a ciegas."""
    try:
        info = page.evaluate("""() => {
            const out = [];
            for (const e of document.querySelectorAll('button, a[role="button"]')) {
                if (e.offsetParent === null) continue;
                const icon = e.querySelector('mat-icon');
                out.push({
                    clase: (e.className || '').toString().slice(0, 80),
                    texto: (e.innerText || '').trim().slice(0, 30),
                    icono: icon ? (icon.innerText || icon.textContent || '').trim() : ''
                });
            }
            return out.slice(0, 20);
        }""")
        log(f"  [debug] url = {page.url}", "warn")
        log(f"  [debug] botones visibles ({len(info)}):", "warn")
        for b in info:
            log(f"    clase={b['clase']!r} texto={b['texto']!r} icono={b['icono']!r}", "warn")
    except Exception as e:
        log(f"  [debug] no se pudo volcar la UI: {type(e).__name__}", "warn")
    if ruta_captura:
        try:
            page.screenshot(path=ruta_captura)
            nombre_cap = os.path.basename(ruta_captura)
            log(f"Captura de pantalla: "
                f"<a href='/api/dicom/captura/{nombre_cap}' target='_blank'>ver imagen</a>",
                "warn")
        except Exception:
            pass


def _menu_fab_abierto(page):
    """True solo si el menú del '+' está REALMENTE desplegado en pantalla.

    Confirmado con un dump real (12/08 15:41): los mini-botones NO tienen
    ningún texto legible "Generar PDF" / "Nuevo Reporte" -- su único
    contenido de texto es el nombre técnico del ícono Material
    ('picture_as_pdf', 'note_add', 'contact_page'), igual estén el menú
    abierto o cerrado (existen en el DOM todo el tiempo). Por eso se
    identifican por ícono, y se confirma que están REALMENTE desplegados
    mirando su tamaño/opacidad/pointer-events en pantalla (con el menú
    cerrado suelen quedar con transform:scale(0), que sí reduce su
    getBoundingClientRect a (casi) cero, aunque sigan en el DOM)."""
    try:
        return bool(page.evaluate("""() => {
            for (const b of document.querySelectorAll('button')) {
                const icon = b.querySelector('mat-icon');
                const t = icon ? (icon.innerText || icon.textContent || '').trim() : '';
                if (t === 'picture_as_pdf' || t === 'note_add' || t === 'contact_page') {
                    const r = b.getBoundingClientRect();
                    const cs = getComputedStyle(b);
                    if (r.width > 5 && r.height > 5 &&
                        parseFloat(cs.opacity) > 0.5 &&
                        cs.visibility !== 'hidden' &&
                        cs.pointerEvents !== 'none') {
                        return true;
                    }
                }
            }
            return false;
        }"""))
    except Exception:
        return False


def _click_fab(page, log, ruta_captura=None):
    """Click en el botón flotante rojo '+' que despliega Generar PDF /
    Nuevo Reporte / Agregar seguimiento.

    Confirmado con un dump real de botones: el "+" principal es el ÚNICO
    con la clase 'fab-main-btn' (ícono 'add'); los otros 3 mini-botones
    del menú (Agregar seguimiento/contact_page, Nuevo Reporte/note_add,
    Generar PDF/picture_as_pdf) YA están en el DOM y visibles ANTES de
    abrir el menú, así que un selector combinado con ':has(mat-icon:...)'
    puede matchear cualquiera de ellos primero -- hay que apuntar
    exclusivamente a 'fab-main-btn' para no clickear el mini-botón
    equivocado por error.

    Además, un click() de Playwright que NO tira error no garantiza que el
    menú se haya abierto de verdad (confirmado con captura real: el menú
    seguía cerrado tras el click). Por eso se verifica el estado real en
    pantalla después de cada click y se reintenta hasta 3 veces."""
    fab = page.locator("button.fab-main-btn").first
    try:
        fab.wait_for(state="visible", timeout=45000)
    except Exception:
        _dump_reportes_ui(page, log, ruta_captura)
        raise

    for intento in range(1, 4):
        fab.click(timeout=8000)
        page.wait_for_timeout(500)
        if _menu_fab_abierto(page):
            if intento > 1:
                log(f"  menú del '+' abierto (intento {intento})", "info")
            return
        log(f"  intento {intento}/3: el menú del '+' no se desplegó, reintentando...", "warn")

    log("  el menú del '+' no se desplegó tras 3 intentos", "warn")
    _dump_reportes_ui(page, log, ruta_captura)


def _seleccionar_boletin_laboral(page, log):
    """Abre el desplegable 'Producto' (mat-select) y elige 'Boletin
    Laboral' de la lista de opciones."""
    page.get_by_label("Producto", exact=False).click(timeout=10000)
    page.get_by_role("option", name="Boletin Laboral", exact=True).click(timeout=8000)


def _generar_reporte(page, rut, log):
    """Escribe el RUT en el formulario y clickea 'Generar Reporte'; espera
    a que el overlay 'Cargando...' aparezca y luego desaparezca."""
    campo = page.get_by_label("Rut", exact=False)
    campo.click(timeout=8000)
    try:
        campo.fill("")
    except Exception:
        pass
    campo.type(rut, delay=40)
    page.get_by_role("button", name="Generar Reporte").click(timeout=8000)
    try:
        page.get_by_text("Cargando", exact=False).wait_for(state="visible", timeout=4000)
    except Exception:
        pass
    page.get_by_text("Cargando", exact=False).wait_for(state="hidden", timeout=60000)


def _descargar_pdf_reporte(page, ruta_dest, log, ruta_captura=None):
    """Click en '+' -> 'Generar PDF' y captura el archivo.

    Confirmado con las DevTools reales (pestaña Network): el PDF NO llega
    por una descarga nativa del navegador -- se pide con una petición
    XHR (fetch) de 200 OK y ~19KB, y recién ahí el JS del portal arma el
    archivo del lado del cliente (blob). Por eso el evento 'download' de
    Playwright nunca disparaba: no hay ninguna navegación/descarga real
    que capturar. En vez de depender de ese evento, se intercepta
    DIRECTAMENTE la respuesta de red que trae los bytes del PDF (mismo
    patrón ya usado con éxito en Previred y en Cartas Previsionales) --
    se guarda el 'download' como respaldo adicional por si en algún caso
    sí dispara uno.
    """
    capturado = []

    def _en_respuesta(response):
        try:
            ct = (response.headers or {}).get("content-type", "")
            if "pdf" in ct.lower():
                body = response.body()
                if body and len(body) > 500:
                    capturado.append(body)
                    log(f"PDF capturado por red ({len(body)} bytes)", "ok")
                    return
            # Diagnóstico: mientras no aparece el PDF por ningún método, se
            # deja un rastro de las respuestas del propio portal (sin
            # archivos estáticos) para saber qué pasa de verdad por la red
            # sin depender de abrir DevTools a mano en cada prueba.
            url = response.url
            if ("equifax" in url) and not url.endswith(
                    (".js", ".css", ".woff", ".woff2", ".svg", ".png", ".ico", ".jpg")):
                log(f"  [red] {response.status} {ct or '(sin content-type)'} {url[:110]}", "info")
                # Confirmado con DevTools reales: el endpoint del PDF es
                # GET .../reports/{id}/pdf. Si ESTA MISMA URL responde con
                # JSON en vez del PDF (como pasó una vez, con content-type
                # application/json), es casi seguro un error del servidor
                # (no listo aún, no autorizado, etc.) -- se imprime el
                # cuerpo para saber la razón exacta en vez de adivinar.
                if "/pdf" in url and "json" in ct.lower():
                    try:
                        texto = response.text()
                        log(f"  [red] cuerpo de esa respuesta JSON: {texto[:500]}", "warn")
                    except Exception:
                        pass
        except Exception:
            pass

    def _en_descarga(dl):
        try:
            import tempfile as _tf
            tmp = _tf.mktemp(suffix=".pdf")
            dl.save_as(tmp)
            with open(tmp, "rb") as fh:
                b = fh.read()
            os.remove(tmp)
            if b and len(b) > 500:
                capturado.append(b)
                log(f"PDF capturado por descarga ({len(b)} bytes)", "ok")
        except Exception:
            pass

    # Respaldo adicional: además de los listeners de red/descarga (que hasta
    # ahora no dispararon ni una vez, con motivo aún no confirmado), el
    # navegador queda configurado (ver descargar_boletines) para guardar
    # cualquier descarga DIRECTO en esta misma carpeta vía CDP
    # (Browser.setDownloadBehavior) -- así que también se vigila la carpeta
    # por si aparece el archivo ahí, sin depender de que Playwright "vea"
    # el evento.
    carpeta_dest = os.path.dirname(ruta_dest)
    try:
        ya_existian = set(os.listdir(carpeta_dest))
    except Exception:
        ya_existian = set()

    page.context.on("response", _en_respuesta)
    page.context.on("download", _en_descarga)
    try:
        _click_fab(page, log, ruta_captura)
        try:
            # El botón "Generar PDF" NO tiene texto legible propio (confirmado
            # con dump real): solo el ícono Material 'picture_as_pdf'. Buscar
            # por get_by_text("Generar PDF") no encuentra el botón real y
            # puede "clickear" con éxito algo que no hace nada (ej. un
            # tooltip fantasma), sin avisar del error.

            # "Sensor" puesto en el botón exacto ANTES del click: un
            # screenshot puede no mostrar ningún cambio visible aunque el
            # click sí haya llegado (si la reacción es solo una petición de
            # red). Esto en cambio confirma, sin depender de lo visual, si
            # el evento click llegó de verdad al botón correcto.
            try:
                page.evaluate("""() => {
                    const btns = Array.from(document.querySelectorAll('button'));
                    const btn = btns.find(b => {
                        const icon = b.querySelector('mat-icon');
                        const t = icon ? (icon.innerText || icon.textContent || '').trim() : '';
                        return t === 'picture_as_pdf';
                    });
                    window.__dicomClickDetectado = false;
                    if (btn) {
                        btn.addEventListener('click', () => { window.__dicomClickDetectado = true; },
                                              {once: true});
                    }
                }""")
            except Exception:
                pass

            page.locator("button:has(mat-icon:text-is('picture_as_pdf'))").first.click(timeout=8000)

            try:
                detectado = page.evaluate("() => window.__dicomClickDetectado === true")
            except Exception:
                detectado = None
            log(f"  click en 'Generar PDF': Playwright OK, ¿el botón recibió el evento? {detectado}",
                "info")
        except Exception:
            _dump_reportes_ui(page, log, ruta_captura)
            raise

        # Captura INMEDIATA (sin esperar nada) justo después del click, para
        # ver en el momento si pasó algo visible (loader, cambio de estado)
        # o si la pantalla queda exactamente igual que antes -- evidencia
        # directa sin tener que esperar los 150s completos.
        if ruta_captura:
            try:
                base, ext = os.path.splitext(ruta_captura)
                ruta_post_click = f"{base}_post_click{ext}"
                os.makedirs(os.path.dirname(ruta_post_click), exist_ok=True)
                page.screenshot(path=ruta_post_click)
                nombre_cap = os.path.basename(ruta_post_click)
                log(f"Captura justo después del click en 'Generar PDF': "
                    f"<a href='/api/dicom/captura/{nombre_cap}' target='_blank'>ver imagen</a>",
                    "info")
            except Exception:
                pass

        ruta_archivo_nuevo = None
        inicio = time.time()
        ultimo_aviso = inicio
        # Confirmado con una respuesta real de red (JSON de
        # .../reports/{id}) que llegó recién a los ~90s: el backend de
        # Equifax tarda bastante en preparar el reporte antes del PDF, así
        # que 90s se quedaba corto justo cuando recién empezaba a responder.
        fin = inicio + 150
        while time.time() < fin:
            if capturado:
                break
            try:
                actuales = set(os.listdir(carpeta_dest))
            except Exception:
                actuales = set()
            nuevos = [f for f in actuales - ya_existian
                      if not f.endswith((".crdownload", ".tmp"))]
            if nuevos:
                ruta_archivo_nuevo = os.path.join(carpeta_dest, nuevos[0])
                break
            if time.time() - ultimo_aviso > 15:
                ultimo_aviso = time.time()
                log(f"  esperando el PDF... ({int(time.time() - inicio)}s)", "info")
            time.sleep(1)

        if ruta_archivo_nuevo:
            # Esperar a que termine de escribirse (tamaño estable) antes de moverlo
            tam_previo = -1
            for _ in range(10):
                try:
                    tam = os.path.getsize(ruta_archivo_nuevo)
                except Exception:
                    tam = -1
                if tam == tam_previo and tam > 0:
                    break
                tam_previo = tam
                time.sleep(0.5)
            log(f"PDF capturado por archivo en disco ({os.path.basename(ruta_archivo_nuevo)})", "ok")
            if os.path.abspath(ruta_archivo_nuevo) != os.path.abspath(ruta_dest):
                os.replace(ruta_archivo_nuevo, ruta_dest)
            return

        if not capturado:
            _dump_reportes_ui(page, log, ruta_captura)
            raise TimeoutError(
                "No se detectó ningún PDF (ni por red, ni por descarga, ni en el disco) en 150s")

        with open(ruta_dest, "wb") as f:
            f.write(capturado[0])
    finally:
        try:
            page.context.remove_listener("response", _en_respuesta)
        except Exception:
            pass
        try:
            page.context.remove_listener("download", _en_descarga)
        except Exception:
            pass


def _cerrar_fab_si_abierto(page):
    """Si el menú del '+' quedó abierto (ej. por un intento anterior que
    falló), lo cierra con Escape para no dejar botones duplicados en
    pantalla (causa de 'resolved to 2 elements' en el siguiente intento)."""
    try:
        page.keyboard.press("Escape")
        page.wait_for_timeout(300)
    except Exception:
        pass


def _nuevo_reporte(page, log, ruta_captura=None):
    """Click en '+' -> 'Nuevo Reporte' para volver al formulario de
    búsqueda y pedir el siguiente RUT."""
    _cerrar_fab_si_abierto(page)
    _click_fab(page, log, ruta_captura)
    try:
        # Mismo caso que "Generar PDF": el botón solo tiene el ícono
        # Material 'note_add', sin texto "Nuevo Reporte" legible.
        page.locator("button:has(mat-icon:text-is('note_add'))").first.click(timeout=8000)
    except Exception:
        _dump_reportes_ui(page, log, ruta_captura)
        raise
    page.wait_for_timeout(500)


def descargar_boletines(correo, clave, ruts, carpeta_dest, log,
                        ruta_captura=None, obtener_codigo=None, debe_cancelar=None):
    """Login completo (incluyendo código de verificación) + descarga el
    Boletin Laboral de cada RUT en `ruts`, uno por uno, guardando cada PDF
    en `carpeta_dest`. Devuelve {"descargados": int, "errores": [rut,...]}.
    """
    display = None
    headless = True
    try:
        from pyvirtualdisplay import Display
        display = Display(visible=0, size=(1440, 1000))
        display.start()
        headless = False
        log("Pantalla virtual iniciada — navegador con pantalla", "info")
    except Exception as e:
        log(f"Sin pantalla virtual ({str(e)[:40]}); se usará modo invisible", "warn")

    os.makedirs(carpeta_dest, exist_ok=True)
    descargados = 0
    errores = []

    try:
      with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=headless,
            args=["--no-sandbox", "--disable-dev-shm-usage",
                  "--disable-blink-features=AutomationControlled",
                  "--disable-features=IsolateOrigins,site-per-process"],
        )
        ctx = browser.new_context(
            viewport={"width": 1400, "height": 900},
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                        "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"),
            locale="es-CL",
            timezone_id="America/Santiago",
            accept_downloads=True,
        )
        ctx.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            Object.defineProperty(navigator, 'languages', {get: () => ['es-CL','es','en']});
            Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3,4,5]});
            window.chrome = window.chrome || {runtime: {}};
        """)

        # El bot corre en el servidor, no en el computador de quien lo usa --
        # no existe una "carpeta de Descargas del usuario" a la que este
        # Chrome pueda escribir. En vez de eso se le indica al navegador, por
        # CDP (protocolo, no evento de Playwright), que guarde cualquier
        # descarga DIRECTO en `carpeta_dest`: así el archivo queda listo ahí
        # mismo para empaquetarlo en el ZIP que se le ofrece a la persona al
        # final, sin depender de que el evento 'download' de Playwright
        # dispare (no lo ha hecho hasta ahora, por un motivo aún sin
        # confirmar).
        try:
            cdp = browser.new_browser_cdp_session()
            cdp.send("Browser.setDownloadBehavior", {
                "behavior": "allow",
                "downloadPath": os.path.abspath(carpeta_dest),
            })
            log("Descargas del navegador redirigidas a la carpeta del servidor", "info")
        except Exception as e:
            log(f"No se pudo forzar la carpeta de descargas: {type(e).__name__}: {e}", "warn")

        page = ctx.new_page()
        page.set_default_timeout(45000)

        def _captura(etiqueta=""):
            if not ruta_captura:
                return
            try:
                os.makedirs(os.path.dirname(ruta_captura), exist_ok=True)
                page.screenshot(path=ruta_captura, full_page=False)
            except Exception:
                pass

        try:
            try:
                hacer_login(page, correo, clave, log, obtener_codigo=obtener_codigo)
            except LoginFallido:
                _captura(" (pantalla del error)")
                _dump_campos(page, log)
                raise

            reportes = _abrir_reportes_interactivos(page, log)

            log(f"__PROGRESO_TOTAL__:{len(ruts)}", "info")

            primero = True
            for idx, rut in enumerate(ruts, 1):
                if debe_cancelar is not None and debe_cancelar():
                    log("Detenido por el usuario", "warn")
                    break

                log(f"── RUT {idx}/{len(ruts)}: {rut}", "info")
                log(f"__PROGRESO_AVANCE__:{idx}", "info")
                rut_limpio = rut.replace(".", "").replace(" ", "")
                try:
                    if not primero:
                        _nuevo_reporte(reportes, log, ruta_captura)
                    primero = False

                    _seleccionar_boletin_laboral(reportes, log)
                    _generar_reporte(reportes, rut, log)

                    nombre = f"Boletin_Laboral_{rut_limpio}.pdf"
                    ruta = os.path.join(carpeta_dest, nombre)
                    _descargar_pdf_reporte(reportes, ruta, log, ruta_captura)

                    log(f"Guardado: {nombre}", "ok")
                    log(f"__ARCHIVO_OK__:{rut}:{nombre}", "ok")
                    descargados += 1
                except Exception as e:
                    motivo = f"{type(e).__name__}: {str(e)[:150]}"
                    log(f"✗ RUT {rut}: {motivo}", "err")
                    log(f"__ARCHIVO_ERROR__:{rut}:{motivo}", "err")
                    errores.append(rut)
                    try:
                        reportes.screenshot(
                            path=os.path.join(carpeta_dest, f"_error_{rut_limpio}.png"))
                    except Exception:
                        pass

            log(f"__RESUMEN__:{descargados}:{len(errores)}", "ok")
            log(f"Descarga completada: {descargados}/{len(ruts)} boletines "
                f"({len(errores)} con error)", "ok")
            return {"descargados": descargados, "errores": errores}
        finally:
            try:
                ctx.close()
            except Exception:
                pass
            browser.close()
    finally:
        if display:
            try:
                display.stop()
            except Exception:
                pass
