"""Bot para consulta masiva de RUTs en Mi DT (midt.dirtrab.cl)."""

import io, time, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

_URL = "https://midt.dirtrab.cl"


# ── Login ──────────────────────────────────────────────────────────────────────

def _tipo_campo(page, locator, valor):
    """Escribe en un campo simulando teclas reales para activar la validación JS."""
    locator.click()
    locator.press("Control+a")
    locator.press("Backspace")
    try:
        locator.press_sequentially(valor, delay=60)
    except AttributeError:
        # Playwright < 1.38 no tiene press_sequentially
        locator.type(valor, delay=60)
    locator.dispatch_event("input")
    locator.dispatch_event("change")
    locator.press("Tab")


def hacer_login(page, run, clave, log):
    log("Abriendo Mi DT...", "info")
    page.goto(_URL, wait_until="networkidle", timeout=45000)

    # Click en botón de login principal
    btn_login = page.locator(
        "a:has-text('Iniciar sesión'), button:has-text('Iniciar sesión'), "
        "a:has-text('Ingresar'), button:has-text('Ingresar'), "
        "a[href*='login'], a[href*='clave']"
    ).first
    try:
        with page.expect_navigation(wait_until="domcontentloaded", timeout=35000):
            btn_login.click(timeout=10000)
    except PWTimeout:
        page.wait_for_load_state("domcontentloaded", timeout=10000)

    # Esperar ClaveÚnica
    if "claveunica" not in page.url.lower():
        page.wait_for_url("*claveunica*", timeout=30000)
        page.wait_for_load_state("domcontentloaded", timeout=20000)

    log("Ingresando ClaveÚnica...", "info")

    # ClaveÚnica puede ser de un paso (RUN+clave juntos) o dos pasos (RUN → Continuar → clave).
    # Detectamos cuál es verificando si hay campo de contraseña visible al inicio.
    run_field = page.locator("#uname, input[name='run'], input[type='text']").first
    run_field.wait_for(state="visible", timeout=15000)

    pw_visible = False
    try:
        pw_loc = page.locator("#pword, input[type='password']").first
        pw_visible = pw_loc.is_visible()
    except Exception:
        pass

    # Escribir RUN con teclas reales (dispara keyup/input requeridos por ClaveÚnica)
    _tipo_campo(page, run_field, run)
    time.sleep(0.5)

    if not pw_visible:
        # Flujo de dos pasos: primero RUN → click "Continuar"
        try:
            continuar = page.locator("button:has-text('Continuar'), button:has-text('CONTINUAR')").first
            continuar.wait_for(state="visible", timeout=8000)
            continuar.click()
            page.wait_for_load_state("domcontentloaded", timeout=15000)
            time.sleep(1)
        except PWTimeout:
            pass

    # Escribir contraseña
    pw_field = page.locator("#pword, input[type='password']").first
    pw_field.wait_for(state="visible", timeout=15000)
    _tipo_campo(page, pw_field, clave)
    time.sleep(0.8)

    # Esperar que el botón INGRESA se habilite (hasta 10s)
    ingresar = page.locator("#login-submit, button:has-text('INGRESA'), button:has-text('Ingresar')").first
    try:
        page.wait_for_function(
            "() => { const b = document.querySelector('#login-submit, button[type=submit]'); "
            "return b && !b.disabled; }",
            timeout=10000
        )
    except PWTimeout:
        pass

    log("Enviando credenciales...", "info")
    try:
        with page.expect_navigation(wait_until="domcontentloaded", timeout=40000):
            ingresar.click(timeout=8000)
    except PWTimeout:
        ingresar.click(timeout=8000, force=True)
        page.wait_for_load_state("domcontentloaded", timeout=20000)

    # Esperar redirección de vuelta a Mi DT
    if "midt.dirtrab" not in page.url.lower():
        page.wait_for_url("*midt.dirtrab*", timeout=35000)
        page.wait_for_load_state("domcontentloaded", timeout=20000)

    log("Login exitoso", "ok")


# ── Selección empresa ──────────────────────────────────────────────────────────

# JS: hace click en el elemento cuyo texto es EXACTAMENTE `texto` (o empieza por él),
# eligiendo el nodo más pequeño y subiendo a su ancestro clickeable. Devuelve el
# texto clickeado o null.
_JS_CLICK_EXACTO = """
(texto) => {
    const objetivo = texto.trim().toUpperCase();
    let best = null;
    for (const el of document.querySelectorAll('button, a, div, span, li, [role=button], .card, .mat-card')) {
        const t = (el.innerText || '').trim().toUpperCase();
        if (t === objetivo || t.startsWith(objetivo + ' ') || t === objetivo.replace(/\\s+/g,'')) {
            if (!best || (el.innerText || '').length < (best.innerText || '').length) best = el;
        }
    }
    if (!best) return null;
    let target = best;
    for (let i = 0; i < 5 && target; i++) {
        const cs = getComputedStyle(target);
        if (target.tagName === 'BUTTON' || target.tagName === 'A' ||
            target.getAttribute('role') === 'button' || target.onclick ||
            cs.cursor === 'pointer') {
            target.click();
            return best.innerText.trim();
        }
        target = target.parentElement;
    }
    best.click();
    return best.innerText.trim();
}
"""

# JS: hace click en el elemento más pequeño cuyo texto CONTIENE `sub` (subcadena).
_JS_CLICK_CONTIENE = """
(sub) => {
    const objetivo = sub.trim().toUpperCase();
    let best = null;
    for (const el of document.querySelectorAll('button, a, div, span, li, [role=button], .card, .mat-card')) {
        const t = (el.innerText || '').trim().toUpperCase();
        if (t.includes(objetivo)) {
            if (!best || (el.innerText || '').length < (best.innerText || '').length) best = el;
        }
    }
    if (!best) return null;
    let target = best;
    for (let i = 0; i < 5 && target; i++) {
        const cs = getComputedStyle(target);
        if (target.tagName === 'BUTTON' || target.tagName === 'A' ||
            target.getAttribute('role') === 'button' || target.onclick ||
            cs.cursor === 'pointer') {
            target.click();
            return best.innerText.trim();
        }
        target = target.parentElement;
    }
    best.click();
    return best.innerText.trim();
}
"""

# JS: hace click en el elemento que CONTIENE el RUT dado (empresa en la lista).
_JS_CLICK_RUT = """
(rut) => {
    const rutClean = rut.replace(/[\\.\\-]/g, '').toUpperCase();
    let best = null;
    for (const el of document.querySelectorAll('button, a, div, li, span, .card, .mat-card, [role=button]')) {
        const txt = (el.innerText || '').replace(/[\\.\\-]/g, '').toUpperCase();
        if (txt.includes(rutClean)) {
            if (!best || (el.innerText || '').length < (best.innerText || '').length) best = el;
        }
    }
    if (!best) return null;
    let target = best;
    for (let i = 0; i < 5 && target; i++) {
        const cs = getComputedStyle(target);
        if (target.tagName === 'BUTTON' || target.tagName === 'A' ||
            target.getAttribute('role') === 'button' || target.onclick ||
            cs.cursor === 'pointer') {
            target.click();
            return best.innerText.trim();
        }
        target = target.parentElement;
    }
    best.click();
    return best.innerText.trim();
}
"""


def _dump_pantalla(page, log, etiqueta):
    """Vuelca al log los elementos clickeables con tag+atributos+texto,
    para poder deducir selectores exactos sin ver el HTML completo."""
    try:
        items = page.evaluate("""() => {
            const out = [];
            const els = document.querySelectorAll(
                'button, a, [role=button], [routerlink], .card, .mat-card, ' +
                'input, [class*=btn], [class*=card], [class*=item]'
            );
            for (const el of els) {
                const t = (el.innerText || el.value || '').trim().replace(/\\s+/g, ' ').slice(0, 45);
                const cs = getComputedStyle(el);
                const clickable = el.tagName === 'BUTTON' || el.tagName === 'A' ||
                    el.onclick || el.getAttribute('role') === 'button' ||
                    el.hasAttribute('routerlink') || cs.cursor === 'pointer';
                if (!t && !clickable) continue;
                let desc = el.tagName.toLowerCase();
                if (el.id) desc += '#' + el.id;
                if (el.className && typeof el.className === 'string')
                    desc += '.' + el.className.trim().split(/\\s+/).slice(0,2).join('.');
                const rl = el.getAttribute('routerlink');
                if (rl) desc += `[routerlink=${rl}]`;
                if (clickable) desc += '*';
                out.push(`${desc} "${t}"`);
            }
            return [...new Set(out)].slice(0, 30);
        }""")
        log(f"[debug] {etiqueta}:", "info")
        for it in items:
            log(f"    {it}", "info")
    except Exception:
        pass


def _click_texto(page, candidatos, log, timeout=15000):
    """Intenta clickear el primer elemento visible que coincida, probando
    role=button → role=link → texto. `candidatos` es lista de textos a probar.
    Usa clicks nativos de Playwright (manejan Angular y esperan actionability).
    Devuelve el texto que funcionó, o None."""
    for texto in candidatos:
        for loc in [
            page.get_by_role("button", name=texto),
            page.get_by_role("link", name=texto),
            page.get_by_text(texto),
        ]:
            try:
                el = loc.first
                el.wait_for(state="visible", timeout=3500)
                el.scroll_into_view_if_needed(timeout=3000)
                el.click(timeout=timeout)
                return texto
            except Exception:
                continue
    return None


def seleccionar_empresa(page, rut_empresa, log, snap=None):
    """Selecciona perfil EMPLEADOR → Empleador Persona Jurídica → empresa por RUT.
    Sin pausas fijas: cada paso espera su condición real y avanza apenas se cumple."""
    # ── Paso 1: click en el perfil EMPLEADOR ───────────────────────────────────
    log("Seleccionando perfil EMPLEADOR...", "info")
    emp = _click_texto(page, ["EMPLEADOR", "Empleador"], log)
    if not emp:
        _dump_pantalla(page, log, "Perfiles disponibles")
        raise Exception("No se encontró el perfil EMPLEADOR en la pantalla de roles")

    # Esperar a que aparezca la pantalla "Indica qué tipo de empleador"
    try:
        page.get_by_text("Persona Jurídica").first.wait_for(state="visible", timeout=20000)
    except PWTimeout:
        pass

    # ── Paso 2: click en "Empleador Persona Jurídica" ──────────────────────────
    log("Seleccionando Empleador Persona Jurídica...", "info")
    pj = _click_texto(page, ["Empleador Persona Jurídica", "Persona Jurídica"], log)
    if not pj:
        _dump_pantalla(page, log, "Opciones tras EMPLEADOR")
        raise Exception("No se encontró 'Empleador Persona Jurídica' tras elegir EMPLEADOR")

    # Esperar a que aparezca la lista de empresas (que contenga el RUT)
    rut_sin = rut_empresa.replace(".", "").replace("-", "")
    rut_con = rut_empresa
    try:
        page.get_by_text(rut_sin[:6]).first.wait_for(state="visible", timeout=20000)
    except PWTimeout:
        pass

    # ── Paso 3: click en la empresa por RUT ────────────────────────────────────
    log(f"Seleccionando empresa {rut_empresa}...", "info")
    empresa = _click_texto(page, [rut_sin, rut_con, rut_empresa.replace("-", "")], log)
    if not empresa:
        _dump_pantalla(page, log, "Empresas tras Persona Jurídica")
        raise Exception(f"No se encontró la empresa {rut_empresa} tras elegir Persona Jurídica")

    # Esperar el home del empleador (condición real, sin sleep fijo)
    try:
        page.wait_for_url("**/empleador/**", timeout=25000)
    except PWTimeout:
        try:
            page.wait_for_load_state("networkidle", timeout=10000)
        except PWTimeout:
            pass
    log("Empresa seleccionada", "ok")


# ── Navegación al formulario ───────────────────────────────────────────────────

_IFRAME_SRC = "front001-registro-electronico-lab.api.dirtrab.cl"

def _get_iframe_frame(page, espera=20):
    """Devuelve el Frame Python del iframe del formulario DT."""
    for _ in range(espera):
        for fr in page.frames:
            if _IFRAME_SRC in (fr.url or ""):
                return fr
        time.sleep(1)
    return None


def _ir_a_formulario(page, log, snap=None):
    """Navega al formulario de registro de contrato dentro del iframe DT."""
    url_destino = f"{_URL}/empleador/registro-electronico-laboral/registroContratoTrabajo"
    page.goto(url_destino, wait_until="domcontentloaded", timeout=45000)

    log("Esperando iframe del formulario...", "info")

    # El formulario está en un iframe cross-origin. Usar frame_locator para interactuar.
    iframe_loc = page.frame_locator(f"iframe[src*='{_IFRAME_SRC}'], iframe").first

    # Esperar y clickear "Registrar" dentro del iframe
    log("Buscando botón Registrar...", "info")
    try:
        btn = iframe_loc.get_by_role("button", name="Registrar")
        btn.wait_for(state="visible", timeout=20000)
        btn.click()
        log("Click en Registrar", "info")
    except Exception as e:
        # Fallback: buscar por texto
        try:
            btn = iframe_loc.get_by_text("Registrar", exact=True)
            btn.wait_for(state="visible", timeout=10000)
            btn.click()
            log("Click en Registrar (fallback texto)", "info")
        except Exception as e2:
            raise Exception(f"No se pudo clickear 'Registrar' en el iframe: {e} / {e2}")

    # Esperar que los campos del formulario estén HABILITADOS (no solo en DOM)
    # Angular puede tener 32 campos en DOM pero todos disabled durante inicialización.
    log("Esperando formulario de RUT...", "info")
    fr = _get_iframe_frame(page, espera=5)
    for _ in range(30):
        if fr:
            try:
                enabled_n = fr.evaluate("""() => {
                    let n = 0;
                    for (const el of document.querySelectorAll('input, select, textarea')) {
                        if (!el.disabled && !el.readOnly) n++;
                    }
                    return n;
                }""")
                if enabled_n > 5:
                    log(f"Formulario listo ({enabled_n} campos habilitados)", "ok")
                    # Clickear radio "Cédula de identidad" para activar sección trabajador
                    try:
                        iframe_loc = page.frame_locator(f"iframe[src*='{_IFRAME_SRC}'], iframe").first
                        radios = iframe_loc.locator("input[type='radio']")
                        if radios.count() > 0:
                            radios.first.click(force=True)
                            time.sleep(0.5)
                    except Exception:
                        pass
                    return page
            except Exception:
                pass
        time.sleep(1)
        fr = _get_iframe_frame(page, espera=1)

    raise Exception("El formulario del contrato no cargó campos habilitados tras 30 segundos")


def _frame_form(page, espera=20, log=None):
    """Devuelve el Frame (main o iframe) que contiene el formulario, detectado
    por ser el que tiene más inputs. Espera hasta `espera` seg a que carguen."""
    for intento in range(espera):
        best, best_n = None, 0
        for fr in page.frames:
            try:
                n = fr.evaluate(
                    "() => document.querySelectorAll('input, select, textarea').length"
                )
            except Exception:
                n = 0
            if n and n > best_n:
                best_n, best = n, fr
        if best and best_n > 0:
            if log and intento == 0:
                log(f"[debug] Frame del formulario: {best.url[:70]} ({best_n} campos)", "info")
            return best
        time.sleep(1)
    if log:
        # Volcar panorama de frames para diagnóstico
        for fr in page.frames:
            try:
                n = fr.evaluate("() => document.querySelectorAll('input,select,textarea').length")
            except Exception:
                n = -1
            log(f"[debug] frame {fr.url[:60]} → {n} campos", "info")
    return page.main_frame


def _dump_iframe(page, log):
    """Vuelca los campos del formulario (main o iframe)."""
    try:
        fr = _frame_form(page, log=log)
        campos = fr.evaluate("""() => {
            const out = [];
            for (const el of document.querySelectorAll('input, select, textarea')) {
                let label = '';
                if (el.id) {
                    const l = document.querySelector(`label[for="${el.id}"]`);
                    if (l) label = l.innerText.trim();
                }
                if (!label && el.getAttribute('aria-label')) label = el.getAttribute('aria-label');
                if (!label) {
                    const p = el.closest('mat-form-field, .form-group, div');
                    if (p) { const l = p.querySelector('label, mat-label'); if (l) label = l.innerText.trim(); }
                }
                const desc = el.tagName.toLowerCase() +
                    (el.type ? `[type=${el.type}]` : '') +
                    (el.id ? `#${el.id}` : '') +
                    (el.name ? `[name=${el.name}]` : '') +
                    (el.getAttribute('formcontrolname') ? `[fcn=${el.getAttribute('formcontrolname')}]` : '') +
                    (el.placeholder ? `[ph="${el.placeholder}"]` : '');
                out.push(`${desc} label="${label}"`);
            }
            return out.slice(0, 40);
        }""")
        log("[debug] Campos del iframe:", "info")
        for c in campos:
            log(f"    {c}", "info")
    except Exception as e:
        log(f"[debug] No se pudo volcar iframe: {e}", "warn")


# ── Consulta de un RUT ─────────────────────────────────────────────────────────

_JS_EXTRAER = """() => {
    const out = {};
    for (const el of document.querySelectorAll('input, select, textarea')) {
        const id = el.id || '';
        let label = (el.getAttribute('aria-label') || '').trim();
        if (!label && id) {
            const l = document.querySelector('label[for="' + id + '"]');
            if (l) label = l.textContent.trim();
        }
        if (!label) {
            const ff = el.closest('mat-form-field, .form-group, .field');
            if (ff) {
                const ml = ff.querySelector('mat-label, label, .label');
                if (ml) label = ml.textContent.trim();
            }
        }
        if (!label) label = (el.getAttribute('formcontrolname') || el.placeholder || '').trim();
        const val = el.tagName === 'SELECT'
            ? (el.options[el.selectedIndex] ? el.options[el.selectedIndex].text.trim() : '')
            : (el.value || '').trim();
        // Guardar por ID (clave exacta) y por label (fallback)
        if (id) out[id] = val;
        if (label) out[label] = val;
    }
    return out;
}"""


def _extraer_datos_frame(frame):
    """Extrae todos los valores del formulario con sus etiquetas."""
    try:
        return frame.evaluate(_JS_EXTRAER) or {}
    except Exception:
        return {}


_IDS_TRAB = ("nombresTrabajador", "apellidosTrabajador", "emailTrabajador",
             "calleTrabajador", "fechaNacimientoTrabajador")


def _firma_trab(datos):
    """Tupla con los valores clave del trabajador, para detectar cambios entre RUTs."""
    return (datos.get("nombresTrabajador", ""), datos.get("emailTrabajador", ""),
            datos.get("calleTrabajador", ""))


def _esperar_datos(frame, max_seg=10, firma_previa=None):
    """Espera a que los campos del TRABAJADOR tengan valor NUEVO (distinto del RUT anterior)."""
    fin = time.time() + max_seg
    while time.time() < fin:
        datos = _extraer_datos_frame(frame)
        if any(datos.get(k) for k in _IDS_TRAB):
            if firma_previa is None or _firma_trab(datos) != firma_previa:
                return datos
        time.sleep(0.4)
    return _extraer_datos_frame(frame)


def _buscar_campo_rut(frame):
    """Localiza #rutTrabajador (ya habilitado porque _ir_a_formulario esperó campos enabled)."""
    # ID exacto conocido del dump del formulario DT
    for sel in ["#rutTrabajador", "input[id*='rutTrab' i]"]:
        try:
            loc = frame.locator(sel).first
            loc.wait_for(state="enabled", timeout=8000)
            return loc
        except Exception:
            continue

    # Fallback JS: primer input de texto habilitado con 'rut' en el id (excluyendo rutEmpleador)
    try:
        field_id = frame.evaluate("""() => {
            for (const el of document.querySelectorAll('input[type="text"]')) {
                if (el.disabled || el.readOnly) continue;
                const id = (el.id || '').toLowerCase();
                if (id === 'rutempleador' || id === 'rutrepresentanteempleador') continue;
                if (id.includes('rut')) return el.id;
            }
            return null;
        }""")
        if field_id:
            return frame.locator(f"#{field_id}").first
    except Exception:
        pass

    return None


# JS: asegura que el radio "Cédula de identidad" esté marcado (habilita la sección trabajador)
_JS_RADIO_CEDULA = """() => {
    const radios = document.querySelectorAll("input[type=radio]");
    if (!radios.length) return null;
    const r = radios[0];
    const info = {total: radios.length, name: r.name || '', value: r.value || '', checked: r.checked};
    if (!r.checked) {
        r.click();
        r.dispatchEvent(new Event('input',  {bubbles: true}));
        r.dispatchEvent(new Event('change', {bubbles: true}));
        info.checked = r.checked;
    }
    return info;
}"""

# JS: busca el botón de búsqueda POR CERCANÍA a #rutTrabajador (sube por los padres),
# en vez de adivinar selectores. Devuelve descripción de lo clickeado o null.
_JS_CLICK_LUPA = """() => {
    const inp = document.querySelector('#rutTrabajador');
    if (!inp) return null;
    let node = inp;
    for (let up = 0; up < 6 && node; up++) {
        node = node.parentElement;
        if (!node) break;
        const btns = node.querySelectorAll(
            "button, [role=button], mat-icon, [class*='search' i], [class*='lupa' i]");
        for (const b of btns) {
            const txt = ((b.innerText || '') + '|' + (b.getAttribute('aria-label') || '') + '|' +
                         (typeof b.className === 'string' ? b.className : '') + '|' +
                         (b.getAttribute('title') || '')).toLowerCase();
            if (txt.includes('buscar') || txt.includes('search') || txt.includes('lupa') ||
                txt.includes('magnif') ||
                b.querySelector("[class*='search' i], [class*='lupa' i]")) {
                b.click();
                return ('match: ' + b.tagName + ' ' + txt).slice(0, 100);
            }
        }
        // Si el contenedor cercano tiene UN solo botón, ese es (lupa junto al campo)
        if (btns.length === 1) {
            btns[0].click();
            return ('unico: ' + btns[0].tagName + ' ' +
                    (typeof btns[0].className === 'string' ? btns[0].className : '')).slice(0, 100);
        }
    }
    return null;
}"""

# JS: vuelca todos los botones del iframe (diagnóstico de un solo disparo)
_JS_BOTONES = """() => {
    const out = [];
    for (const b of document.querySelectorAll("button, [role=button], mat-icon, a")) {
        let d = b.tagName.toLowerCase();
        if (b.id) d += '#' + b.id;
        const cls = (typeof b.className === 'string' ? b.className : '')
            .trim().split(/\\s+/).slice(0, 3).join('.');
        if (cls) d += '.' + cls;
        const t = ((b.innerText || '').trim().replace(/\\s+/g, ' ')).slice(0, 30);
        const aria = b.getAttribute('aria-label') || '';
        out.push(d + ' "' + t + '"' + (aria ? ' aria="' + aria + '"' : ''));
    }
    return [...new Set(out)].slice(0, 45);
}"""


# JS: lee la COMUNA del trabajador. En el form DT es un mat-select (Angular
# Material), que _JS_EXTRAER no capta porque no es <input>/<select> nativo.
_JS_LEER_COMUNA = """() => {
    const norm = s => (s || '').toLowerCase();
    const limpiar = s => (s || '').replace(/\\s+/g, ' ').trim();
    // 1) control cuyo id/formcontrolname mencione 'comuna' Y 'trab' (evita empleador)
    for (const el of document.querySelectorAll('mat-select, select, input, [formcontrolname]')) {
        const key = norm((el.id || '') + ' ' + (el.getAttribute('formcontrolname') || ''));
        if (!(key.includes('comuna') && key.includes('trab'))) continue;
        if (el.tagName === 'SELECT') {
            const o = el.options[el.selectedIndex];
            if (o && limpiar(o.text)) return limpiar(o.text);
        }
        const vt = el.querySelector('.mat-select-value-text, .mat-mdc-select-value-text, [class*=select-value]');
        if (vt && limpiar(vt.innerText)) return limpiar(vt.innerText);
        if (el.value && limpiar(el.value)) return limpiar(el.value);
    }
    // 2) por etiqueta 'comuna' (evitando la del empleador, que suele ir arriba)
    const labs = [...document.querySelectorAll('label, mat-label')].filter(
        l => norm(l.innerText).includes('comuna') && !norm(l.innerText).includes('emplea'));
    for (const lab of labs.reverse()) {   // la del trabajador suele ir más abajo
        const cont = lab.closest('mat-form-field, .form-group, .field, .row, div');
        if (!cont) continue;
        const sel = cont.querySelector('select');
        if (sel) { const o = sel.options[sel.selectedIndex]; if (o && limpiar(o.text)) return limpiar(o.text); }
        const vt = cont.querySelector('.mat-select-value-text, .mat-mdc-select-value-text, [class*=select-value]');
        if (vt && limpiar(vt.innerText)) return limpiar(vt.innerText);
        const inp = cont.querySelector('input');
        if (inp && limpiar(inp.value)) return limpiar(inp.value);
    }
    return '';
}"""


def _aplanar_json(obj, out, clave=""):
    """Aplana un JSON anidado a {clave_minuscula: valor_str}."""
    if isinstance(obj, dict):
        for k, v in obj.items():
            _aplanar_json(v, out, k)
    elif isinstance(obj, list):
        for v in obj[:25]:
            _aplanar_json(v, out, clave)
    elif obj is not None and obj != "":
        k = clave.lower()
        if k and k not in out:
            out[k] = str(obj).strip()


def _extraer_de_json(respuestas, log):
    """Busca datos de persona en las respuestas API capturadas (la más reciente primero).
    Lee el JSON crudo de la API de DT — no depende de cómo Angular pinte el formulario."""
    for resp in reversed(respuestas[-15:]):
        try:
            body = resp.json()
        except Exception:
            continue
        plano = {}
        _aplanar_json(body, plano)

        def _k(*terms, excluir=()):
            for t in terms:
                for k, v in plano.items():
                    if t in k and v and not any(x in k for x in excluir):
                        return v
            return ""

        nombres   = _k("nombres", excluir=("razon", "empleador", "empresa")) or \
                    _k("nombre", excluir=("razon", "empleador", "empresa", "usuario"))
        ap_p      = _k("paterno")
        ap_m      = _k("materno")
        apellidos = f"{ap_p} {ap_m}".strip() or _k("apellido")
        correo    = _k("mail", "correo", excluir=("empleador",))
        calle     = _k("calle", excluir=("empleador",)) or _k("direccion", excluir=("empleador",))
        numero    = _k("numero", excluir=("telefono", "fono", "empleador", "documento"))
        # La comuna suele venir anidada ("comuna": {"nombre": "..."}), y el
        # aplanado pierde ese contexto. Se busca recursivamente respetando la
        # anidación; el _k plano queda sólo como respaldo.
        comuna    = _buscar_comuna_json(body) or _k("comuna", excluir=("empleador", "juri"))
        fnac      = _k("nacimiento")

        if nombres or correo or calle:
            try:
                log(f"  [api] datos desde: ...{resp.url.split('?')[0][-70:]}", "info")
            except Exception:
                pass
            return {"nombres": nombres, "apellidos": apellidos, "fecha_nac": fnac,
                    "correo": correo, "calle": calle, "numero": numero, "comuna": comuna}
    return None


def _valor_comuna(v):
    """Extrae el texto de una comuna dado un valor: string directo, u objeto
    anidado como {codigo, nombre} / {descripcion} / {glosa}. Descarta códigos."""
    if isinstance(v, str):
        s = v.strip()
        # descartar códigos numéricos puros (ej. "05801")
        return s if (s and not s.replace(".", "").replace("-", "").isdigit()) else ""
    if isinstance(v, dict):
        for nk in ("nombrecomuna", "nombre_comuna", "nombre", "descripcion",
                   "glosa", "valor", "nombrecompleto"):
            for k2, val in v.items():
                if k2.lower().replace(" ", "") == nk and isinstance(val, str) and val.strip():
                    return val.strip()
        # comuna anidada como {"comuna": "X"}
        for k2, val in v.items():
            if "comuna" in k2.lower() and isinstance(val, str) and val.strip() \
                    and not val.strip().isdigit():
                return val.strip()
    return ""


def _buscar_comuna_json(obj, _prof=0):
    """Recorre el JSON de la API buscando la comuna del trabajador (clave que
    contenga 'comuna' a cualquier profundidad), respetando la anidación."""
    if _prof > 14:
        return ""
    if isinstance(obj, dict):
        # 1) claves 'comuna' directas en este nivel (excluye empleador/jurídica)
        for k, v in obj.items():
            kl = k.lower()
            if "comuna" in kl and "emplea" not in kl and "juri" not in kl and "empresa" not in kl:
                nombre = _valor_comuna(v)
                if nombre:
                    return nombre
        # 2) descender, saltando ramas del empleador/empresa
        for k, v in obj.items():
            kl = k.lower()
            if "emplea" in kl or "empresa" in kl:
                continue
            r = _buscar_comuna_json(v, _prof + 1)
            if r:
                return r
    elif isinstance(obj, list):
        for v in obj[:40]:
            r = _buscar_comuna_json(v, _prof + 1)
            if r:
                return r
    return ""


def _snippet_direccion(body):
    """Fragmento JSON del nodo de dirección (para la columna DEBUG)."""
    import json as _json

    def buscar(o, _p=0):
        if _p > 12:
            return None
        if isinstance(o, dict):
            ks = " ".join(o.keys()).lower()
            if "comuna" in ks or "calle" in ks or "direccion" in ks:
                return o
            for v in o.values():
                r = buscar(v, _p + 1)
                if r:
                    return r
        elif isinstance(o, list):
            for v in o[:40]:
                r = buscar(v, _p + 1)
                if r:
                    return r
        return None

    node = buscar(body)
    if node is None:
        return ""
    try:
        return _json.dumps(node, ensure_ascii=False)[:350]
    except Exception:
        return str(node)[:350]


def _consultar_rut(page, rut, log, primera=False, snap=None):
    """Rellena el RUT del trabajador, dispara la búsqueda y extrae los datos.
    Doble vía: campos del formulario + captura directa de la respuesta API."""
    frame = _get_iframe_frame(page, espera=10)
    if frame is None:
        frame = _frame_form(page)

    # 1. Asegurar radio "Cédula de identidad" marcado (habilita la sección trabajador)
    try:
        radio_info = frame.evaluate(_JS_RADIO_CEDULA)
        if primera:
            log(f"  [debug] radio doc: {radio_info}", "info")
    except Exception:
        pass

    # 2. Campo RUT
    rut_input = _buscar_campo_rut(frame)
    if rut_input is None:
        raise Exception("No se encontró campo RUT en el formulario")
    rut_input.scroll_into_view_if_needed(timeout=5000)

    # Firma de los datos actuales (para detectar cambio real entre RUTs consecutivos)
    firma_previa = _firma_trab(_extraer_datos_frame(frame))

    # 3. Capturar las respuestas API de DT mientras dura la consulta
    respuestas = []

    def _cap(r):
        try:
            u = r.url
            if ("dirtrab" in u or "dt.gob" in u) and \
                    r.request.resource_type in ("xhr", "fetch"):
                respuestas.append(r)
        except Exception:
            pass

    page.on("response", _cap)
    try:
        # 4. Escribir RUT con teclas reales (incluye Tab/blur al final)
        _tipo_campo(None, rut_input, rut)

        # 5. Disparar búsqueda: lupa junto al campo (por cercanía) y Enter de respaldo
        lupa = None
        try:
            lupa = frame.evaluate(_JS_CLICK_LUPA)
        except Exception:
            pass
        if primera:
            log(f"  [debug] lupa: {lupa}", "info")
        if not lupa:
            try:
                rut_input.press("Enter")
            except Exception:
                pass

        # 6. Esperar datos NUEVOS del trabajador (sale apenas llegan, máx 10s)
        datos = _esperar_datos(frame, max_seg=10, firma_previa=firma_previa)
    finally:
        try:
            page.remove_listener("response", _cap)
        except Exception:
            pass

    # Si los campos siguen con los valores del RUT anterior, NO son de esta persona
    if any(firma_previa) and _firma_trab(datos) == firma_previa:
        datos = {}

    # 7. Extracción por ID exacto del formulario DT
    nombres   = datos.get("nombresTrabajador", "")
    apellidos = datos.get("apellidosTrabajador", "")
    fecha_nac = datos.get("fechaNacimientoTrabajador", "")
    correo    = datos.get("emailTrabajador", "")
    calle     = datos.get("calleTrabajador", "")
    numero    = datos.get("numeroTrabajador", "")
    comuna    = datos.get("comunaTrabajador", "")

    # 8. Vía alternativa: leer la respuesta JSON de la API directamente
    if not (nombres or correo or calle):
        api = _extraer_de_json(respuestas, log)
        if api:
            nombres   = api["nombres"]
            apellidos = api["apellidos"]
            fecha_nac = api["fecha_nac"]
            correo    = api["correo"]
            calle     = api["calle"]
            numero    = api["numero"]
            comuna    = api["comuna"]

    # 8b. Comuna: resolución independiente (el form la muestra en un mat-select
    # que _JS_EXTRAER no capta, y la vía API de arriba sólo corre si TODO vino
    # vacío). Se resuelve aunque el resto de datos ya haya llegado del formulario.
    if not comuna:
        # buscar en los datos del form cualquier clave que mencione 'comuna'
        for k, v in datos.items():
            kl = k.lower()
            if "comuna" in kl and "emplea" not in kl and "juri" not in kl and v:
                comuna = v
                break
    if not comuna:
        try:
            comuna = (frame.evaluate(_JS_LEER_COMUNA) or "").strip()
        except Exception:
            pass
    if not comuna and respuestas:
        api = _extraer_de_json(respuestas, log)
        if api and api.get("comuna"):
            comuna = api["comuna"]
    # Buscador recursivo directo sobre el JSON crudo (comuna anidada)
    if not comuna and respuestas:
        for resp in reversed(respuestas[-15:]):
            try:
                c = _buscar_comuna_json(resp.json())
            except Exception:
                c = ""
            if c:
                comuna = c
                break

    # Diagnóstico de comuna (solo primer RUT): mostrar dónde podría estar
    if primera:
        log(f"  [debug] comuna resuelta: '{comuna}'", "info")
        try:
            _terms = ("comuna", "region", "direccion", "calle", "ciudad", "localidad")
            campos_addr = [f"{k}={v}" for k, v in datos.items()
                           if any(t in k.lower() for t in _terms) and v][:15]
            log(f"  [debug] campos form dirección: {campos_addr}", "warn")
        except Exception:
            pass
        try:
            for resp in reversed(respuestas[-15:]):
                try:
                    body = resp.json()
                except Exception:
                    continue
                plano = {}
                _aplanar_json(body, plano)
                addr = {k: v for k, v in plano.items()
                        if any(t in k for t in ("comuna", "region", "direccion", "calle", "ciudad"))}
                if addr:
                    log(f"  [debug] API {resp.url.split('?')[0][-45:]} → {addr}", "warn")
        except Exception:
            pass

    # 9. Diagnóstico completo de un solo disparo (solo primer RUT sin datos)
    if primera and not (nombres or correo or calle):
        try:
            log("  [debug] botones del iframe:", "warn")
            for b in frame.evaluate(_JS_BOTONES):
                log(f"    {b}", "warn")
        except Exception:
            pass
        if respuestas:
            log("  [debug] llamadas API durante la consulta:", "warn")
            for r in respuestas[-10:]:
                try:
                    log(f"    {r.status} {r.url[:95]}", "warn")
                except Exception:
                    pass
        else:
            log("  [debug] NINGUNA llamada API se disparó — la búsqueda no se está gatillando", "warn")
        if snap:
            snap("rut_sin_datos")

    # DEBUG temporal: nodo de dirección crudo del JSON de la API (para ubicar comuna)
    _dbg = []
    try:
        for resp in reversed(respuestas[-8:]):
            try:
                body = resp.json()
            except Exception:
                continue
            snip = _snippet_direccion(body)
            if snip:
                _dbg.append(snip)
                break
    except Exception:
        pass

    return {
        "RUT Trabajador": rut,
        "Nombres":        nombres,
        "Apellidos":      apellidos,
        "Fecha Nac.":     fecha_nac,
        "Correo":         correo,
        "Calle":          f"{calle} {numero}".strip(),
        "Comuna":         comuna,
        "Error":          "" if (nombres or correo or calle) else "sin datos",
        "DEBUG":          " | ".join(_dbg[:8]),
    }


# ── Función principal ──────────────────────────────────────────────────────────

def consultar_ruts(run, clave, rut_empresa, lista_ruts, log, debug_dir=None):
    """Consulta todos los RUTs y devuelve lista de dicts."""
    resultados = []

    # Pantalla virtual (xvfb): permite correr el navegador CON pantalla en el
    # servidor. El portal DT no renderiza el formulario en modo headless, así
    # que arrancamos un display virtual y lanzamos Chromium headed. Si no está
    # disponible (p. ej. en local Windows), caemos a headless.
    _display = None
    _headless = True
    try:
        from pyvirtualdisplay import Display
        _display = Display(visible=0, size=(1440, 1000))
        _display.start()
        _headless = False
        log("[debug] Pantalla virtual (xvfb) iniciada — navegador con pantalla", "info")
    except Exception as e:
        log(f"[debug] Sin pantalla virtual ({str(e)[:50]}); usando headless", "warn")

    try:
      with sync_playwright() as pw:
        # Medidas anti-detección: el portal DT no renderiza el formulario si
        # detecta un navegador automatizado (headless / navigator.webdriver).
        browser = pw.chromium.launch(
            headless=_headless,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--disable-features=IsolateOrigins,site-per-process",
                "--no-sandbox",
            ],
        )
        ctx = browser.new_context(
            viewport={"width": 1400, "height": 900},
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/125.0.0.0 Safari/537.36"),
            locale="es-CL",
            timezone_id="America/Santiago",
        )
        # Ocultar señales de automatización antes de cargar cualquier página
        ctx.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            Object.defineProperty(navigator, 'languages', {get: () => ['es-CL','es']});
            Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3,4,5]});
            window.chrome = window.chrome || {runtime: {}};
        """)
        page    = ctx.new_page()
        page.set_default_timeout(45000)
        page.set_default_navigation_timeout(45000)

        # La página "activa" puede cambiar si el formulario abre pestaña nueva
        estado = {"page": page}

        def snap(nombre):
            if not debug_dir:
                return
            import os as _os
            pg = estado["page"]
            try:
                ruta = _os.path.join(debug_dir, f"midt_{nombre}.png")
                pg.screenshot(path=ruta, full_page=True)
                log(f"[shot] captura guardada: midt_{nombre}.png", "info")
            except Exception:
                pass
            # Guardar también el HTML de la pantalla (para inspeccionar el DOM real)
            try:
                ruta_html = _os.path.join(debug_dir, f"midt_{nombre}.html")
                with open(ruta_html, "w", encoding="utf-8") as fh:
                    fh.write(pg.content())
            except Exception:
                pass
            # Y el HTML del iframe del formulario (donde vive el DOM que importa)
            try:
                for fr in pg.frames:
                    if _IFRAME_SRC in (fr.url or ""):
                        ruta_if = _os.path.join(debug_dir, f"midt_{nombre}_iframe.html")
                        with open(ruta_if, "w", encoding="utf-8") as fh:
                            fh.write(fr.content())
                        break
            except Exception:
                pass

        try:
            hacer_login(page, run, clave, log)
            snap("1_roles")
            seleccionar_empresa(page, rut_empresa, log, snap)
            snap("2_post_empresa")
            page_form = _ir_a_formulario(page, log, snap)
            estado["page"] = page_form
            snap("3_formulario")
            _dump_iframe(page_form, log)   # volcar campos del formulario una vez

            for i, rut in enumerate(lista_ruts, 1):
                rut = rut.strip()
                if not rut:
                    continue
                log(f"[{i}/{len(lista_ruts)}] {rut}...", "info")
                try:
                    datos = _consultar_rut(page_form, rut, log, primera=(i == 1), snap=snap)
                    resultados.append(datos)
                    ok = datos.get("Nombres") or datos.get("Calle")
                    log(f"[{i}] {'✓' if ok else '⚠'} {rut} → "
                        f"{datos.get('Calle','')} {datos.get('Comuna','')}", "ok" if ok else "warn")
                except Exception as e:
                    log(f"[{i}] Error {rut}: {e}", "warn")
                    resultados.append({"RUT Trabajador": rut, "Error": str(e)})
                    # Reintentar reabriendo el formulario
                    try:
                        page_form = _ir_a_formulario(page, log)
                        estado["page"] = page_form
                    except Exception:
                        pass

        except Exception as e:
            log(f"Error fatal: {e}", "err")
            try:
                snap("error")
            except Exception:
                pass
            raise
        finally:
            browser.close()
    finally:
        if _display:
            try:
                _display.stop()
            except Exception:
                pass

    return resultados


# ── Excel ──────────────────────────────────────────────────────────────────────

COLS = ["RUT Trabajador", "Nombres", "Apellidos", "Fecha Nac.", "Correo", "Calle", "Comuna", "Error", "DEBUG"]
ANCHOS = [16, 22, 22, 14, 30, 30, 18, 20, 70]


def generar_excel(resultados: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consulta RUTs DT"

    fill_h = PatternFill("solid", start_color="1E3A5F")
    font_h = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    borde  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col, nombre in enumerate(COLS, 1):
        c = ws.cell(row=1, column=col, value=nombre)
        c.font = font_h; c.fill = fill_h; c.border = borde
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    font_d   = Font(name="Arial", size=9)
    fill_par = PatternFill("solid", start_color="EFF6FF")

    for fi, fila in enumerate(resultados, 2):
        fill = fill_par if fi % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        for col, key in enumerate(COLS, 1):
            c = ws.cell(row=fi, column=col, value=fila.get(key, ""))
            c.font = font_d; c.border = borde; c.fill = fill
            c.alignment = Alignment(horizontal="left", vertical="center")

    for col, ancho in enumerate(ANCHOS, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = ancho
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
