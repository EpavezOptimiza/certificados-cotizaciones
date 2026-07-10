"""Convierte PDFs de planillas Previred a Excel unificado."""
import io, os, re
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNAS = [
    "RUT Empresa", "Razón Social", "RUT", "Nombre",
    "Rem. Imponible", "Cot. Obligatoria", "SIS", "Cot. Voluntaria",
    "N° Contrato APVI", "Dep. Convenido", "Dep. Cta. Ahorro",
    "Rem. Imp. Cesantia", "Cot. Afiliado", "Cot. Empleador",
    "Cod.", "Fecha Inicio", "Fecha Termino",
    "AFP", "Periodo", "Nomina"
]

COLS_PESOS = {
    "Rem. Imponible", "Cot. Obligatoria", "SIS", "Cot. Voluntaria",
    "Dep. Convenido", "Dep. Cta. Ahorro",
    "Rem. Imp. Cesantia", "Cot. Afiliado", "Cot. Empleador"
}

ANCHOS = [15,30,15,40,16,16,12,16,14,14,14,16,14,14,8,14,14,15,12,25]


def _detectar_afp(texto: str) -> str:
    for afp in ["AFP Provida","AFP Capital","AFP Habitat","AFP Modelo",
                "AFP Uno","AFP PlanVital","AFP Cuprum","AFP Crecer"]:
        if afp.lower() in texto.lower():
            return afp
    return "AFP"


def _extraer_periodo(nombre: str) -> str:
    m = re.match(r"(\d{4})-(\d{2})-", nombre)
    return f"{m.group(2)}/{m.group(1)}" if m else ""


def _extraer_nomina(nombre: str) -> str:
    m = re.match(r"\d{4}-\d{2}-(.+)\.pdf", nombre)
    return m.group(1) if m else nombre


def _limpiar_num(val):
    try:
        return int(str(val).replace(".", "").replace(",", "").strip())
    except Exception:
        return 0


def _formato_pesos(val) -> str:
    try:
        return f"$ {int(val):,}".replace(",", ".")
    except Exception:
        return str(val)


_SKIP_KW = ['rut empresa', 'período', 'periodo', 'institución', 'institucion',
            'nombre o', 'folio', 'página', 'pagina', 'detalle de pago', 'totales']
_RUT_RE  = re.compile(r'(\d{1,2}\.\d{3}\.\d{3}-[\dkK])')


def _extraer_empresa_del_pdf(texto: str, log=None):
    """Extrae RUT empresa y razón social del texto extraído por pdfplumber."""
    rut = ""
    razon = ""
    lineas = texto.split("\n")

    if log:
        preview = " | ".join(
            f"[{i}]{l.strip()[:55]}" for i, l in enumerate(lineas[:20]) if l.strip()
        )
        log(f"[empresa-debug] {preview}", "debug")

    # Estrategia 1: línea con "razón social" → buscar empresa en las siguientes 5 líneas
    for i, linea in enumerate(lineas):
        s = linea.strip()
        if not ("razón social" in s.lower() or "razon social" in s.lower()):
            continue
        for j in range(i + 1, min(i + 6, len(lineas))):
            sig = lineas[j].strip()
            if not sig:
                continue
            m = _RUT_RE.search(sig)
            if m:
                candidate = sig[:m.start()].strip()
                if len(candidate) > 5:
                    rut = m.group(1)
                    razon = candidate
                    if log:
                        log(f"[empresa] strat1 → {razon} | {rut}", "info")
                    return rut, razon

    # Estrategia 2: cualquier línea con nombre >10 chars + RUT (no empieza con dígito)
    for linea in lineas:
        s = linea.strip()
        m = _RUT_RE.search(s)
        if not m:
            continue
        before = s[:m.start()].strip()
        if (len(before) > 10
                and not re.match(r'^\d', before)
                and not any(kw in before.lower() for kw in _SKIP_KW)):
            rut = m.group(1)
            razon = before
            if log:
                log(f"[empresa] strat2 → {razon} | {rut}", "info")
            return rut, razon

    if log:
        log("[empresa] no se pudo extraer empresa del PDF", "warn")
    return rut, razon


def extraer_trabajadores(ruta_pdf: str, nombre_archivo: str,
                         rut_empresa: str, razon_social: str, log=None) -> list:
    filas = []
    periodo = _extraer_periodo(nombre_archivo)
    nomina  = _extraer_nomina(nombre_archivo)

    try:
        with pdfplumber.open(ruta_pdf) as pdf:
            # Extraer empresa de la primera página (sin filtro DETALLE DE PAGO)
            if (not rut_empresa or not razon_social) and pdf.pages:
                texto_p1 = pdf.pages[0].extract_text() or ""
                rut_pdf, razon_pdf = _extraer_empresa_del_pdf(texto_p1, log)
                if not rut_empresa and rut_pdf:
                    rut_empresa = rut_pdf
                if not razon_social and razon_pdf:
                    razon_social = razon_pdf

            for page in pdf.pages:
                texto = page.extract_text()
                if not texto or "DETALLE DE PAGO" not in texto.upper():
                    continue
                # Si aún no tenemos empresa, intentar con esta página también
                if not rut_empresa or not razon_social:
                    rut_pdf, razon_pdf = _extraer_empresa_del_pdf(texto, log)
                    if not rut_empresa and rut_pdf:
                        rut_empresa = rut_pdf
                    if not razon_social and razon_pdf:
                        razon_social = razon_pdf
                afp = _detectar_afp(texto)
                for linea in texto.split("\n"):
                    linea = linea.strip()
                    if not linea:
                        continue
                    if not re.match(r"^\d[\d\.]{6,11}-[\dkK]\s", linea):
                        continue
                    if "TOTALES" in linea.upper():
                        continue
                    partes = linea.split()
                    if len(partes) < 4:
                        continue
                    rut = partes[0]
                    idx_num = None
                    for i, p in enumerate(partes[1:], 1):
                        if re.match(r"^\d[\d\.]*$", p):
                            idx_num = i
                            break
                    if idx_num is None:
                        continue
                    nombre = " ".join(partes[1:idx_num])
                    nums   = partes[idx_num:]

                    def n(i):
                        try:
                            v = nums[i]
                            if re.match(r"\d{2}/\d{2}/\d{4}", v):
                                return v
                            return _limpiar_num(v)
                        except Exception:
                            return 0

                    cod = 0
                    fecha_inicio = ""
                    fecha_term   = ""
                    try:
                        cod_val = nums[9] if len(nums) > 9 else "0"
                        if not re.match(r"\d{2}/\d{2}/\d{4}", cod_val):
                            cod = int(cod_val)
                    except Exception:
                        cod = 0
                    for v in nums[10:]:
                        if re.match(r"\d{2}/\d{2}/\d{4}", v):
                            if not fecha_inicio:
                                fecha_inicio = v
                            elif not fecha_term:
                                fecha_term = v

                    filas.append({
                        "RUT Empresa":        rut_empresa,
                        "Razón Social":       razon_social,
                        "RUT":                rut,
                        "Nombre":             nombre,
                        "Rem. Imponible":     n(0),
                        "Cot. Obligatoria":   n(1),
                        "SIS":                n(2),
                        "Cot. Voluntaria":    n(3),
                        "N° Contrato APVI":   n(4),
                        "Dep. Convenido":     n(5),
                        "Dep. Cta. Ahorro":   n(6),
                        "Rem. Imp. Cesantia": n(6),
                        "Cot. Afiliado":      n(7),
                        "Cot. Empleador":     n(8),
                        "Cod.":               cod,
                        "Fecha Inicio":       fecha_inicio,
                        "Fecha Termino":      fecha_term,
                        "AFP":                afp,
                        "Periodo":            periodo,
                        "Nomina":             nomina,
                    })
    except Exception as e:
        print(f"Error en {nombre_archivo}: {e}")
    return filas


def generar_excel_bytes(rutas_pdf: list, rut_empresa: str,
                        razon_social: str, log=None) -> bytes:
    """Procesa los PDFs y devuelve el Excel como bytes."""
    todas = []
    for i, ruta in enumerate(rutas_pdf, 1):
        nombre = os.path.basename(ruta)
        filas = extraer_trabajadores(ruta, nombre, rut_empresa, razon_social, log)
        if log:
            log(f"[{i}/{len(rutas_pdf)}] {nombre} → {len(filas)} trabajadores", "info")
        todas.extend(filas)

    if not todas:
        raise ValueError("No se encontraron datos en los PDFs")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planillas Unificadas"

    fill_h = PatternFill("solid", start_color="4B0082")
    font_h = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    borde  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col, nombre_col in enumerate(COLUMNAS, 1):
        c = ws.cell(row=1, column=col, value=nombre_col)
        c.font      = font_h
        c.fill      = fill_h
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = borde
    ws.row_dimensions[1].height = 30

    font_d   = Font(name="Arial", size=9)
    fill_par = PatternFill("solid", start_color="F3F0FF")
    fill_imp = PatternFill("solid", start_color="FFFFFF")

    for fi, fila in enumerate(todas, 2):
        fill = fill_par if fi % 2 == 0 else fill_imp
        for col, key in enumerate(COLUMNAS, 1):
            val = fila.get(key, "")
            if key in COLS_PESOS:
                val = _formato_pesos(val) if val and val != 0 else ""
            c = ws.cell(row=fi, column=col, value=val)
            c.font      = font_d
            c.border    = borde
            c.fill      = fill
            c.alignment = Alignment(horizontal="center", vertical="center")

    for col, ancho in enumerate(ANCHOS, 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
