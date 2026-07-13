# -*- coding: utf-8 -*-
"""
Lógica de lectura del Excel de cierre y generación de datos para gráficos.
Lee las tablas pivot de la hoja "Cierre ..." o la hoja raw "afc-afp" / "Base AFP".
"""
import io
import re
import openpyxl

MESES_ES = ['','Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']

def _fmt_monto(v):
    try:
        return int(float(str(v).replace('$','').replace(',','')))
    except:
        return 0

def _norm(s):
    """Normaliza texto: minúsculas, reemplaza ñ corrupta (?) para comparar."""
    return str(s or '').strip().lower().replace('?', '\xf1')

def leer_cierre(wb, nombre_hoja):
    ws = wb[nombre_hoja]
    datos = {}

    filas = []
    for row in ws.iter_rows(values_only=True):
        if any(c is not None for c in row):
            filas.append(list(row))
        else:
            filas.append(None)

    def buscar_tabla(header_col0, max_buscar=200):
        """Busca tabla cuyo encabezado col-A coincide con header_col0."""
        target = _norm(header_col0)
        for i, fila in enumerate(filas[:max_buscar]):
            if fila is None: continue
            if _norm(fila[0]) == target:
                tabla = {}
                for j in range(i+1, min(i+50, len(filas))):
                    f = filas[j]
                    if f is None: break
                    k = str(f[0] or '').strip()
                    if not k: break
                    v = _fmt_monto(f[1]) if len(f) > 1 else 0
                    if k.lower() == 'total general':
                        tabla['__total__'] = v
                    else:
                        tabla[k] = v
                return tabla
        return {}

    # 1. Por institución
    datos['por_institucion'] = buscar_tabla('institucion')
    if not datos['por_institucion']:
        datos['por_institucion'] = buscar_tabla('afp')

    # 2. Por año — maneja 'años', 'año' y variantes con ñ corrupta
    datos['por_anio'] = (buscar_tabla('años') or
                         buscar_tabla('año') or
                         buscar_tabla('a�s') or
                         buscar_tabla('a�'))

    # 3. Por estatus
    datos['por_estatus'] = buscar_tabla('estatus')

    # 4. Por motivo de deuda
    datos['por_motivo'] = (buscar_tabla('26_ motivos de deuda') or
                           buscar_tabla('motivos de deuda'))

    # 5. Tabla resumen AFP-AFC vs ISAPRE vs Total
    datos['resumen_tabla'] = _buscar_resumen(filas)

    # 6. Fee (columna "Suma de 22_ FEE% MONTO") junto al estatus
    datos['por_fee'] = _buscar_fee(filas)

    # 7. Header del cliente (NV, checkmarks de entregables)
    datos['header_cliente'] = _buscar_header(filas)

    return datos


def _buscar_resumen(filas):
    """Busca la tabla con columnas Estatus / AFP-AFC / ISAPRE / Total."""
    for i, fila in enumerate(filas[:30]):
        if fila is None: continue
        for j, v in enumerate(fila):
            if _norm(v) == 'estatus':
                tabla = {'headers': [], 'filas': []}
                for k in range(j+1, min(j+4, len(fila))):
                    tabla['headers'].append(str(fila[k] or '').strip())
                for m in range(i+1, min(i+15, len(filas))):
                    f = filas[m]
                    if f is None or (j < len(f) and f[j] is None): continue
                    etiqueta = str(f[j] if j < len(f) else '').strip()
                    if not etiqueta: continue
                    vals = [_fmt_monto(f[k]) for k in range(j+1, min(j+4, len(f)))]
                    tabla['filas'].append({'label': etiqueta, 'valores': vals})
                if tabla['filas']:
                    return tabla
    return {}


def _buscar_fee(filas):
    """
    Extrae la tabla de estatus + fee buscando la columna 'Suma de 22_ FEE% MONTO'.
    Retorna {estatus: fee_amount, '__total__': total_fee}.
    """
    for i, fila in enumerate(filas):
        if fila is None: continue
        fee_col = None
        for j, v in enumerate(fila):
            s = _norm(v)
            if 'suma de 22' in s or '22_ fee' in s:
                fee_col = j
                break
        if fee_col is None or fee_col < 1:
            continue

        label_col = max(0, fee_col - 2)
        monto_col = fee_col - 1

        tabla = {}
        for k in range(i+1, min(i+20, len(filas))):
            f = filas[k]
            if f is None: break
            if len(f) <= fee_col: break
            label = str(f[label_col] if label_col < len(f) else '').strip()
            if not label: break
            fee = _fmt_monto(f[fee_col])
            if label.lower() == 'total general':
                tabla['__total__'] = fee
                break
            else:
                tabla[label] = fee
        if tabla:
            return tabla
    return {}


def _buscar_header(filas):
    """Extrae info del cliente de la fila de encabezado (NV, cliente, checks)."""
    for i, fila in enumerate(filas[:10]):
        if fila is None: continue
        for j, v in enumerate(fila):
            if str(v or '').strip().upper() == 'CLIENTE':
                info = {}
                if i+1 < len(filas) and filas[i+1]:
                    heads = fila
                    vals  = filas[i+1]
                    for k in range(j, min(len(heads), len(vals))):
                        h = str(heads[k] or '').strip()
                        val = vals[k]
                        if h:
                            info[h] = val
                return info
    return {}


def detectar_hojas_cierre(wb):
    """Retorna hojas de cierre (excluye hojas auxiliares como 'Cierre Isapre')."""
    excluir = {'isapre'}
    result = []
    for n in wb.sheetnames:
        nl = n.lower()
        if 'cierre' in nl and not any(ex in nl for ex in excluir):
            result.append(n)
    return result


def _detectar_columnas_raw(ws):
    """Detecta índices de columna en la hoja raw (afc-afp o Base AFP) leyendo la fila 1."""
    hdr = {}
    for col in range(1, min(ws.max_column + 1, 60)):
        h = str(ws.cell(1, col).value or '').strip().upper()
        hn = h.replace('_', '').replace(' ', '')
        if 'INSTITUCION' in hn or 'INSTITUCIÓN' in hn or '10_' in h:
            hdr['inst'] = col
        elif 'PERIODOAR' in hn or 'PERIODO' in hn or 'PERÍODO' in hn or '11_' in h:
            hdr['per'] = col
        elif 'MONTOINTER' in hn or 'INTERÉS' in hn or 'INTERES' in hn or '13_' in h:
            hdr['monto'] = col
        elif 'FEE' in hn or '22_' in h:
            hdr['fee'] = col
        elif 'MOTIVO' in hn or '26_' in h:
            hdr['motivo'] = col
        elif 'ESTATUS' in hn or 'ESTADO' in hn or '34_' in h:
            hdr['estatus'] = col
    # Fallback para "Base AFP" generado por app (columnas fijas conocidas)
    if 'inst' not in hdr:
        hdr = {'inst': 6, 'per': 7, 'monto': 10, 'estatus': 8}
    return hdr


def leer_base_afp(wb):
    """
    Lee la hoja raw 'afc-afp' o 'Base AFP' y computa los mismos pivots que leer_cierre().
    Permite mostrar Reportes sin necesitar un Excel con hojas Cierre.
    """
    import datetime as _dt

    hoja_nombre = None
    for candidato in ['afc-afp', 'Base AFP']:
        if candidato in wb.sheetnames:
            hoja_nombre = candidato
            break
    if not hoja_nombre:
        return None

    ws = wb[hoja_nombre]
    hdr = _detectar_columnas_raw(ws)

    datos = {
        'por_institucion': {},
        'por_anio': {},
        'por_estatus': {},
        'por_motivo': {},
        'resumen_tabla': {},
        'por_fee': {},
        'header_cliente': {},
    }

    def _anio_de(val):
        if not val:
            return ''
        if isinstance(val, (_dt.date, _dt.datetime)):
            return str(val.year)
        s = str(val).strip()
        m = re.search(r'(\d{4})', s)
        return m.group(1) if m else ''

    for r in range(2, ws.max_row + 1):
        inst = str(ws.cell(r, hdr.get('inst', 6)).value or '').strip()
        if not inst:
            continue

        monto   = _fmt_monto(ws.cell(r, hdr.get('monto', 10)).value)
        estatus = str(ws.cell(r, hdr.get('estatus', 8)).value or '').strip()
        anio    = _anio_de(ws.cell(r, hdr.get('per', 7)).value) if 'per' in hdr else ''
        motivo  = str(ws.cell(r, hdr.get('motivo', 0)).value or '').strip() if 'motivo' in hdr else ''
        fee     = _fmt_monto(ws.cell(r, hdr.get('fee', 0)).value) if 'fee' in hdr else 0

        if monto == 0:
            continue

        datos['por_institucion'][inst] = datos['por_institucion'].get(inst, 0) + monto
        if anio:
            datos['por_anio'][anio] = datos['por_anio'].get(anio, 0) + monto
        if estatus:
            datos['por_estatus'][estatus] = datos['por_estatus'].get(estatus, 0) + monto
        if motivo:
            datos['por_motivo'][motivo] = datos['por_motivo'].get(motivo, 0) + monto
        if fee and estatus:
            datos['por_fee'][estatus] = datos['por_fee'].get(estatus, 0) + fee

    # Ordenar años
    datos['por_anio'] = dict(sorted(datos['por_anio'].items()))

    return datos


def detectar_tipo_excel(wb):
    """Detecta si el Excel tiene hojas Cierre o es una Base de Deudas raw."""
    nombres_lower = [n.lower() for n in wb.sheetnames]
    for n in nombres_lower:
        if 'cierre' in n and 'isapre' not in n:
            return 'cierre'
    if 'afc-afp' in nombres_lower or 'base afp' in nombres_lower:
        return 'base'
    return 'desconocido'


def leer_excel_cierre(excel_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    tipo = detectar_tipo_excel(wb)

    if tipo == 'base':
        datos = leer_base_afp(wb)
        if datos:
            return {'Base de Deudas': datos}, wb.sheetnames
        return {}, wb.sheetnames

    hojas_cierre = detectar_hojas_cierre(wb)
    resultado = {}
    for h in hojas_cierre:
        try:
            resultado[h] = leer_cierre(wb, h)
        except Exception as e:
            resultado[h] = {'error': str(e)}
    return resultado, wb.sheetnames


def datos_a_chartjs(datos):
    """Convierte los datos del cierre a formato Chart.js."""
    charts = {}

    COLORES_INST = [
        '#6366f1','#8b5cf6','#a78bfa','#c4b5fd',
        '#4f46e5','#7c3aed','#2563eb','#0ea5e9',
        '#10b981','#f59e0b','#ef4444','#ec4899',
    ]
    COLORES_ESTATUS = {
        'Corresponde Pago':         '#10b981',
        'Corresponde pago':         '#10b981',
        'En gestion':               '#f59e0b',
        'En Gestion':               '#f59e0b',
        'En gestión':               '#f59e0b',
        'Regularizado':             '#6366f1',
        'Regularizado Jun-26':      '#6366f1',
        'Regularizado May-26':      '#6366f1',
        'Regularizado - Informado': '#a78bfa',
        'Regularizado-Informado':   '#a78bfa',
        'Pagado en aclaracion':     '#0ea5e9',
        'Sin deuda':                '#94a3b8',
    }

    # Gráfico 1: Por institución (barra horizontal)
    inst = {k: v for k, v in datos.get('por_institucion', {}).items() if k != '__total__'}
    if inst:
        charts['por_institucion'] = {
            'type': 'bar',
            'labels': list(inst.keys()),
            'datasets': [{'label': 'Monto ($)', 'data': list(inst.values()),
                          'backgroundColor': COLORES_INST[:len(inst)], 'borderRadius': 6}],
            'options': {'indexAxis': 'y', 'title': 'Deuda por institución'},
        }

    # Gráfico 2: Por estatus (doughnut)
    est = {k: v for k, v in datos.get('por_estatus', {}).items() if k != '__total__' and v > 0}
    if est:
        charts['por_estatus'] = {
            'type': 'doughnut',
            'labels': list(est.keys()),
            'datasets': [{'data': list(est.values()),
                          'backgroundColor': [COLORES_ESTATUS.get(k, '#94a3b8') for k in est],
                          'hoverOffset': 6}],
            'options': {'title': 'Estado de la deuda'},
        }

    # Gráfico 3: Por año (barra vertical)
    anios = {k: v for k, v in datos.get('por_anio', {}).items() if k != '__total__'}
    if anios:
        charts['por_anio'] = {
            'type': 'bar',
            'labels': [str(k) for k in anios.keys()],
            'datasets': [{'label': 'Monto ($)', 'data': list(anios.values()),
                          'backgroundColor': '#6366f1', 'borderRadius': 4}],
            'options': {'title': 'Deuda por año'},
        }

    # Gráfico 4: Por motivo (doughnut)
    mot = {k: v for k, v in datos.get('por_motivo', {}).items() if k != '__total__' and v > 0}
    if mot:
        charts['por_motivo'] = {
            'type': 'doughnut',
            'labels': list(mot.keys()),
            'datasets': [{'data': list(mot.values()),
                          'backgroundColor': COLORES_INST[:len(mot)], 'hoverOffset': 6}],
            'options': {'title': 'Motivos de deuda'},
        }

    return charts
