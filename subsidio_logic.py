"""Visor 'Subsidio a la Contratación — Línea de Activación Laboral'.

Lee la hoja 'Linea contrata' del Excel de Gestión de Ingresos (SharePoint) y la
entrega en forma de tabla + resumen para mostrarla de forma clara.

El enlace compartido del Excel se guarda en la variable de entorno SUBSIDIO_URL
o, si no existe, en la tabla app_config (clave 'subsidio_url') — NO va en el
código porque el repositorio es público. Acepta el share-link tal cual lo
entrega SharePoint y lo convierte al endpoint de descarga directa.
"""

import io
import os
import re
import time
import threading
import unicodedata
import http.cookiejar
import urllib.request
import datetime as _dt
from collections import Counter, OrderedDict

import openpyxl

HOJA        = "Linea contrata"
REFRESCO_SEG = 600  # 10 minutos de cache

_CACHE = {"data": None, "ts": 0, "error": None}
_LOCK  = threading.Lock()

# Columnas que muestra el visor (nombre en la hoja → etiqueta amigable)
COLUMNAS = [
    ("GRUPO",              "Grupo"),
    ("Responsable",        "Responsable"),
    ("RUT",                "RUT"),
    ("Estatus cliente",    "Estatus cliente"),
    ("EMPRESA",            "Empresa"),
    ("Estatus",            "Estatus"),
    ("Estatus de gestión", "Estatus de gestión"),
    ("Estatus de Carga",   "Estatus de carga"),
]


def url_guardada():
    url = os.environ.get("SUBSIDIO_URL", "").strip()
    if url:
        return url
    try:
        from database import get_conn
        with get_conn() as conn:
            row = conn.execute(
                "SELECT valor FROM app_config WHERE clave='subsidio_url'").fetchone()
            return (row["valor"] or "").strip() if row else ""
    except Exception:
        return ""


def guardar_url(url):
    from database import get_conn
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO app_config(clave, valor) VALUES('subsidio_url', ?) "
            "ON CONFLICT(clave) DO UPDATE SET valor=excluded.valor", (url.strip(),))
    with _LOCK:
        _CACHE.update({"data": None, "ts": 0, "error": None})


def _url_descarga(url):
    """Convierte el share-link de SharePoint al endpoint de descarga directa."""
    if not url:
        return None
    if "download.aspx" in url:
        return url
    try:
        if "/:x:/g/personal/" in url:
            dominio = url.split("/:x:/")[0]
            resto   = url.split("/:x:/g/personal/")[1]
            usuario, token = resto.split("/", 1)
            token = token.split("?")[0]
            return f"{dominio}/personal/{usuario}/_layouts/15/download.aspx?share={token}"
    except Exception:
        pass
    return url


def _descargar():
    url = _url_descarga(url_guardada())
    if not url:
        raise Exception("Falta configurar el enlace del Excel (pégalo en Configuración).")
    cj = http.cookiejar.CookieJar()
    op = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    op.addheaders = [("User-Agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64)")]
    with op.open(url, timeout=90) as r:
        data = r.read()
    if data[:2] != b"PK":
        raise Exception("SharePoint no devolvió el Excel — revisa que el enlace siga "
                        "vigente y compartido como 'Cualquier persona con el enlace'.")
    return data


def _fmt(v):
    if v is None:
        return ""
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _parsear(data):
    """Lee la hoja 'Linea contrata' → (filas como dicts con etiquetas amigables)."""
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = None
        for nombre in wb.sheetnames:
            if nombre.strip().lower() == HOJA.lower():
                ws = wb[nombre]; break
        if ws is None:  # tolerante: buscar por palabras clave
            for nombre in wb.sheetnames:
                if "linea" in nombre.lower() and "contrat" in nombre.lower():
                    ws = wb[nombre]; break
        if ws is None:
            raise Exception(f"No se encontró la hoja '{HOJA}' en el Excel.")

        gen = ws.iter_rows(values_only=True)
        header = []
        for row in gen:
            header = [str(h).strip() if h else "" for h in row]
            break

        # Mapear cada columna pedida a su índice en la hoja (match tolerante)
        idx_por_etiqueta = OrderedDict()
        for col_hoja, etiqueta in COLUMNAS:
            ci = None
            for i, h in enumerate(header):
                if h and h.strip().lower() == col_hoja.lower():
                    ci = i; break
            if ci is None:  # match parcial
                for i, h in enumerate(header):
                    if h and col_hoja.lower() in h.strip().lower():
                        ci = i; break
            idx_por_etiqueta[etiqueta] = ci

        filas = []
        for row in gen:
            if not any(v is not None and str(v).strip() for v in row):
                continue
            d = {}
            for etiqueta, ci in idx_por_etiqueta.items():
                d[etiqueta] = _fmt(row[ci]) if (ci is not None and ci < len(row)) else ""
            # descartar filas totalmente vacías en las columnas clave
            if d.get("RUT") or d.get("Empresa") or d.get("Grupo"):
                filas.append(d)
        return filas
    finally:
        wb.close()


def _resumen(filas):
    """Conteos para las tarjetas del visor."""
    por_gestion = Counter(f.get("Estatus de gestión") or "(vacío)" for f in filas)
    por_resp    = Counter((f.get("Responsable") or "(vacío)").strip().title() for f in filas)
    por_estatus = Counter(f.get("Estatus") or "(vacío)" for f in filas)
    grupos      = {f.get("Grupo") for f in filas if f.get("Grupo")}
    return {
        "total":         len(filas),
        "grupos":        len(grupos),
        "por_gestion":   por_gestion.most_common(),
        "por_responsable": por_resp.most_common(),
        "por_estatus":   por_estatus.most_common(),
    }


# ── Detalle de Gestiones (subida manual del Excel detallePostulacion) ──────────

DET_COLS = ["NumeroPostulacionEmpresa", "NumeroPostulacionDupla", "FechaPostulacion",
            "RutTrabajador", "DvTrabajador", "Nombres", "Apellidos", "Estado", "Motivo"]


def _fmt_rut(rut, dv):
    """Une RUT + DV con puntos: 16936974 + 7 → 16.936.974-7."""
    s = _fmt(rut)
    if not s:
        return ""
    dvs = _fmt(dv)
    s = s.replace(".", "").replace("-", "")
    try:
        s = f"{int(s):,}".replace(",", ".")
    except Exception:
        pass
    return f"{s}-{dvs}" if dvs else s


def parsear_detalle(data):
    """Lee la hoja 'detallePostulacion...' de un Excel subido y devuelve el detalle
    de trabajadores postulados + un resumen por número de postulación."""
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = None
        for nombre in wb.sheetnames:
            if "detallepostulacion" in nombre.lower().replace(" ", ""):
                ws = wb[nombre]; break
        if ws is None:  # buscar por la columna clave en la primera fila
            for w in wb.worksheets:
                first = next(w.iter_rows(values_only=True), None)
                if first and any(h and "numeropostulacionempresa" in str(h).lower().replace(" ", "")
                                 for h in first):
                    ws = w; break
        if ws is None:
            raise Exception("No se encontró la hoja de detalle de postulación "
                            "(debe tener la columna 'NumeroPostulacionEmpresa').")

        gen = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h else "" for h in next(gen)]

        def idx(name):
            for i, h in enumerate(header):
                if h.lower() == name.lower():
                    return i
            for i, h in enumerate(header):
                if name.lower() in h.lower():
                    return i
            return None

        I = {c: idx(c) for c in DET_COLS}

        def g(row, c):
            i = I.get(c)
            return row[i] if (i is not None and i < len(row)) else None

        trabajadores = []
        for row in gen:
            if not any(v is not None and str(v).strip() for v in row):
                continue
            trabajadores.append({
                "postulacion": _fmt(g(row, "NumeroPostulacionEmpresa")),
                "dupla":       _fmt(g(row, "NumeroPostulacionDupla")),
                "fecha":       _fmt(g(row, "FechaPostulacion")),
                "rut":         _fmt_rut(g(row, "RutTrabajador"), g(row, "DvTrabajador")),
                "nombre":      f"{_fmt(g(row, 'Nombres'))} {_fmt(g(row, 'Apellidos'))}".strip(),
                "estado":      _fmt(g(row, "Estado")),
                "motivo":      _fmt(g(row, "Motivo")),
            })

        posts = OrderedDict()
        for t in trabajadores:
            p = t["postulacion"] or "(sin nº)"
            if p not in posts:
                posts[p] = {"numero": p, "fecha": t["fecha"], "total": 0, "estados": Counter()}
            posts[p]["total"] += 1
            posts[p]["estados"][t["estado"] or "(sin estado)"] += 1

        postulaciones = [{
            "numero":  v["numero"],
            "fecha":   v["fecha"],
            "total":   v["total"],
            "estados": v["estados"].most_common(),
        } for v in posts.values()]

        return {
            "postulaciones": postulaciones,
            "trabajadores":  trabajadores,
            "total":         len(trabajadores),
        }
    finally:
        wb.close()


# ── Cruce con BASE MADRE ───────────────────────────────────────────────────────

def _rut_norm(r):
    return re.sub(r"[.\-\s]", "", str(r or "")).upper()


def _norm_txt(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return s.strip().lower()


# Solo se consideran empresas con estos Estatus cliente (el resto no aparece)
_ESTATUS_PERMITIDOS = ("vigente", "estudio inicial")


def _estatus_ok(val):
    n = _norm_txt(val)
    return any(n == p or n.startswith(p) for p in _ESTATUS_PERMITIDOS)


def _cruzar_base_madre(filas, forzar=False):
    """Agrega al listado las empresas que están en BASE MADRE pero NO en la base
    de subsidios (cruce por RUT). Se marcan con Estatus de gestión 'NO APLICA'.
    Motivo automático 'No tiene gestión de ingresos' si no tienen consultor de
    ingreso asignado; si lo tienen, el motivo queda vacío para completarlo a mano.
    Devuelve (filas, nº agregadas). Si base madre no está disponible, no falla."""
    try:
        from base_madre_logic import obtener_clientes
        cols, filas_bm, _ts, _err = obtener_clientes(forzar=forzar)
    except Exception:
        return filas, 0
    if not filas_bm:
        return filas, 0

    def col(*terms):
        for c in (cols or []):
            cn = _norm_txt(c)
            if all(t in cn for t in terms):
                return c
        return None

    c_rut     = col("rut")
    c_razon   = col("razon", "social") or col("empresa")
    c_grupo   = col("grupo")
    c_ingreso = col("consultor", "ingreso")
    c_estcli  = col("estatus", "cliente") or col("estatus")
    if not c_rut:
        return filas, 0

    presentes = {_rut_norm(f.get("RUT")) for f in filas}
    vistos, nuevos = set(), 0
    for f in filas_bm:
        rn = _rut_norm(f.get(c_rut))
        if not rn or rn in presentes or rn in vistos:
            continue
        # Solo empresas VIGENTE o ESTUDIO INICIAL de base madre
        if c_estcli and not _estatus_ok(f.get(c_estcli)):
            continue
        vistos.add(rn)
        consultor = (f.get(c_ingreso) or "").strip() if c_ingreso else ""
        filas.append({
            "Grupo":              (f.get(c_grupo) or "").strip() if c_grupo else "",
            "Responsable":        consultor,
            "RUT":                (f.get(c_rut) or "").strip(),
            "Estatus cliente":    (f.get(c_estcli) or "").strip() if c_estcli else "",
            "Empresa":            (f.get(c_razon) or "").strip() if c_razon else "",
            "Estatus":            "",
            "Estatus de gestión": "NO APLICA",
            "Estatus de carga":   "",
            "Motivo":             "No tiene gestión de ingresos" if not consultor else "",
            "_no_aplica":         True,   # marca fila agregada por el cruce (motivo editable)
        })
        nuevos += 1
    return filas, nuevos


def obtener(forzar=False):
    """Devuelve dict {columnas, filas, resumen, ts, error}. Cache de REFRESCO_SEG."""
    with _LOCK:
        fresco = _CACHE["data"] is not None and (time.time() - _CACHE["ts"]) < REFRESCO_SEG
        if fresco and not forzar:
            return {**_CACHE["data"], "ts": _CACHE["ts"], "error": _CACHE["error"]}
        try:
            data  = _descargar()
            filas = _parsear(data)
            # Solo empresas con Estatus cliente VIGENTE o ESTUDIO INICIAL
            filas = [f for f in filas if _estatus_ok(f.get("Estatus cliente"))]
            filas, cruzadas = _cruzar_base_madre(filas, forzar=forzar)  # agrega NO APLICA desde base madre
            # Unificar variantes de 'no aplica' (ej. 'No aplica' de la hoja vs
            # 'NO APLICA' del cruce) para que no salgan dos tarjetas separadas.
            for _f in filas:
                if _norm_txt(_f.get("Estatus de gestión")) == "no aplica":
                    _f["Estatus de gestión"] = "NO APLICA"
            payload = {
                "columnas": [et for _, et in COLUMNAS] + ["Motivo"],
                "filas":    filas,
                "resumen":  _resumen(filas),
                "cruzadas": cruzadas,
            }
            _CACHE.update({"data": payload, "ts": time.time(), "error": None})
        except Exception as e:
            _CACHE["error"] = str(e)
            if _CACHE["data"] is None:
                return {"columnas": [], "filas": [], "resumen": {}, "ts": 0, "error": str(e)}
        return {**_CACHE["data"], "ts": _CACHE["ts"], "error": _CACHE["error"]}
